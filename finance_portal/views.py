import calendar
import json
from datetime import date

from django.contrib import messages
from django.db.models import Q
from django.db.models.functions import Lower
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_exempt

from companies.models import Company
from core.email_utils import send_payslip_email
from core.utils import save_pdf_to_model
from employees.models import Employee, Payslip
from employees.payroll_utils import calculate_payslip_breakdown, num2words_flexible

from .decorators import finance_manager_required
from .models import FinanceAuditLog, PayrollBatch


def get_client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    ip = x_forwarded_for.split(",")[0] if x_forwarded_for else request.META.get("REMOTE_ADDR")
    return ip


@finance_manager_required
def dashboard(request):
    """Centralized Finance Portal Dashboard for cross-company payroll"""
    companies = Company.objects.filter(is_active=True).order_by("name")

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))
    selected_company_id = request.GET.get("company", "all")

    # Month range calculation
    num_days = calendar.monthrange(selected_year, selected_month)[1]
    month_start = date(selected_year, selected_month, 1)
    month_end = date(selected_year, selected_month, num_days)

    # Base query for employees who were employed during the selected month
    employees = Employee.objects.filter(Q(date_of_joining__isnull=True) | Q(date_of_joining__lte=month_end)).filter(
        Q(is_active=True) | Q(employment_status="ACTIVE") | (Q(exit_date__isnull=False) & Q(exit_date__gte=month_start))
    )

    if selected_company_id and selected_company_id != "all":
        employees = employees.filter(company_id=selected_company_id)

    employees = employees.select_related("user", "company", "location").order_by(
        Lower("company__name"), Lower("user__first_name"), Lower("user__last_name")
    )

    # Get payslips for selected month/year
    existing_payslips = Payslip.objects.filter(month__month=selected_month, month__year=selected_year)
    if selected_company_id and selected_company_id != "all":
        existing_payslips = existing_payslips.filter(employee__company_id=selected_company_id)

    payslip_map = {slip.employee_id: slip for slip in existing_payslips}

    # Package employee details with payslips, separating active from ex-employees
    active_employee_data = []
    ex_employee_data = []
    active_calculated_count = 0
    ex_calculated_count = 0

    for emp in employees:
        emp.monthly_ctc = float(emp.annual_ctc or 0.0) / 12
        slip = payslip_map.get(emp.id)
        is_active = emp.is_active and emp.employment_status == "ACTIVE" and emp.exit_date is None
        item = {"employee": emp, "payslip": slip}

        if is_active:
            if slip:
                active_calculated_count += 1
            active_employee_data.append(item)
        else:
            if slip:
                ex_calculated_count += 1
            ex_employee_data.append(item)

    months = [{"value": i, "name": calendar.month_name[i]} for i in range(1, 13)]
    years = range(today.year - 2, today.year + 2)

    # Add days in month to context for default worked days
    days_in_month = num_days

    # Payroll cycle: previous month's day count is used for calculation denominator
    if selected_month == 1:
        prev_month_days = calendar.monthrange(selected_year - 1, 12)[1]
    else:
        prev_month_days = calendar.monthrange(selected_year, selected_month - 1)[1]

    # Fetch recent payroll batches
    batches = PayrollBatch.objects.all().prefetch_related("companies")[:10]
    audit_logs = FinanceAuditLog.objects.all().select_related("user", "company")[:15]

    context = {
        "title": "Centralized Finance Portal",
        "companies": companies,
        "active_employee_data": active_employee_data,
        "ex_employee_data": ex_employee_data,
        "active_calculated_count": active_calculated_count,
        "ex_calculated_count": ex_calculated_count,
        "selected_month": selected_month,
        "selected_year": selected_year,
        "selected_company_id": selected_company_id,
        "months": months,
        "years": years,
        "batches": batches,
        "audit_logs": audit_logs,
        "days_in_month": days_in_month,
        "prev_month_days": prev_month_days,
    }
    return render(request, "finance_portal/dashboard.html", context)


def generate_payslip_internal(
    employee, month, year, worked_days=None, monthly_gross=None, is_draft=False, travel_allowance=0.0, tds_deduction=0.0
):
    """
    Core payroll generation helper (mirroring core/views.py process_payslip_generation).
    Calculates breakdown, saves Payslip, and optionally renders PDF (skipped if is_draft=True).
    """
    total_days = calendar.monthrange(year, month)[1]
    month_date = date(year, month, 1)
    travel_allowance = float(travel_allowance or 0.0)
    tds_deduction = float(tds_deduction or 0.0)

    if worked_days is None:
        worked_days = float(total_days)

    # If custom monthly gross is supplied (e.g. from Excel upload), use it to calculate annual CTC
    annual_ctc = float(monthly_gross) * 12 if monthly_gross is not None else float(employee.annual_ctc or 0.0)

    # Calculate breakdown
    breakdown = calculate_payslip_breakdown(
        annual_ctc,
        worked_days,
        total_days,
        employee.pf_enabled,
        location=employee.location,
        company=employee.company,
        month=month,
        year=year,
        travel_allowance=travel_allowance,
        tds_deduction=tds_deduction,
    )

    payslip, created = Payslip.objects.get_or_create(employee=employee, month=month_date)

    # Set payslip fields
    payslip.basic = breakdown["basic"]
    payslip.hra = breakdown["hra"]
    payslip.lta = breakdown["lta"]
    payslip.other_allowance = breakdown["other_allowance"]
    payslip.travel_allowance = breakdown["travel_allowance"]
    payslip.tds_deduction = breakdown["tds_deduction"]

    # Map location specific allowances
    if breakdown.get("country_code", "IN") == "IN":
        payslip.conveyance_allowance = breakdown.get("lta", 0.0)
        payslip.special_allowance = breakdown.get("other_allowance", 0.0)
    else:
        payslip.conveyance_allowance = breakdown.get("conveyance", 0.0)
        payslip.special_allowance = breakdown.get("medical", 0.0)

    payslip.monthly_gross = breakdown["full_monthly_gross"]
    payslip.gross_salary = breakdown["gross_monthly"]
    payslip.employee_pf = breakdown["employee_pf"]
    payslip.employer_pf = breakdown["employer_pf"]
    payslip.professional_tax = breakdown["professional_tax"]
    payslip.net_salary = breakdown["net_salary"]
    payslip.worked_days = worked_days
    payslip.total_days = breakdown.get("display_days", total_days)
    payslip.is_draft = is_draft
    payslip.save()

    if not is_draft:
        # Generate branding and currency details for PDF context
        currency = "INR"
        currency_name = "Rupees"
        if employee.location:
            currency = employee.location.currency or "INR"
            if employee.location.country_code == "IN" or currency == "INR":
                currency_name = "Rupees"
            elif employee.location.country_code == "BD" or currency == "BDT":
                currency_name = "Taka"
            elif employee.location.country_code == "US" or currency == "USD":
                currency_name = "Dollars"
            else:
                currency_name = currency

        cname_upper = employee.company.name.upper()
        branding = {
            "name": "PETABYTZ TECHNOLOGY SERVICES PVT LTD",
            "address": "PLOT NO 201 & 202, 1ST FLOOR, DMR CORPORATE, KAVURI HILLS RD, HYDERABAD, TELANGANA 500081.",
        }
        if "SOFTSTANDARD" in cname_upper or "RMINDS" in cname_upper:
            branding["name"] = "SOFTSTANDARD SOLUTIONS"

        if employee.location and employee.location.address_line1:
            loc = employee.location
            addr = f"{loc.address_line1}"
            if loc.address_line2:
                addr += f", {loc.address_line2}"
            addr += f", {loc.city}"
            if loc.state:
                addr += f", {loc.state}"
            if loc.postal_code:
                addr += f" {loc.postal_code}"
            branding["address"] = addr

        context = {
            "payslip": payslip,
            "company": employee.company,
            "branding": branding,
            "net_salary_words": num2words_flexible(round(payslip.net_salary or 0), currency_name),
            "currency": currency,
            "basic_rounded": round(payslip.basic or 0),
            "hra_rounded": round(payslip.hra or 0),
            "conveyance_rounded": round(payslip.conveyance_allowance or 0),
            "special_rounded": round(payslip.special_allowance or 0),
            "employer_pf_rounded": round(payslip.employer_pf or 0),
            "employee_pf_rounded": round(payslip.employee_pf or 0),
            "professional_tax_rounded": round(payslip.professional_tax or 0),
            "total_earnings_ctc": round((payslip.gross_salary or 0) + (payslip.employer_pf or 0)),
            "total_contributions": round((payslip.employee_pf or 0) + (payslip.employer_pf or 0)),
            "net_salary_rounded": round(payslip.net_salary or 0),
            "payable_units": f"{int(payslip.worked_days)} Days",
        }

        filename = f"payslip_{employee.badge_id}_{month_date.strftime('%b_%Y')}.pdf"
        save_pdf_to_model(payslip, "employees/payslip_pdf.html", context, filename)
    else:
        # Clear existing PDF files when generating draft
        payslip.pdf_file = None
        payslip.save()

    return payslip


def finalize_payslip_internal(payslip):
    """
    Core payroll finalization helper. Renders the PDF using values currently stored in the
    Payslip record, changes is_draft to False, and saves it.
    """
    employee = payslip.employee
    month_date = payslip.month

    # Generate branding and currency details for PDF context
    currency = "INR"
    currency_name = "Rupees"
    if employee.location:
        currency = employee.location.currency or "INR"
        if employee.location.country_code == "IN" or currency == "INR":
            currency_name = "Rupees"
        elif employee.location.country_code == "BD" or currency == "BDT":
            currency_name = "Taka"
        elif employee.location.country_code == "US" or currency == "USD":
            currency_name = "Dollars"
        else:
            currency_name = currency

    cname_upper = employee.company.name.upper()
    branding = {
        "name": "PETABYTZ TECHNOLOGY SERVICES PVT LTD",
        "address": "PLOT NO 201 & 202, 1ST FLOOR, DMR CORPORATE, KAVURI HILLS RD, HYDERABAD, TELANGANA 500081.",
    }
    if "SOFTSTANDARD" in cname_upper or "RMINDS" in cname_upper:
        branding["name"] = "SOFTSTANDARD SOLUTIONS"

    if employee.location and employee.location.address_line1:
        loc = employee.location
        addr = f"{loc.address_line1}"
        if loc.address_line2:
            addr += f", {loc.address_line2}"
        addr += f", {loc.city}"
        if loc.state:
            addr += f", {loc.state}"
        if loc.postal_code:
            addr += f" {loc.postal_code}"
        branding["address"] = addr

    context = {
        "payslip": payslip,
        "company": employee.company,
        "branding": branding,
        "net_salary_words": num2words_flexible(round(payslip.net_salary or 0), currency_name),
        "currency": currency,
        "basic_rounded": round(payslip.basic or 0),
        "hra_rounded": round(payslip.hra or 0),
        "conveyance_rounded": round(payslip.conveyance_allowance or 0),
        "special_rounded": round(payslip.special_allowance or 0),
        "employer_pf_rounded": round(payslip.employer_pf or 0),
        "employee_pf_rounded": round(payslip.employee_pf or 0),
        "professional_tax_rounded": round(payslip.professional_tax or 0),
        "tds_deduction_rounded": round(payslip.tds_deduction or 0),
        "total_earnings_ctc": round((payslip.gross_salary or 0) + (payslip.employer_pf or 0)),
        "total_contributions": round((payslip.employee_pf or 0) + (payslip.employer_pf or 0)),
        "total_taxes_deductions_rounded": round((payslip.professional_tax or 0) + (payslip.tds_deduction or 0)),
        "net_salary_rounded": round(payslip.net_salary or 0),
        "payable_units": f"{int(payslip.worked_days)} Days",
    }

    filename = f"payslip_{employee.badge_id}_{month_date.strftime('%b_%Y')}.pdf"
    save_pdf_to_model(payslip, "employees/payslip_pdf.html", context, filename)

    payslip.is_draft = False
    payslip.save()
    return payslip


@finance_manager_required
def process_draft_payroll(request):
    """Process Phase 1 bulk draft payroll for a company or all companies with backend attendance count"""
    if request.method != "POST":
        return redirect("finance_portal:dashboard")

    company_id = request.POST.get("company_id", "all")
    month = int(request.POST.get("month"))
    year = int(request.POST.get("year"))

    # Calculate range
    num_days = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, num_days)

    # Filter employees
    employees = Employee.objects.filter(Q(date_of_joining__isnull=True) | Q(date_of_joining__lte=month_end)).filter(
        Q(is_active=True) | Q(employment_status="ACTIVE") | (Q(exit_date__isnull=False) & Q(exit_date__gte=month_start))
    )

    # Apply company filter
    target_companies = Company.objects.filter(is_active=True)
    if company_id and company_id != "all":
        employees = employees.filter(company_id=company_id)
        target_companies = target_companies.filter(id=company_id)

    if not employees.exists():
        messages.error(request, "No eligible employees found for the selected company/companies.")
        return redirect(f"/finance/?company={company_id}&month={month}&year={year}")

    from employees.models import Attendance

    processed_count = 0
    errors = []

    for emp in employees:
        try:
            # Count days with clock_in or status present in selected month/year
            worked_days = (
                Attendance.objects.filter(employee=emp, date__year=year, date__month=month)
                .filter(Q(clock_in__isnull=False) | Q(status__in=["PRESENT", "WFH", "ON_DUTY", "HYBRID"]))
                .count()
            )

            # Default to total days if employee has no attendance records at all
            if worked_days == 0:
                worked_days = num_days

            # Generate Phase 1 draft payslip
            generate_payslip_internal(emp, month, year, worked_days=float(worked_days), is_draft=True)
            processed_count += 1
        except Exception as e:
            errors.append(f"Error for {emp.user.get_full_name()}: {str(e)}")

    # Audit log
    comp_name_str = "All Companies" if company_id == "all" else target_companies.first().name
    details = f"Processed Phase 1 Draft payroll for {comp_name_str}. Period: {month}/{year}. Total: {employees.count()}, Success: {processed_count}."
    if errors:
        details += f" Errors: {len(errors)}"

    FinanceAuditLog.objects.create(
        user=request.user,
        action="BULK_PAYROLL_DRAFT",
        company=None if company_id == "all" else target_companies.first(),
        details=details,
        ip_address=get_client_ip(request),
    )

    if errors:
        messages.warning(
            request, f"Draft payroll generated with some errors ({len(errors)} failed out of {employees.count()})."
        )
    else:
        messages.success(request, f"Successfully generated Phase 1 Draft payroll for {processed_count} employees!")

    return redirect(f"/finance/?company={company_id}&month={month}&year={year}")


@finance_manager_required
@csrf_exempt
def save_draft_payslip(request):
    """
    AJAX endpoint to update draft payslip fields dynamically.
    Recalculates net and gross based on edited values.
    Also saves annual_ctc back to the Employee if provided.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Only POST requests allowed"}, status=405)

    try:
        data = json.loads(request.body)
        payslip_id = data.get("payslip_id")
        payslip = get_object_or_404(Payslip, id=payslip_id)

        # Update customizable fields
        payslip.worked_days = float(data.get("worked_days", payslip.worked_days))
        payslip.basic = float(data.get("basic", payslip.basic))
        payslip.hra = float(data.get("hra", payslip.hra))
        payslip.conveyance_allowance = float(data.get("conveyance_allowance", payslip.conveyance_allowance))
        payslip.special_allowance = float(data.get("special_allowance", payslip.special_allowance))
        payslip.employee_pf = float(data.get("employee_pf", payslip.employee_pf))
        payslip.employer_pf = float(data.get("employer_pf", payslip.employer_pf))
        payslip.professional_tax = float(data.get("professional_tax", payslip.professional_tax))

        if "travel_allowance" in data:
            payslip.travel_allowance = float(data["travel_allowance"])
        if "tds_deduction" in data:
            payslip.tds_deduction = float(data["tds_deduction"])

        # Dynamic salary calculations based on manual overrides
        if "gross_salary" in data:
            payslip.gross_salary = float(data["gross_salary"])
            payslip.monthly_gross = payslip.gross_salary
        else:
            payslip.gross_salary = (
                payslip.basic
                + payslip.hra
                + payslip.conveyance_allowance
                + payslip.special_allowance
                + float(payslip.travel_allowance or 0.0)
            )
            payslip.monthly_gross = payslip.gross_salary

        if "net_salary" in data:
            payslip.net_salary = float(data["net_salary"])
        else:
            payslip.net_salary = (
                payslip.gross_salary
                - payslip.employee_pf
                - payslip.employer_pf
                - payslip.professional_tax
                - float(payslip.tds_deduction or 0.0)
            )
        payslip.save()

        # Also save annual_ctc to Employee if provided
        new_ctc = data.get("annual_ctc")
        if new_ctc is not None:
            employee = payslip.employee
            employee.annual_ctc = float(new_ctc)
            employee.save(update_fields=["annual_ctc"])

        # Log change to security audit log
        FinanceAuditLog.objects.create(
            user=request.user,
            company=payslip.employee.company,
            action="EDIT_DRAFT_PAYSLIP",
            details=f"Manually edited draft payslip for {payslip.employee.user.get_full_name()} ({payslip.month.strftime('%b %Y')}). Net: {payslip.net_salary}.",
            ip_address=get_client_ip(request),
        )

        return JsonResponse(
            {
                "status": "success",
                "net_salary": payslip.net_salary,
                "gross_salary": payslip.gross_salary,
                "message": "Draft saved successfully!",
            }
        )

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)


@finance_manager_required
@csrf_exempt
def update_employee_ctc(request):
    """
    AJAX endpoint: given an employee_id and new annual_ctc,
    recalculates the full payslip breakdown and returns all component values
    so the frontend can auto-populate the Edit Draft form fields.
    Also persists annual_ctc on the Employee record.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
        payslip_id = data.get("payslip_id")
        new_ctc = float(data.get("annual_ctc", 0))
        payslip = get_object_or_404(Payslip, id=payslip_id)
        employee = payslip.employee

        if new_ctc <= 0:
            return JsonResponse({"status": "error", "message": "CTC must be a positive number"}, status=400)

        # Save CTC on employee
        employee.annual_ctc = new_ctc
        employee.save(update_fields=["annual_ctc"])

        # Recalculate breakdown using existing worked_days
        worked_days = float(payslip.worked_days or 30)
        month_obj = payslip.month  # date(year, month, 1)
        total_days = calendar.monthrange(month_obj.year, month_obj.month)[1]

        breakdown = calculate_payslip_breakdown(
            new_ctc,
            worked_days,
            total_days,
            employee.pf_enabled,
            location=employee.location,
            company=employee.company,
            month=month_obj.month,
            year=month_obj.year,
        )

        # Determine which allowance fields map to conveyance/special
        if breakdown.get("country_code", "IN") == "IN":
            conveyance = breakdown.get("lta", 0.0)
            special = breakdown.get("other_allowance", 0.0)
        else:
            conveyance = breakdown.get("conveyance", 0.0)
            special = breakdown.get("medical", 0.0)

        basic = breakdown["basic"]
        hra = breakdown["hra"]
        employee_pf = breakdown["employee_pf"]
        employer_pf = breakdown["employer_pf"]
        professional_tax = breakdown["professional_tax"]
        gross = breakdown["gross_monthly"]
        net = breakdown["net_salary"]
        monthly_gross = breakdown["full_monthly_gross"]

        # Update payslip with recalculated values
        payslip.basic = basic
        payslip.hra = hra
        payslip.conveyance_allowance = conveyance
        payslip.special_allowance = special
        payslip.employee_pf = employee_pf
        payslip.employer_pf = employer_pf
        payslip.professional_tax = professional_tax
        payslip.gross_salary = gross
        payslip.monthly_gross = monthly_gross
        payslip.net_salary = net
        payslip.save()

        # Audit log
        FinanceAuditLog.objects.create(
            user=request.user,
            company=employee.company,
            action="UPDATE_CTC",
            details=f"CTC updated for {employee.user.get_full_name()} to {new_ctc} ({month_obj.strftime('%b %Y')}). Recalculated Net: {net}.",
            ip_address=get_client_ip(request),
        )

        return JsonResponse(
            {
                "status": "success",
                "annual_ctc": new_ctc,
                "monthly_gross": round(new_ctc / 12, 2),
                "basic": round(basic, 2),
                "hra": round(hra, 2),
                "conveyance": round(conveyance, 2),
                "special": round(special, 2),
                "employee_pf": round(employee_pf, 2),
                "employer_pf": round(employer_pf, 2),
                "professional_tax": round(professional_tax, 2),
                "gross": round(gross, 2),
                "net": round(net, 2),
            }
        )

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)


@finance_manager_required
def process_bulk_payroll(request):
    """Process bulk payroll or finalize existing Phase 1 drafts for a company or all companies"""
    if request.method != "POST":
        return redirect("finance_portal:dashboard")

    company_id = request.POST.get("company_id", "all")
    month = int(request.POST.get("month"))
    year = int(request.POST.get("year"))
    auto_send_email = request.POST.get("auto_send_email") == "on"

    # Calculate range
    num_days = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, num_days)

    # Filter employees
    employees = Employee.objects.filter(Q(date_of_joining__isnull=True) | Q(date_of_joining__lte=month_end)).filter(
        Q(is_active=True) | Q(employment_status="ACTIVE") | (Q(exit_date__isnull=False) & Q(exit_date__gte=month_start))
    )

    # Apply company filter
    target_companies = Company.objects.filter(is_active=True)
    if company_id and company_id != "all":
        employees = employees.filter(company_id=company_id)
        target_companies = target_companies.filter(id=company_id)

    if not employees.exists():
        messages.error(request, "No eligible employees found for the selected company/companies.")
        return redirect(f"/finance/?company={company_id}&month={month}&year={year}")

    # Check for existing draft payslips for this employee subset and period
    draft_payslips = Payslip.objects.filter(employee__in=employees, month__month=month, month__year=year, is_draft=True)

    if draft_payslips.exists():
        # Phase 2: Finalize drafts
        processed_count = 0
        total_amount = 0.0
        errors = []

        batch = PayrollBatch.objects.create(
            month=month,
            year=year,
            status="PROCESSING",
            total_employees=draft_payslips.count(),
            processed_employees=0,
            total_amount=0.0,
            created_by=request.user,
        )
        batch.companies.add(*target_companies)

        for slip in draft_payslips:
            try:
                # Finalize draft (saves is_draft=False and renders PDF)
                finalize_payslip_internal(slip)

                # Send email if auto-send requested
                if auto_send_email:
                    send_payslip_email(slip)

                processed_count += 1
                if slip.net_salary:
                    total_amount += float(slip.net_salary)

                batch.processed_employees = processed_count
                batch.total_amount = total_amount
                batch.save()
            except Exception as e:
                errors.append(f"Error finalizing draft for {slip.employee.user.get_full_name()}: {str(e)}")

        if errors and processed_count == 0:
            batch.status = "FAILED"
        else:
            batch.status = "COMPLETED"
        batch.save()

        # Audit log
        company_log = None if company_id == "all" else target_companies.first()
        comp_name_str = "All Companies" if company_id == "all" else target_companies.first().name
        details = f"Finalized bulk payroll from drafts for {comp_name_str}. Period: {month}/{year}. Processed: {processed_count}/{batch.total_employees}."
        if errors:
            details += f" Errors: {len(errors)}"

        FinanceAuditLog.objects.create(
            user=request.user,
            action="BULK_PAYROLL_FINALIZE",
            company=company_log,
            details=details,
            ip_address=get_client_ip(request),
        )

        if errors:
            messages.warning(request, f"Payroll finalized with some errors ({len(errors)} failed).")
        else:
            messages.success(request, f"Successfully finalized and sent payroll for {processed_count} employees!")

        return redirect(f"/finance/?company={company_id}&month={month}&year={year}")

    # Fallback to direct bulk process if no drafts exist
    # Create the Batch record
    batch = PayrollBatch.objects.create(
        month=month,
        year=year,
        status="PROCESSING",
        total_employees=employees.count(),
        processed_employees=0,
        total_amount=0.0,
        created_by=request.user,
    )
    batch.companies.add(*target_companies)

    processed_count = 0
    total_amount = 0.0
    errors = []

    for emp in employees:
        try:
            # Generate payslip
            payslip = generate_payslip_internal(emp, month, year)

            # Send email if requested
            if auto_send_email:
                send_payslip_email(payslip)

            processed_count += 1
            if payslip.net_salary:
                total_amount += float(payslip.net_salary)

            # Update batch progress
            batch.processed_employees = processed_count
            batch.total_amount = total_amount
            batch.save()

        except Exception as e:
            errors.append(f"Error for {emp.user.get_full_name()}: {str(e)}")

    # Finalize batch
    if errors and processed_count == 0:
        batch.status = "FAILED"
    else:
        batch.status = "COMPLETED"
    batch.save()

    # Audit log
    company_log = None if company_id == "all" else target_companies.first()
    comp_name_str = "All Companies" if company_id == "all" else target_companies.first().name
    details = f"Processed bulk payroll for {comp_name_str}. Period: {month}/{year}. Total Employees: {batch.total_employees}, Processed: {processed_count}, Net Payroll: {total_amount:.2f}."
    if errors:
        details += f" Errors encountered: {len(errors)}"

    FinanceAuditLog.objects.create(
        user=request.user,
        action="BULK_PAYROLL_PROCESS",
        company=company_log,
        details=details,
        ip_address=get_client_ip(request),
    )

    if errors:
        messages.warning(
            request, f"Payroll batch processed with some errors ({len(errors)} failed out of {batch.total_employees})."
        )
    else:
        messages.success(request, f"Successfully processed payroll batch for {processed_count} employees!")

    return redirect(f"/finance/?company={company_id}&month={month}&year={year}")


@finance_manager_required
def send_single_payslip_email_view(request, payslip_id):
    """View to email a single generated payslip manually from the dashboard"""
    payslip = get_object_or_404(Payslip, id=payslip_id)
    if not payslip.pdf_file:
        # Generate the PDF if it's missing for some reason
        try:
            generate_payslip_internal(payslip.employee, payslip.month.month, payslip.month.year)
            payslip.refresh_from_db()
        except Exception as e:
            return JsonResponse({"status": "error", "message": f"Could not generate PDF: {str(e)}"}, status=500)

    success = send_payslip_email(payslip)

    # Audit log
    FinanceAuditLog.objects.create(
        user=request.user,
        action="SEND_SINGLE_PAYSLIP",
        company=payslip.employee.company,
        details=f"Emailed payslip to {payslip.employee.user.get_full_name()} for {payslip.month.strftime('%B %Y')}.",
        ip_address=get_client_ip(request),
    )

    if success:
        if request.headers.get("x-requested-with") == "XMLHttpRequest" or request.GET.get("ajax"):
            return JsonResponse({"status": "success", "message": "Email sent successfully!"})
        messages.success(request, f"Payslip successfully emailed to {payslip.employee.user.get_full_name()}!")
    else:
        if request.headers.get("x-requested-with") == "XMLHttpRequest" or request.GET.get("ajax"):
            return JsonResponse({"status": "error", "message": "Failed to send email. Check logs."}, status=500)
        messages.error(
            request, f"Failed to send email to {payslip.employee.user.get_full_name()}. Please check configuration."
        )

    return redirect(
        f"/finance/?company={payslip.employee.company.id}&month={payslip.month.month}&year={payslip.month.year}"
    )


@finance_manager_required
def process_bulk_excel_upload(request):
    """
    Process cross-company or single-company payroll via Excel upload.
    Matches active employees by badge number, calculates breakdown,
    saves PDF payslips, and optionally dispatches payslips to user emails immediately.
    """
    if request.method != "POST":
        return redirect("finance_portal:dashboard")

    import openpyxl

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        messages.error(request, "Please upload a valid Excel file.")
        return redirect("finance_portal:dashboard")

    company_id = request.POST.get("company_id", "all")
    month = int(request.POST.get("month", date.today().month))
    year = int(request.POST.get("year", date.today().year))
    auto_send_email = request.POST.get("auto_send_email") == "on"

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Extract headers mapping from Row 2 or Row 1
        header_map = {}
        header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        if not header_row or not any(header_row):
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            start_row = 2
        else:
            start_row = 3

        for idx, value in enumerate(header_row):
            if value:
                header_map[str(value).strip().lower()] = idx

        # Support flexible / alternative names for matching columns
        badge_key = None
        gross_key = None
        days_key = None

        for key in header_map:
            if any(term in key for term in ["employee", "badge", "id", "number", "emp"]):
                badge_key = key
            if any(term in key for term in ["gross", "amount", "salary", "ctc", "earned"]):
                gross_key = key
            if any(term in key for term in ["day", "hour", "unit", "payable", "worked"]):
                days_key = key

        # Fill standard defaults if not matched
        if not badge_key:
            badge_key = "employee number"
        if not gross_key:
            gross_key = "monthly earned gross"
        if not days_key:
            days_key = "no of payable units (days / hours / units)"

        # Validate critical headers exist
        missing = []
        for col in [badge_key, gross_key]:
            if col not in header_map:
                missing.append(col)
        if missing:
            messages.error(
                request,
                "Could not find required columns in your file. Ensure headers for employee number/badge ID and gross salary are present.",
            )
            return redirect("finance_portal:dashboard")

        success_count = 0
        error_count = 0

        # Initialize PayrollBatch
        batch = PayrollBatch.objects.create(
            month=month,
            year=year,
            total_employees=0,
            processed_employees=0,
            total_amount=0.0,
            status="PROCESSING",
            created_by=request.user,
        )
        selected_company = None
        if company_id != "all":
            selected_company = Company.objects.get(id=company_id)
            batch.companies.add(selected_company)

        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if len(row) <= max(header_map[badge_key], header_map[gross_key]):
                continue

            badge_val = row[header_map[badge_key]]
            if badge_val is None:
                continue

            badge_id = str(badge_val).strip()
            gross_val = row[header_map[gross_key]]
            days_val = row[header_map[days_key]] if days_key in header_map else None

            if gross_val is None:
                continue

            try:
                # Query active employee matching the Badge ID
                emp_query = Employee.objects.filter(badge_id=badge_id, is_active=True)
                if company_id != "all":
                    emp_query = emp_query.filter(company_id=company_id)
                employee = emp_query.first()

                if not employee:
                    error_count += 1
                    continue

                monthly_gross = float(gross_val)
                worked_days = float(days_val) if days_val is not None else None

                # Perform the full payslip generation cycle
                payslip = generate_payslip_internal(
                    employee=employee, month=month, year=year, worked_days=worked_days, monthly_gross=monthly_gross
                )

                # Email automatically with dynamic name and branding
                if auto_send_email:
                    send_payslip_email(payslip)

                success_count += 1
                batch.total_amount += float(payslip.net_salary or 0.0)

            except Exception:
                import traceback

                traceback.print_exc()
                error_count += 1

        # Conclude batch state
        batch.total_employees = success_count + error_count
        batch.processed_employees = success_count
        batch.status = "COMPLETED" if success_count > 0 else "FAILED"
        batch.save()

        # Log to FinanceAuditLog
        FinanceAuditLog.objects.create(
            user=request.user,
            company=selected_company,
            action="EXCEL_BULK_PAYROLL",
            ip_address=get_client_ip(request),
            details=f"Processed Excel bulk upload. Calculated and processed: {success_count}, Errors/Mismatches: {error_count}. Batch ID: {batch.batch_id}.",
        )

        messages.success(
            request,
            f"Successfully processed {success_count} payslips from Excel. {error_count} records failed or had no matching active employee.",
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        messages.error(request, f"Failed to read or parse Excel file: {str(e)}")

    return redirect(f"/finance/?company={company_id}&month={month}&year={year}")


@finance_manager_required
def preview_draft_payslip(request, payslip_id):
    """
    Dynamically renders and streams an on-the-fly PDF preview of a draft or final payslip.
    """
    from django.http import Http404, HttpResponse

    from core.utils import render_to_pdf

    try:
        payslip = Payslip.objects.get(pk=payslip_id)
    except Payslip.DoesNotExist:
        raise Http404("Payslip not found.")

    employee = payslip.employee

    # Generate branding and currency details for PDF context
    currency = "INR"
    currency_name = "Rupees"
    if employee.location:
        currency = employee.location.currency or "INR"
        if employee.location.country_code == "IN" or currency == "INR":
            currency_name = "Rupees"
        elif employee.location.country_code == "BD" or currency == "BDT":
            currency_name = "Taka"
        elif employee.location.country_code == "US" or currency == "USD":
            currency_name = "Dollars"
        else:
            currency_name = currency

    cname_upper = employee.company.name.upper()
    branding = {
        "name": "PETABYTZ TECHNOLOGY SERVICES PVT LTD",
        "address": "PLOT NO 201 & 202, 1ST FLOOR, DMR CORPORATE, KAVURI HILLS RD, HYDERABAD, TELANGANA 500081.",
    }
    if "SOFTSTANDARD" in cname_upper or "RMINDS" in cname_upper:
        branding["name"] = "SOFTSTANDARD SOLUTIONS"

    if employee.location and employee.location.address_line1:
        loc = employee.location
        addr = f"{loc.address_line1}"
        if loc.address_line2:
            addr += f", {loc.address_line2}"
        addr += f", {loc.city}"
        if loc.state:
            addr += f", {loc.state}"
        if loc.postal_code:
            addr += f" {loc.postal_code}"
        branding["address"] = addr

    context = {
        "payslip": payslip,
        "company": employee.company,
        "branding": branding,
        "net_salary_words": num2words_flexible(round(payslip.net_salary or 0), currency_name),
        "currency": currency,
        "basic_rounded": round(payslip.basic or 0),
        "hra_rounded": round(payslip.hra or 0),
        "conveyance_rounded": round(payslip.conveyance_allowance or 0),
        "special_rounded": round(payslip.special_allowance or 0),
        "employer_pf_rounded": round(payslip.employer_pf or 0),
        "employee_pf_rounded": round(payslip.employee_pf or 0),
        "professional_tax_rounded": round(payslip.professional_tax or 0),
        "tds_deduction_rounded": round(payslip.tds_deduction or 0),
        "total_earnings_ctc": round((payslip.gross_salary or 0) + (payslip.employer_pf or 0)),
        "total_contributions": round((payslip.employee_pf or 0) + (payslip.employer_pf or 0)),
        "total_taxes_deductions_rounded": round((payslip.professional_tax or 0) + (payslip.tds_deduction or 0)),
        "net_salary_rounded": round(payslip.net_salary or 0),
        "payable_units": f"{int(payslip.worked_days)} Days",
    }

    pdf_content = render_to_pdf("employees/payslip_pdf.html", context)
    if pdf_content:
        response = HttpResponse(pdf_content, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="draft_payslip_{payslip.id}.pdf"'
        return response
    else:
        return HttpResponse("Error generating draft preview PDF.", status=500)


@finance_manager_required
@csrf_exempt
def recalculate_components(request):
    """
    AJAX endpoint called when Annual CTC or Worked Days change.
    Recalculates the full payslip breakdown using the new/current CTC and worked days,
    updates the draft payslip record in the database, and returns all recalculated components.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
        payslip_id = data.get("payslip_id")
        payslip = get_object_or_404(Payslip, id=payslip_id)
        employee = payslip.employee

        # If annual_ctc is passed, update it on the employee
        new_ctc_val = data.get("annual_ctc")
        if new_ctc_val is not None:
            new_ctc = float(new_ctc_val)
            employee.annual_ctc = new_ctc
            employee.save(update_fields=["annual_ctc"])
        else:
            new_ctc = float(employee.annual_ctc or 0.0)

        # Update worked_days
        new_worked_days_val = data.get("worked_days")
        if new_worked_days_val is not None:
            worked_days = float(new_worked_days_val)
            payslip.worked_days = worked_days
            payslip.save(update_fields=["worked_days"])
        else:
            worked_days = float(payslip.worked_days or 0.0)

        # Perform recalculation
        month_obj = payslip.month
        total_days = calendar.monthrange(month_obj.year, month_obj.month)[1]

        breakdown = calculate_payslip_breakdown(
            new_ctc,
            worked_days,
            total_days,
            employee.pf_enabled,
            location=employee.location,
            company=employee.company,
            month=month_obj.month,
            year=month_obj.year,
        )

        if breakdown.get("country_code", "IN") == "IN":
            conveyance = breakdown.get("lta", 0.0)
            special = breakdown.get("other_allowance", 0.0)
        else:
            conveyance = breakdown.get("conveyance", 0.0)
            special = breakdown.get("medical", 0.0)

        basic = breakdown["basic"]
        hra = breakdown["hra"]
        employee_pf = breakdown["employee_pf"]
        employer_pf = breakdown["employer_pf"]
        professional_tax = breakdown["professional_tax"]
        gross = breakdown["gross_monthly"]
        net = breakdown["net_salary"]
        monthly_gross = breakdown["full_monthly_gross"]

        # Update travel_allowance & tds_deduction if passed or keep existing
        travel_allowance_val = data.get("travel_allowance")
        if travel_allowance_val is not None:
            payslip.travel_allowance = float(travel_allowance_val)

        tds_deduction_val = data.get("tds_deduction")
        if tds_deduction_val is not None:
            payslip.tds_deduction = float(tds_deduction_val)

        travel_allowance = float(payslip.travel_allowance or 0.0)
        tds_deduction = float(payslip.tds_deduction or 0.0)

        # Update payslip with recalculated values
        payslip.basic = basic
        payslip.hra = hra
        payslip.conveyance_allowance = conveyance
        payslip.special_allowance = special
        payslip.employee_pf = employee_pf
        payslip.employer_pf = employer_pf
        payslip.professional_tax = professional_tax
        payslip.gross_salary = gross + travel_allowance
        payslip.monthly_gross = monthly_gross + travel_allowance
        payslip.net_salary = net + travel_allowance - tds_deduction
        payslip.save()

        # Audit log
        FinanceAuditLog.objects.create(
            user=request.user,
            company=employee.company,
            action="RECALCULATE_DRAFT",
            details=f"Recalculated draft payslip for {employee.user.get_full_name()} (CTC: {new_ctc}, Worked Days: {worked_days}). Net: {payslip.net_salary}.",
            ip_address=get_client_ip(request),
        )

        return JsonResponse(
            {
                "status": "success",
                "annual_ctc": new_ctc,
                "worked_days": worked_days,
                "monthly_gross": round(new_ctc / 12, 2),
                "basic": round(basic, 2),
                "hra": round(hra, 2),
                "conveyance": round(conveyance, 2),
                "special": round(special, 2),
                "employee_pf": round(employee_pf, 2),
                "employer_pf": round(employer_pf, 2),
                "professional_tax": round(professional_tax, 2),
                "gross": round(payslip.gross_salary, 2),
                "net": round(payslip.net_salary, 2),
                "travel_allowance": round(travel_allowance, 2),
                "tds_deduction": round(tds_deduction, 2),
            }
        )

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)


@finance_manager_required
def company_payroll_settings(request, company_id):
    """
    Configure statutory rules, salary component mapping, and per-location PT slabs
    specifically for the given company_id. Accessible by finance managers.
    """
    from companies.models import LocationProfessionalTax, PayrollConfiguration

    company = get_object_or_404(Company, id=company_id)
    config, created = PayrollConfiguration.objects.get_or_create(company=company)

    # Fetch all active locations for this company
    locations = company.locations.filter(is_active=True).order_by("name")

    # Ensure every location has a PT config row (auto-create with defaults)
    for loc in locations:
        LocationProfessionalTax.objects.get_or_create(location=loc)

    # Reload with PT configs attached
    location_pt_configs = (
        LocationProfessionalTax.objects.filter(location__company=company, location__is_active=True)
        .select_related("location")
        .order_by("location__name")
    )

    if request.method == "POST":
        # Extract decimal values safely
        def get_decimal(key, default):
            val = request.POST.get(key)
            if val is None or val == "":
                return default
            try:
                return float(val)
            except ValueError:
                return default

        # India Rates
        config.pf_employer_rate = get_decimal("pf_employer_rate", config.pf_employer_rate)
        config.pf_employee_rate = get_decimal("pf_employee_rate", config.pf_employee_rate)
        config.pf_ceiling = get_decimal("pf_ceiling", config.pf_ceiling)
        config.esi_employer_rate = get_decimal("esi_employer_rate", config.esi_employer_rate)
        config.esi_employee_rate = get_decimal("esi_employee_rate", config.esi_employee_rate)
        config.esi_ceiling = get_decimal("esi_ceiling", config.esi_ceiling)
        config.pt_threshold = get_decimal("pt_threshold", config.pt_threshold)
        config.pt_amount_below = get_decimal("pt_amount_below", config.pt_amount_below)
        config.pt_amount_above = get_decimal("pt_amount_above", config.pt_amount_above)

        # Salary Components
        config.basic_percentage = get_decimal("basic_percentage", config.basic_percentage)
        config.hra_percentage = get_decimal("hra_percentage", config.hra_percentage)
        config.lta_percentage = get_decimal("lta_percentage", config.lta_percentage)
        config.special_allowance_percentage = get_decimal(
            "special_allowance_percentage", config.special_allowance_percentage
        )

        # BD Rates
        config.bd_basic_percentage = get_decimal("bd_basic_percentage", config.bd_basic_percentage)
        config.bd_hra_percentage = get_decimal("bd_hra_percentage", config.bd_hra_percentage)
        config.bd_medical_percentage = get_decimal("bd_medical_percentage", config.bd_medical_percentage)
        config.bd_conveyance_percentage = get_decimal("bd_conveyance_percentage", config.bd_conveyance_percentage)

        # US Rates
        config.us_basic_percentage = get_decimal("us_basic_percentage", config.us_basic_percentage)
        config.us_tax_percentage = get_decimal("us_tax_percentage", config.us_tax_percentage)

        config.save()

        # Save per-location PT configs
        for lpt in location_pt_configs:
            loc_id = lpt.location.id
            lpt.pt_threshold = get_decimal(f"loc_pt_threshold_{loc_id}", lpt.pt_threshold)
            lpt.pt_amount_below = get_decimal(f"loc_pt_below_{loc_id}", lpt.pt_amount_below)
            lpt.pt_amount_above = get_decimal(f"loc_pt_above_{loc_id}", lpt.pt_amount_above)
            lpt.is_active = request.POST.get(f"loc_pt_active_{loc_id}") == "on"
            lpt.save()

        # Log change to security audit log
        FinanceAuditLog.objects.create(
            user=request.user,
            company=company,
            action="UPDATE_PAYROLL_CONFIG",
            details=f"Updated payroll configuration for {company.name}.",
            ip_address=get_client_ip(request),
        )

        messages.success(request, f"Payroll configuration for {company.name} updated successfully.")
        return redirect(f"/finance/?company={company.id}")

    return render(
        request,
        "core/payroll_settings.html",
        {
            "config": config,
            "company": company,
            "location_pt_configs": location_pt_configs,
        },
    )


@finance_manager_required
def search_employees_finance(request):
    """API endpoint for searching employees within the finance portal scope"""
    query = request.GET.get("q", "").strip()
    company_id = request.GET.get("company", "all")

    month_val = request.GET.get("month")
    year_val = request.GET.get("year")

    if not query or len(query) < 1:
        return JsonResponse({"employees": [], "debug": "Query too short"})

    employees = Employee.objects.all()
    if company_id and company_id != "all":
        employees = employees.filter(company_id=company_id)

    employees = employees.filter(
        Q(user__first_name__icontains=query) | Q(user__last_name__icontains=query) | Q(badge_id__icontains=query)
    ).select_related("user", "location", "company")[:15]

    results = []
    for emp in employees:
        has_payslip = False
        saved_worked_days = None
        saved_travel_allowance = 0.0
        saved_tds_deduction = 0.0
        if month_val and year_val:
            try:
                from datetime import date

                payslip_date = date(int(year_val), int(month_val), 1)
                payslip = Payslip.objects.filter(employee=emp, month=payslip_date).first()
                if payslip:
                    has_payslip = True
                    saved_worked_days = float(payslip.worked_days) if payslip.worked_days is not None else None
                    saved_travel_allowance = float(payslip.travel_allowance or 0.0)
                    saved_tds_deduction = float(payslip.tds_deduction or 0.0)
            except (ValueError, TypeError):
                pass

        result = {
            "id": emp.id,
            "name": emp.user.get_full_name() or "No Name",
            "employee_id": emp.badge_id or f"EMP-{emp.id}",
            "department": emp.department or "N/A",
            "location": emp.location.name if emp.location else "N/A",
            "designation": emp.designation or "N/A",
            "email": emp.user.email,
            "phone": emp.mobile_number or "N/A",
            "status": emp.employment_status,
            "exit_date": emp.exit_date.strftime("%Y-%m-%d") if emp.exit_date else None,
            "annual_ctc": float(emp.annual_ctc) if emp.annual_ctc else 0,
            "pf_enabled": emp.pf_enabled,
            "currency": emp.location.currency
            if emp.location and hasattr(emp.location, "currency")
            else emp.company.currency
            if hasattr(emp.company, "currency")
            else "INR",
            "country_code": emp.location.country_code.upper()
            if emp.location and hasattr(emp.location, "country_code")
            else "IN",
            "has_payslip": has_payslip,
            "saved_worked_days": saved_worked_days,
            "saved_travel_allowance": saved_travel_allowance,
            "saved_tds_deduction": saved_tds_deduction,
        }
        results.append(result)

    return JsonResponse({"employees": results, "count": len(results)})


@finance_manager_required
@csrf_exempt
def calculate_payslip_preview(request):
    """API to calculate payslip breakdown without saving in the finance portal"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            employee_id = data.get("employee_id")
            annual_ctc = data.get("annual_ctc")
            worked_days = data.get("worked_days")
            month = int(data.get("month"))
            year = int(data.get("year"))
            pf_enabled = data.get("pf_enabled")
            travel_allowance = float(data.get("travel_allowance", 0.0) or 0.0)
            tds_deduction = float(data.get("tds_deduction", 0.0) or 0.0)

            # A finance manager can manage any employee in the active companies
            employee = get_object_or_404(Employee, id=employee_id)
            if pf_enabled is None:
                pf_enabled = employee.pf_enabled

            total_days = calendar.monthrange(year, month)[1]

            breakdown = calculate_payslip_breakdown(
                annual_ctc,
                worked_days,
                total_days,
                pf_enabled,
                location=employee.location,
                company=employee.company,
                month=month,
                year=year,
                travel_allowance=travel_allowance,
                tds_deduction=tds_deduction,
            )
            return JsonResponse({"status": "success", "breakdown": breakdown})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)


@finance_manager_required
def process_single_payroll(request):
    """Process single employee payroll from the calculator (generates draft payslip)"""
    if request.method != "POST":
        return redirect("finance_portal:dashboard")

    employee_id = request.POST.get("employee_id")
    annual_ctc = request.POST.get("annual_ctc")
    worked_days = float(request.POST.get("worked_days") or 0)
    month = int(request.POST.get("month"))
    year = int(request.POST.get("year"))
    pf_enabled = request.POST.get("pf_enabled") == "on"
    travel_allowance = 0.0
    if request.POST.get("travel_allowance_enabled") == "on":
        travel_allowance = float(request.POST.get("travel_allowance") or 0.0)
    tds_deduction = 0.0
    if request.POST.get("tds_deduction_enabled") == "on":
        tds_deduction = float(request.POST.get("tds_deduction") or 0.0)

    try:
        employee = get_object_or_404(Employee, id=employee_id)

        # Update employee's annual CTC and PF status if changed in the calculator
        if annual_ctc:
            new_ctc = float(str(annual_ctc).replace(",", ""))
            if float(employee.annual_ctc or 0) != new_ctc:
                employee.annual_ctc = new_ctc
        employee.pf_enabled = pf_enabled
        employee.save()

        # Generate Phase 1 draft payslip (is_draft=True)
        generate_payslip_internal(
            employee,
            month,
            year,
            worked_days=worked_days,
            travel_allowance=travel_allowance,
            tds_deduction=tds_deduction,
            is_draft=True,
        )
        messages.success(request, f"Successfully generated draft payslip for {employee.user.get_full_name()}!")

        # Log change to security audit log
        FinanceAuditLog.objects.create(
            user=request.user,
            company=employee.company,
            action="SINGLE_PAYROLL_DRAFT",
            details=f"Generated draft payslip for {employee.user.get_full_name()} ({month}/{year}) via calculator.",
            ip_address=get_client_ip(request),
        )

    except Exception as e:
        messages.error(request, f"Error generating payslip: {str(e)}")

    selected_company_id = request.GET.get("company", "all")
    return redirect(f"/finance/?company={selected_company_id}&month={month}&year={year}")

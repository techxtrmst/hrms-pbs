import contextlib
import json
import logging
from datetime import timedelta

import pytz
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.views.generic import CreateView, DeleteView, FormView, ListView, UpdateView
from loguru import logger
from timezonefinder import TimezoneFinder

from accounts.models import User
from core.error_handling import (
    capture_exception,
    safe_get_employee_profile,
    safe_parse_location,
    safe_queryset_filter,
)
from core.utils import normalize_timezone

from .forms import (
    EmployeeBulkImportForm,
    EmployeeCreationForm,
    EmployeeUpdateForm,
    LeaveApplicationForm,
    RegularizationRequestForm,
)
from .models import (
    Attendance,
    AttendanceSession,
    Employee,
    EmployeeIDProof,
    LeaveBalance,
    LeaveRequest,
    LocationLog,
    RegularizationRequest,
)


def detect_timezone_from_coordinates(lat, lng):
    """
    Detect timezone from latitude and longitude coordinates
    """
    try:
        tf = TimezoneFinder()
        timezone_name = tf.timezone_at(lat=float(lat), lng=float(lng))
        return timezone_name if timezone_name else "Asia/Kolkata"
    except Exception as e:
        logger.warning("Error detecting timezone", lat=lat, lng=lng, error=str(e))
        return "Asia/Kolkata"


class CompanyAdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.role == User.Role.COMPANY_ADMIN


class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = "employees/employee_list.html"
    context_object_name = "employees"

    def get_queryset(self):
        user = self.request.user

        # Base queryset based on role
        if user.role == User.Role.COMPANY_ADMIN:
            queryset = Employee.objects.filter(company=user.company)
        elif user.role == User.Role.MANAGER:
            # Show only employees who report to this manager AND are in the same location
            try:
                manager_employee = Employee.objects.get(user=user)
                manager_location = manager_employee.location

                # Filter by manager AND same location
                queryset = Employee.objects.filter(
                    manager=user,
                    location=manager_location,  # Only same location employees
                )
            except Employee.DoesNotExist:
                queryset = Employee.objects.none()
        else:
            queryset = safe_queryset_filter(Employee, user=user)

        # Apply employment status filter
        status_filter = self.request.GET.get("status", "active")

        if status_filter == "active":
            queryset = queryset.filter(is_active=True)
        elif status_filter == "inactive":
            queryset = queryset.filter(is_active=False)
        elif status_filter == "resigned":
            queryset = queryset.filter(employment_status="RESIGNED")
        elif status_filter == "absconded":
            queryset = queryset.filter(employment_status="ABSCONDED")
        elif status_filter == "terminated":
            queryset = queryset.filter(employment_status="TERMINATED")
        # 'all' shows everyone

        return queryset.select_related("user", "company", "manager", "location")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employees = self.get_queryset()
        departments = employees.values_list("department", flat=True).distinct()
        departments_count = len([d for d in departments if d])
        context["departments_count"] = departments_count

        # Add counts for different statuses
        user = self.request.user
        if user.role == User.Role.COMPANY_ADMIN:
            all_employees = Employee.objects.filter(company=user.company)
        elif user.role == User.Role.MANAGER:
            # Show only employees who report to this manager AND are in the same location
            try:
                manager_employee = Employee.objects.get(user=user)
                manager_location = manager_employee.location
                all_employees = Employee.objects.filter(
                    manager=user,
                    location=manager_location,  # Only same location employees
                )
            except Employee.DoesNotExist:
                all_employees = Employee.objects.none()
        else:
            all_employees = safe_queryset_filter(Employee, user=user)

        context["active_count"] = all_employees.filter(is_active=True).count()
        context["inactive_count"] = all_employees.filter(is_active=False).count()
        context["selected_filter"] = self.request.GET.get("status", "active")

        # Check if pseudo name should be shown (for Bluebix and Softstandard)
        show_pseudo_name = False
        if user.role == User.Role.COMPANY_ADMIN and user.company:
            company_name = user.company.name.lower()
            if "bluebix" in company_name or "softstandard" in company_name:
                show_pseudo_name = True
        elif user.role == User.Role.SUPERADMIN:
            # For superadmins, check if any employee in the current queryset belongs to these companies
            show_pseudo_name = employees.filter(
                Q(company__name__icontains="bluebix") | Q(company__name__icontains="softstandard")
            ).exists()
        context["show_pseudo_name"] = show_pseudo_name

        return context


class EmployeeCreateView(LoginRequiredMixin, CompanyAdminRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeCreationForm
    template_name = "employees/employee_form.html"
    success_url = reverse_lazy("employee_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        return super().form_valid(form)


class EmployeeUpdateView(LoginRequiredMixin, CompanyAdminRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeUpdateForm
    template_name = "employees/employee_form.html"
    success_url = reverse_lazy("employee_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_queryset(self):
        return Employee.objects.filter(company=self.request.user.company)

    def form_valid(self, form):
        response = super().form_valid(form)
        employee = self.object

        # Handle emergency contacts
        from .models import EmergencyContact

        # Collect emergency contacts from POST data
        contact_index = 0
        processed_contact_ids = []

        while True:
            name_key = f"emergency_contact_name_{contact_index}"

            if name_key not in self.request.POST:
                break

            name = self.request.POST.get(name_key, "").strip()
            phone = self.request.POST.get(f"emergency_contact_phone_{contact_index}", "").strip()
            relationship = self.request.POST.get(f"emergency_contact_relationship_{contact_index}", "").strip()
            is_primary = f"emergency_contact_primary_{contact_index}" in self.request.POST
            contact_id = self.request.POST.get(f"emergency_contact_id_{contact_index}", "").strip()

            # Only process if at least name and phone are provided
            if name and phone:
                if contact_id:
                    # Update existing contact
                    try:
                        contact = EmergencyContact.objects.get(id=contact_id, employee=employee)
                        contact.name = name
                        contact.phone_number = phone
                        contact.relationship = relationship
                        contact.is_primary = is_primary
                        contact.save()
                        processed_contact_ids.append(int(contact_id))
                    except EmergencyContact.DoesNotExist:
                        pass
                else:
                    # Create new contact
                    contact = EmergencyContact.objects.create(
                        employee=employee,
                        name=name,
                        phone_number=phone,
                        relationship=relationship,
                        is_primary=is_primary,
                    )
                    processed_contact_ids.append(contact.id)

            contact_index += 1

        # Delete contacts that were marked for deletion
        for key in self.request.POST:
            if key.startswith("emergency_contact_delete_"):
                contact_id = key.replace("emergency_contact_delete_", "")
                try:
                    EmergencyContact.objects.filter(id=contact_id, employee=employee).delete()
                except Exception as e:
                    logger.warning(
                        "Failed to delete emergency contact",
                        contact_id=contact_id,
                        error=str(e),
                    )

        # Delete contacts that were removed (not in processed list)
        # This handles contacts that were removed from the form
        existing_contacts = EmergencyContact.objects.filter(employee=employee)
        for contact in existing_contacts:
            if (
                contact.id not in processed_contact_ids
                and f"emergency_contact_delete_{contact.id}" in self.request.POST
            ):
                # Check if it's marked for deletion
                contact.delete()

        return response


class EmployeeDeleteView(LoginRequiredMixin, CompanyAdminRequiredMixin, DeleteView):
    model = Employee
    template_name = "employees/employee_confirm_delete.html"
    success_url = reverse_lazy("employee_list")

    def get_queryset(self):
        return Employee.objects.filter(company=self.request.user.company)

    @transaction.atomic
    def form_valid(self, form):
        """
        Perform complete permanent deletion of employee and all related data.
        This mimics Keka's permanent delete functionality.
        """
        from django.contrib import messages
        from django.shortcuts import redirect

        employee = self.get_object()
        employee_name = employee.user.get_full_name() if employee.user else "Employee"

        try:
            # Log the deletion for audit purposes
            logger.info(
                f"Permanent deletion initiated for employee: {employee_name} (ID: {employee.id})",
                user=self.request.user.email,
            )

            # Delete all related data in the correct order to avoid foreign key constraints

            # 1. Delete Attendance Sessions and Location Logs
            if hasattr(employee, "attendances"):
                for attendance in employee.attendances.all():
                    # Delete session location logs
                    if hasattr(attendance, "sessions"):
                        for session in attendance.sessions.all():
                            # Delete session-specific location logs
                            if hasattr(session, "location_logs"):
                                session.location_logs.all().delete()
                            session.delete()
                    attendance.delete()

            # 2. Delete Location Logs (general)
            if hasattr(employee, "location_logs"):
                employee.location_logs.all().delete()

            # 3. Delete Leave Requests
            if hasattr(employee, "leave_requests"):
                employee.leave_requests.all().delete()

            # 4. Delete Leave Balances
            if hasattr(employee, "leave_balances"):
                employee.leave_balances.all().delete()

            # 5. Delete Regularization Requests
            if hasattr(employee, "regularization_requests"):
                employee.regularization_requests.all().delete()

            # 6. Delete Emergency Contacts
            if hasattr(employee, "emergency_contacts"):
                employee.emergency_contacts.all().delete()

            # 7. Delete ID Proofs
            if hasattr(employee, "id_proofs"):
                try:
                    employee.id_proofs.delete()
                except Exception as e:
                    logger.warning(f"Error deleting ID proofs: {e}")

            # 8. Delete the User account (this will cascade delete the Employee due to CASCADE)
            user = employee.user
            if user:
                user_email = user.email
                user.delete()  # This will also delete the employee due to CASCADE
                logger.info(f"Successfully deleted user account: {user_email}")
            else:
                # If no user exists, delete employee directly
                employee.delete()
                logger.info(f"Successfully deleted employee: {employee_name}")

            # Success message
            messages.success(
                self.request,
                f"Employee '{employee_name}' and all associated data have been permanently deleted.",
            )

            logger.info(f"Permanent deletion completed for: {employee_name}")

        except Exception as e:
            logger.error(f"Error during permanent deletion: {str(e)}", exc_info=True)
            messages.error(self.request, f"An error occurred during deletion: {str(e)}")

        return redirect(self.success_url)


@login_required
def resend_welcome_email(request, pk):
    """
    Resend welcome email with activation link to the employee.
    """
    if request.user.role != User.Role.COMPANY_ADMIN:
        messages.error(request, "Permission denied.")
        return redirect("employee_list")

    try:
        employee = Employee.objects.get(pk=pk, company=request.user.company)
        from core.email_utils import send_welcome_email_with_link

        domain = request.get_host()
        if send_welcome_email_with_link(employee, domain):
            messages.success(request, f"Welcome email resent to {employee.user.email}")
        else:
            messages.error(request, "Failed to send email. Please check email settings.")

    except Employee.DoesNotExist:
        messages.error(request, "Employee not found.")

    return redirect("employee_list")


# --- Attendance & Tracking Views ---


@csrf_exempt
@login_required
def clock_in(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            lat = data.get("latitude")
            lng = data.get("longitude")
            accuracy = data.get("accuracy")
            clock_in_type = data.get("type", "office")  # 'office' or 'remote'

            # Ensure employee profile exists
            if not hasattr(request.user, "employee_profile"):
                # Auto-create for Company Admin to prevent setup deadlock
                if request.user.role == User.Role.COMPANY_ADMIN and request.user.company:
                    from .models import Employee

                    try:
                        Employee.objects.create(
                            user=request.user,
                            company=request.user.company,
                            designation="Administrator",
                            department="Management",
                            badge_id=f"ADM{request.user.id}",
                        )
                        # Refresh user to get the profile
                        request.user.refresh_from_db()
                    except Exception as e:
                        logger.error(f"Failed to auto-create profile in clock-in: {e}")
                        return JsonResponse(
                            {
                                "status": "error",
                                "message": "No employee profile found. Please contact support.",
                            },
                            status=400,
                        )
                else:
                    return JsonResponse(
                        {
                            "status": "error",
                            "message": "No employee profile found. Please set up your profile first.",
                        },
                        status=400,
                    )

            employee = request.user.employee_profile

            # --- TIMEZONE RESOLUTION (Assigned Work Location is Boss) ---
            from core.utils import get_user_timezone

            # Get the official timezone assigned to the employee's work location
            user_timezone = get_user_timezone(request.user, getattr(request, "company", None))

            # Only use browser detected timezone if the employee has NO assigned location/timezone in the system
            # This prevents a laptop set to 'USA' from overriding the official 'India' work location.
            if not user_timezone:
                detected_tz = data.get("timezone")
                if not detected_tz:
                    detected_tz = detect_timezone_from_coordinates(lat, lng)
                user_timezone = detected_tz or "Asia/Kolkata"

            # Calculate today's work date strictly based on the Official Timezone
            import pytz

            try:
                tz = pytz.timezone(user_timezone)
                today = timezone.now().astimezone(tz).date()
            except Exception:
                # Absolute fallback to India if anything fails
                user_timezone = "Asia/Kolkata"
                tz = pytz.timezone(user_timezone)
                today = timezone.now().astimezone(tz).date()

            # Check if employee is already clocked in for ANY date (Night Shift Support)
            active_attendance = Attendance.objects.filter(employee=employee, is_currently_clocked_in=True).first()
            if active_attendance:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "You are already clocked in. Please clock out first.",
                        "already_clocked_in": True,
                        "clock_in_date": active_attendance.date.strftime("%Y-%m-%d"),
                    }
                )

            # Get or create attendance record for today
            attendance, created = Attendance.objects.get_or_create(
                employee=employee,
                date=today,
                defaults={
                    "status": "ABSENT",
                    "daily_sessions_count": 0,
                    "is_currently_clocked_in": False,
                    "user_timezone": user_timezone,
                },
            )

            # Update timezone if it was created/fetched without it or if it changed
            if attendance.user_timezone != user_timezone:
                attendance.user_timezone = user_timezone
                attendance.save(update_fields=["user_timezone"])

            # Check if employee can clock in
            if not attendance.can_clock_in() and attendance.is_currently_clocked_in:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "You are already clocked in. Please clock out first.",
                        "already_clocked_in": True,
                    }
                )

            # Determine session type and status
            session_type = "WEB" if clock_in_type == "office" else "REMOTE"

            # Use database transaction to prevent race conditions
            from django.db import transaction

            try:
                with transaction.atomic():
                    # Refresh attendance from database to get latest state
                    attendance.refresh_from_db()

                    # Create new session with proper session number
                    session_number = attendance.daily_sessions_count + 1

                    # Check if session already exists (race condition protection)
                    existing_session = AttendanceSession.objects.filter(
                        employee=employee, date=today, session_number=session_number
                    ).first()

                    if existing_session:
                        return JsonResponse(
                            {
                                "status": "error",
                                "message": "Session already exists. Please refresh the page.",
                            }
                        )

                    session = AttendanceSession.objects.create(
                        employee=employee,
                        date=today,
                        session_number=session_number,
                        clock_in=timezone.now(),
                        session_type=session_type,
                        clock_in_latitude=lat if lat is not None else None,
                        clock_in_longitude=lng if lng is not None else None,
                        is_active=True,
                    )

                    # Log clock-in location only if coordinates are available
                    if lat is not None and lng is not None:
                        LocationLog.objects.create(
                            employee=employee,
                            attendance_session=session,
                            latitude=lat,
                            longitude=lng,
                            accuracy=accuracy if accuracy is not None else 9999,  # Use high value for unknown accuracy
                            log_type="CLOCK_IN",
                            is_valid=True,
                        )
                    else:
                        # Log with null coordinates to track that location was unavailable
                        logger.warning(f"Clock-in without location data for {employee.user.get_full_name()}")

                    # Update attendance record
                    attendance.daily_sessions_count = session_number
                    attendance.is_currently_clocked_in = True
                    attendance.current_session_type = session_type
                    attendance.user_timezone = user_timezone

                    # Set first clock-in of the day
                    if not attendance.clock_in:
                        attendance.clock_in = session.clock_in
                        attendance.location_in = f"{lat},{lng}" if lat is not None and lng is not None else "N/A"

                    # Determine overall status
                    if session_number == 1:
                        attendance.status = "WFH" if session_type == "REMOTE" else "PRESENT"
                    else:
                        # Multiple sessions - check if mixed types
                        session_types = set(
                            AttendanceSession.objects.filter(employee=employee, date=today).values_list(
                                "session_type", flat=True
                            )
                        )
                        if len(session_types) > 1:
                            attendance.status = "HYBRID"
                        else:
                            attendance.status = "WFH" if session_type == "REMOTE" else "PRESENT"

                    # Start location tracking
                    attendance.location_tracking_active = True

                    # Calculate location tracking end time based on shift duration
                    shift = employee.assigned_shift
                    if shift:
                        if hasattr(shift, "get_shift_duration_timedelta"):
                            shift_duration = shift.get_shift_duration_timedelta()
                        else:
                            from datetime import datetime, timedelta

                            today_date = timezone.localdate()
                            s_start = datetime.combine(today_date, shift.start_time)
                            s_end = datetime.combine(today_date, shift.end_time)
                            if s_end < s_start:
                                s_end += timedelta(days=1)
                            shift_duration = s_end - s_start

                        attendance.location_tracking_end_time = session.clock_in + shift_duration
                    else:
                        # Default to 9 hours if no shift assigned
                        from datetime import timedelta

                        attendance.location_tracking_end_time = session.clock_in + timedelta(hours=9)

                    # Calculate late arrival for first session only
                    if session_number == 1:
                        attendance.calculate_late_arrival()

                    attendance.save()

            except Exception as db_error:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": f"Database error: {str(db_error)}. Please try again.",
                    },
                    status=500,
                )

            # Check for late clock-ins in the last 7 days
            late_warning = None
            if session_number == 1 and (attendance.is_late or attendance.is_grace_used):
                from datetime import timedelta

                seven_days_ago = today - timedelta(days=7)

                # Count late clock-ins in the last 7 days (excluding today)
                late_count = (
                    Attendance.objects.filter(
                        employee=employee,
                        date__gte=seven_days_ago,
                        date__lt=today,  # Exclude today
                    )
                    .filter(Q(is_late=True) | Q(is_grace_used=True))
                    .count()
                )

                # Include today's late arrival in the count
                total_late_count = late_count + 1

                # Prepare warning message
                if total_late_count >= 5:
                    late_warning = {
                        "show_warning": True,
                        "late_count": total_late_count,
                        "message": f"You have been late {total_late_count} times in the last 7 days.",
                        "action": "LOP will be applied. Please ensure timely attendance.",
                        "severity": "critical",
                    }
                else:
                    late_warning = {
                        "show_warning": True,
                        "late_count": total_late_count,
                        "message": f"You have been late {total_late_count} times in the last 7 days.",
                        "action": "Please ensure timely attendance to avoid LOP.",
                        "severity": "warning",
                    }

            response_data = {
                "status": "success",
                "message": f"Successfully clocked in for session {session_number} ({session_type.lower()})",
                "session_number": session_number,
                "session_type": session_type,
                "clock_in_time": session.clock_in.strftime("%H:%M:%S"),
                "total_sessions_today": attendance.daily_sessions_count,
                "max_sessions": attendance.max_daily_sessions,
            }

            # Add late warning if applicable
            if late_warning:
                response_data["late_warning"] = late_warning

            return JsonResponse(response_data)

        except Exception as e:
            logger.error(f"Clock-in error: {str(e)}", exc_info=True)
            # print(f"Clock-in error: {str(e)}")
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)


@csrf_exempt
@login_required
def clock_out(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            lat = data.get("latitude")
            lng = data.get("longitude")
            accuracy = data.get("accuracy")
            force_clockout = data.get("force_clockout", False)

            if not hasattr(request.user, "employee_profile"):
                return JsonResponse(
                    {"status": "error", "message": "No employee profile found"},
                    status=400,
                )

            employee = request.user.employee_profile

            # Find active attendance regardless of current date (Keka-style Night Shift Logic)
            attendance = Attendance.objects.filter(employee=employee, is_currently_clocked_in=True).first()

            # Determine today based on employee's location timezone
            from core.utils import get_user_timezone

            user_timezone = get_user_timezone(request.user, getattr(request, "company", None))

            # Override with active record's timezone if available
            if attendance and attendance.user_timezone:
                user_timezone = attendance.user_timezone

            try:
                tz = pytz.timezone(user_timezone)
                today = timezone.now().astimezone(tz).date()
            except Exception:
                today = timezone.localdate()

            try:
                # Fallback to today's record if no active session found (to show appropriate error)
                if not attendance:
                    attendance = Attendance.objects.filter(employee=employee, date=today).first()

                if not attendance:
                    return JsonResponse(
                        {
                            "status": "error",
                            "message": "No attendance record found for today",
                        }
                    )

                # Check if currently clocked in
                if not attendance.is_currently_clocked_in:
                    return JsonResponse(
                        {
                            "status": "error",
                            "message": "You are not currently clocked in.",
                        }
                    )

                # Get current active session
                current_session = attendance.get_current_session()
                # DEBUG logging removed
                # if current_session:
                #     pass

                if not current_session:
                    # Auto-correct state: Attendance says clocked in but no session is active
                    attendance.is_currently_clocked_in = False
                    attendance.save(update_fields=["is_currently_clocked_in"])

                    return JsonResponse(
                        {
                            "status": "success",
                            "message": "System sync corrected. You are clocked out.",
                        }
                    )

                # Check if shift is complete (unless forced)
                if not force_clockout and current_session.clock_in:
                    # Calculate cumulative working hours from all sessions including current one
                    worked_hours = attendance.get_cumulative_working_hours_including_current()
                    expected_hours = attendance.get_shift_duration_hours()
                    shift = employee.assigned_shift

                    if worked_hours < expected_hours and not (shift and getattr(shift, "is_flexible", False)):
                        completion_percentage = (worked_hours / expected_hours) * 100
                        remaining_hours = expected_hours - worked_hours

                        return JsonResponse(
                            {
                                "status": "confirmation_required",
                                "requires_confirmation": True,
                                "message": f"Your {expected_hours:.1f}-hour shift is not completed yet. Do you want to clock out?",
                                "worked_hours": round(worked_hours, 1),
                                "expected_hours": round(expected_hours, 1),
                                "completion_percentage": round(completion_percentage, 1),
                                "remaining_hours": round(remaining_hours, 1),
                            }
                        )

                # Process clock-out for current session
                current_session.clock_out = timezone.now()
                current_session.clock_out_latitude = lat if lat is not None else None
                current_session.clock_out_longitude = lng if lng is not None else None
                current_session.is_active = False
                current_session.save()  # This will auto-calculate duration

                # Log clock-out location only if coordinates are available
                if lat is not None and lng is not None:
                    LocationLog.objects.create(
                        employee=employee,
                        attendance_session=current_session,
                        latitude=lat,
                        longitude=lng,
                        accuracy=accuracy if accuracy is not None else 9999,  # Use high value for unknown accuracy
                        log_type="CLOCK_OUT",
                        is_valid=True,
                    )
                else:
                    # Log that location was unavailable
                    logger.warning(f"Clock-out without location data for {employee.user.get_full_name()}")

                # Update attendance record
                attendance.is_currently_clocked_in = False
                attendance.current_session_type = None
                attendance.clock_out = current_session.clock_out  # Update last clock-out
                attendance.location_out = f"{lat},{lng}" if lat is not None and lng is not None else "N/A"

                # Stop location tracking
                attendance.location_tracking_active = False

                # Calculate total working hours
                attendance.calculate_total_working_hours()
                attendance.save()

                return JsonResponse(
                    {
                        "status": "success",
                        "message": f"Successfully clocked out from session {current_session.session_number} ({current_session.session_type.lower()})",
                        "session_number": current_session.session_number,
                        "session_type": current_session.session_type,
                        "session_duration": current_session.duration_hours,
                        "clock_out_time": current_session.clock_out.strftime("%H:%M:%S"),
                        "total_working_hours": attendance.total_working_hours,
                        "sessions_remaining": attendance.max_daily_sessions - attendance.daily_sessions_count,
                    }
                )

            except Attendance.DoesNotExist:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "No attendance record found for today",
                    }
                )

        except Exception as e:
            logger.error(f"Clock-out error: {str(e)}", exc_info=True)
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)


def perform_auto_clock_out(attendance, session, lat, lng):
    """
    Helper to perform system-triggered clock out when shift ends
    """
    try:
        current_time = timezone.now()

        # Process clock-out
        session.clock_out = current_time
        session.clock_out_latitude = lat
        session.clock_out_longitude = lng
        session.is_active = False
        session.save()

        # Update attendance
        attendance.is_currently_clocked_in = False
        attendance.current_session_type = None
        attendance.clock_out = current_time
        attendance.location_out = f"{lat},{lng}"
        attendance.location_tracking_active = False

        attendance.calculate_total_working_hours()
        attendance.save()

        return True
    except Exception as e:
        logger.error(f"Auto clock-out failed: {str(e)}")
        return False


@csrf_exempt
@login_required
def update_location(request):
    """API endpoint to update employee's location (or log coordinates).
    Accepts POST JSON with either 'location_id' to set location, or 'latitude'/'longitude' to log coordinates.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)
    try:
        data = json.loads(request.body)
        # Ensure employee profile exists
        if not hasattr(request.user, "employee_profile"):
            return JsonResponse({"status": "error", "message": "No employee profile found"}, status=400)
        employee = request.user.employee_profile

        # Update location if provided
        location_id = data.get("location_id")
        if location_id:
            from companies.models import Location

            try:
                location = Location.objects.get(id=location_id, company=employee.company)
                employee.location = location
                employee.save()
                return JsonResponse({"status": "success", "message": "Location updated"})
            except Location.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Location not found"}, status=404)

        # Otherwise log latitude/longitude (only if tracking is active)
        lat = data.get("latitude")
        lng = data.get("longitude")
        accuracy = data.get("accuracy")

        if lat is not None and lng is not None:
            # Check if location tracking is active for today's attendance
            today = timezone.localdate()
            try:
                attendance = Attendance.objects.get(employee=employee, date=today)

                # Check if currently clocked in and has active session
                if not attendance.is_currently_clocked_in:
                    return JsonResponse(
                        {
                            "status": "not_clocked_in",
                            "message": "Not currently clocked in",
                            "location_tracking_active": False,
                        }
                    )

                # Get current active session
                current_session = attendance.get_current_session()
                if not current_session:
                    return JsonResponse(
                        {
                            "status": "no_active_session",
                            "message": "No active session found",
                            "location_tracking_active": False,
                        }
                    )

                # Log location for current session (Removed auto-clockout as per requirement)
                if current_session.clock_in:
                    pass  # Keep the clock_in check if needed for other logic, but removed the 9-hour constraint

                # Log location for current session
                if attendance.location_tracking_active:
                    # Create session-specific location log
                    from employees.models import SessionLocationLog

                    SessionLocationLog.objects.create(
                        session=current_session,
                        latitude=lat,
                        longitude=lng,
                        accuracy=accuracy,
                    )

                    # Log to generic LocationLog as well - ONLY if accuracy is good
                    # Filter out poor accuracy (likely network based or bad signal) to avoid "fake" look
                    # Threshold: 2500 meters (Relaxed to ensure tracking works for all users/devices)
                    is_accurate = True
                    if accuracy and float(accuracy) > 2500:
                        is_accurate = False

                    if is_accurate:
                        # Also create general location log for backward compatibility
                        LocationLog.objects.create(employee=employee, latitude=str(lat), longitude=str(lng))

                    # Prepare response
                    response_data = {
                        "status": "success",
                        "message": f"Location logged for Session {current_session.session_number}",
                        "location_tracking_active": True,
                        "session_number": current_session.session_number,
                        "session_type": current_session.session_type,
                    }

                    # Check session duration and provide notifications
                    session_duration = timezone.now() - current_session.clock_in
                    session_hours = session_duration.total_seconds() / 3600
                    if session_hours >= 8:
                        response_data["session_completed"] = True
                        response_data["notification"] = (
                            f"Session {current_session.session_number} completed ({session_hours:.1f} hours). Consider clocking out."
                        )
                    elif session_hours >= 4:
                        response_data["session_progress"] = (
                            f"Session {current_session.session_number} in progress ({session_hours:.1f} hours)"
                        )

                    return JsonResponse(response_data)
                else:
                    return JsonResponse(
                        {
                            "status": "tracking_inactive",
                            "message": "Location tracking is not active",
                            "location_tracking_active": False,
                        }
                    )

            except Attendance.DoesNotExist:
                # No attendance record, don't log location
                return JsonResponse(
                    {
                        "status": "no_attendance",
                        "message": "No attendance record found for today",
                        "location_tracking_active": False,
                    }
                )

        # Return 200 even if no data to prevent log spam
        return JsonResponse({"status": "ignored", "message": "No valid data provided"}, status=200)
    except Exception as e:
        err_logger = logging.getLogger(__name__)
        err_logger.error(f"Update location error: {str(e)}", exc_info=True)
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
def employee_profile(request):
    user = request.user

    # Try to get or create employee profile if User is a Company Admin/Manager
    # This prevents the "blank page" issue for the initial admin user.
    employee = safe_get_employee_profile(user)

    if not employee:
        if user.company:
            # Auto-create basic profile for Admin/Manager to avoid UI block
            employee = Employee.objects.create(
                user=user,
                company=user.company,
                designation="Administrator" if user.role == User.Role.COMPANY_ADMIN else "Employee",
                department="Management",
                badge_id=f"ADM{user.id}",  # Simple fallback ID
            )
        else:
            # Fallback if no company (shouldn't happen for active users)
            # We must pass the 'content' block to a template that extends base.html
            return render(
                request,
                "core/general_message.html",
                {
                    "title": "Profile Not Found",
                    "message": "No employee profile found. Please contact your administrator.",
                },
            )

    id_proofs, created = EmployeeIDProof.objects.get_or_create(employee=employee)
    locations = employee.company.locations.all()

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "update_profile":
            is_admin = request.user.role == User.Role.COMPANY_ADMIN
            if employee.profile_edited and not is_admin:
                messages.error(request, "You have already edited your profile once.")
                return redirect("employee_profile")

            try:
                # Personal Details
                employee.mobile_number = request.POST.get("mobile_number")
                employee.personal_email = request.POST.get("personal_email")
                employee.pseudo_name = request.POST.get("pseudo_name")
                dob_str = request.POST.get("dob")
                if dob_str:
                    employee.dob = dob_str
                employee.gender = request.POST.get("gender")
                employee.marital_status = request.POST.get("marital_status")

                # Address Book
                employee.current_address = request.POST.get("current_address")
                employee.permanent_address = request.POST.get("permanent_address")

                employee.save()

                # Emergency Contacts - Clear existing and add new
                from .models import EmergencyContact

                employee.emergency_contacts.all().delete()

                # Contact 1
                c1_name = request.POST.get("contact_name_1")
                c1_phone = request.POST.get("contact_phone_1")
                c1_rel = request.POST.get("contact_rel_1")
                if c1_name and c1_phone:
                    EmergencyContact.objects.create(
                        employee=employee, name=c1_name, phone_number=c1_phone, relationship=c1_rel, is_primary=True
                    )

                # Contact 2
                c2_name = request.POST.get("contact_name_2")
                c2_phone = request.POST.get("contact_phone_2")
                c2_rel = request.POST.get("contact_rel_2")
                if c2_name and c2_phone:
                    EmergencyContact.objects.create(
                        employee=employee, name=c2_name, phone_number=c2_phone, relationship=c2_rel, is_primary=False
                    )

                employee.profile_edited = True
                employee.save(update_fields=["profile_edited"])

                messages.success(request, "Profile updated successfully.")
                return redirect("employee_profile")
            except Exception as e:
                messages.error(request, f"Error updating profile: {str(e)}")
                return redirect("employee_profile")

        elif action == "assign_shift":
            # Only allow Company Admins to assign shifts
            if request.user.role != User.Role.COMPANY_ADMIN and not request.user.is_superuser:
                messages.error(request, "You don't have permission to assign shifts.")
                return redirect("employee_profile")

            try:
                from companies.models import ShiftSchedule

                shift_id = request.POST.get("assigned_shift")
                if shift_id:
                    shift = ShiftSchedule.objects.get(id=shift_id, company=request.user.company)
                    old_shift_name = employee.assigned_shift.name if employee.assigned_shift else "None"
                    employee.assigned_shift = shift
                    employee.save(update_fields=["assigned_shift"])

                    from django.contrib import messages

                    messages.success(
                        request,
                        f"Shift assignment updated from '{old_shift_name}' to '{shift.name}' for {employee.user.get_full_name()}.",
                    )
                else:
                    # Remove shift assignment
                    old_shift_name = employee.assigned_shift.name if employee.assigned_shift else "None"
                    employee.assigned_shift = None
                    employee.save(update_fields=["assigned_shift"])

                    from django.contrib import messages

                    messages.success(
                        request,
                        f"Shift assignment removed from {employee.user.get_full_name()}. Previous shift: '{old_shift_name}'.",
                    )

                return redirect("employee_profile")

            except ShiftSchedule.DoesNotExist:
                from django.contrib import messages

                messages.error(request, "Selected shift not found.")
                return redirect("employee_profile")
            except Exception as e:
                from django.contrib import messages

                messages.error(request, f"Error assigning shift: {str(e)}")
                return redirect("employee_profile")

        # Handle Profile Picture Upload
        if "profile_picture" in request.FILES:
            employee.profile_picture = request.FILES["profile_picture"]
            employee.save()
            return redirect("employee_profile")

        is_admin = request.user.role == User.Role.COMPANY_ADMIN

        # Helper to check if upload allowed
        def can_upload(current_file):
            return not current_file or is_admin

        if "aadhar_front" in request.FILES and can_upload(id_proofs.aadhar_front):
            id_proofs.aadhar_front = request.FILES["aadhar_front"]

        if "aadhar_back" in request.FILES and can_upload(id_proofs.aadhar_back):
            id_proofs.aadhar_back = request.FILES["aadhar_back"]

        if "pan_card" in request.FILES and can_upload(id_proofs.pan_card):
            id_proofs.pan_card = request.FILES["pan_card"]

        id_proofs.save()
        from django.contrib import messages

        messages.success(request, "ID Documents updated successfully.")

        # Deletion logic for Admins
        if is_admin:
            if request.POST.get("delete_aadhar_front") == "on":
                id_proofs.aadhar_front.delete(save=False)
                id_proofs.aadhar_front = None
            if request.POST.get("delete_aadhar_back") == "on":
                id_proofs.aadhar_back.delete(save=False)
                id_proofs.aadhar_back = None
            if request.POST.get("delete_pan_card") == "on":
                id_proofs.pan_card.delete(save=False)
                id_proofs.pan_card = None
            id_proofs.save()

        return redirect("employee_profile")

    # Get emergency contacts for this employee
    emergency_contacts = employee.emergency_contacts.all().order_by("-is_primary", "created_at")

    # Get probation status
    probation_status = employee.get_probation_status() if employee.date_of_joining else "IN_PROBATION"
    probation_date = employee.get_probation_end_date() if employee.date_of_joining else None

    # Get available shifts for the company (for shift assignment)
    available_shifts = []
    if request.user.role == User.Role.COMPANY_ADMIN or request.user.is_superuser:
        from companies.models import ShiftSchedule

        available_shifts = ShiftSchedule.objects.filter(company=request.user.company, is_active=True).order_by("name")

    # Get work history
    from .models import WorkHistory

    work_history = WorkHistory.objects.filter(employee=employee).order_by("-date", "-created_at")

    return render(
        request,
        "employees/employee_profile.html",
        {
            "employee": employee,
            "id_proofs": id_proofs,
            "is_admin": request.user.role == User.Role.COMPANY_ADMIN,
            "locations": locations,
            "emergency_contacts": emergency_contacts,
            "probation_status": probation_status,
            "probation_date": probation_date,
            "available_shifts": available_shifts,
            "work_history": work_history,
        },
    )


@login_required
def set_location(request):
    """Handle location update from personal details form."""
    if request.method != "POST":
        return redirect("employee_profile")
    location_id = request.POST.get("location_id")
    if not location_id:
        messages.error(request, "No location selected.")
        return redirect("employee_profile")
    try:
        from .models import Location

        location = Location.objects.get(id=location_id, company=request.user.company)
        employee = request.user.employee_profile
        employee.location = location
        employee.save()
        messages.success(request, f"Location updated to {location.name}.")
    except Location.DoesNotExist:
        messages.error(request, "Location not found.")
    except Exception as e:
        messages.error(request, f"Error updating location: {str(e)}")
    return redirect("employee_profile")


# --- Leave Management Views ---


class LeaveApplyView(LoginRequiredMixin, CreateView):
    model = LeaveRequest
    form_class = LeaveApplicationForm
    template_name = "employees/leave_form.html"
    success_url = reverse_lazy("employee_profile")  # Or dashboard

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            employee = self.request.user.employee_profile
            if hasattr(employee, "leave_balance"):
                balance = employee.leave_balance
                context["cl_balance"] = balance.casual_leave_balance
                context["sl_balance"] = balance.sick_leave_balance
                context["leave_balance"] = balance
            else:
                context["cl_balance"] = 0
                context["sl_balance"] = 0
        except Exception:
            # Fallback if something goes wrong (e.g. no profile)
            context["cl_balance"] = 0
            context["sl_balance"] = 0
        return context

    def form_valid(self, form):
        employee = self.request.user.employee_profile
        form.instance.employee = employee

        # Multi-level Workflow Integration
        from core.models import ApprovalWorkflow

        workflow = ApprovalWorkflow.objects.filter(
            company=employee.company, workflow_type="LEAVE", is_active=True
        ).first()

        if workflow:
            form.instance.workflow = workflow
            form.instance.current_step = 1
            # Update approval_level based on first step config if needed
            # For now, default to MANAGER as per existing logic but inside workflow

        # Server-side duplicate prevention: Check for recent duplicate submissions
        from datetime import timedelta

        from django.utils import timezone

        recent_duplicate = LeaveRequest.objects.filter(
            employee=employee,
            start_date=form.cleaned_data["start_date"],
            end_date=form.cleaned_data["end_date"],
            leave_type=form.cleaned_data["leave_type"],
            created_at__gte=timezone.now() - timedelta(seconds=10),
        ).exists()

        if recent_duplicate:
            from django.contrib import messages

            messages.warning(
                self.request,
                "You just submitted a leave request for these dates. Please wait before submitting again.",
            )
            return self.form_invalid(form)

        # Check if this is a confirmation submission
        confirm_lop = self.request.POST.get("confirm_lop", "false").lower() == "true"

        # Validate leave application before saving
        temp_leave_request = LeaveRequest(
            employee=form.instance.employee,
            leave_type=form.cleaned_data["leave_type"],
            start_date=form.cleaned_data["start_date"],
            end_date=form.cleaned_data["end_date"],
            duration=form.cleaned_data.get("duration", "FULL"),
        )

        validation = temp_leave_request.validate_leave_application()

        # If validation shows issues and user hasn't confirmed, ask for confirmation
        if validation.get("will_be_lop", False) and form.cleaned_data["leave_type"] != "UL" and not confirm_lop:
            from django.contrib import messages

            messages.error(
                self.request,
                f"⚠️ Insufficient Leave Balance: {validation['message']} Please confirm if you want to proceed with LOP (Loss of Pay).",
            )

            # Return form with validation warning for user confirmation
            return self.render_to_response(
                self.get_context_data(form=form, validation_warning=validation, show_confirmation=True)
            )

        # Add a comment to the leave request if it involves LOP
        if validation.get("will_be_lop", False) and form.cleaned_data["leave_type"] != "UL":
            original_reason = form.cleaned_data.get("reason", "")
            lop_note = f"\n\n[System Note: This application involves {validation.get('shortfall', 0)} days of LOP due to insufficient balance. Available: {validation.get('available_balance', 0)} days, Requested: {validation.get('requested_days', 0)} days]"
            form.instance.reason = original_reason + lop_note

        # Proceed with normal save
        response = super().form_valid(form)

        # Send email notifications asynchronously to avoid blocking the response
        import threading

        def send_email_async():
            try:
                import logging

                from core.email_utils import send_leave_request_notification

                logger = logging.getLogger(__name__)
                result = send_leave_request_notification(self.object)
                if not result.get("hr", False):
                    logger.error(f"Failed to send leave request email to HR for {self.object.id}")
            except Exception as e:
                import logging

                logger = logging.getLogger(__name__)
                logger.error(f"Leave request email error: {str(e)}")

        # Start email sending in background thread
        email_thread = threading.Thread(target=send_email_async)
        email_thread.daemon = True
        email_thread.start()

        return response


@csrf_exempt
@login_required
def check_leave_balance(request):
    """AJAX endpoint to check leave balance for real-time validation"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            leave_type = data.get("leave_type")
            start_date = data.get("start_date")
            end_date = data.get("end_date")
            duration = data.get("duration", "FULL")

            if not hasattr(request.user, "employee_profile"):
                return JsonResponse(
                    {"status": "error", "message": "No employee profile found"},
                    status=400,
                )

            employee = request.user.employee_profile

            # Create temporary leave request for validation
            from datetime import datetime

            temp_leave = LeaveRequest(
                employee=employee,
                leave_type=leave_type,
                start_date=datetime.strptime(start_date, "%Y-%m-%d").date(),
                end_date=datetime.strptime(end_date, "%Y-%m-%d").date(),
                duration=duration,
            )

            validation = temp_leave.validate_leave_application()

            # Get current balances
            balance = employee.leave_balance
            balances = {
                "CL": balance.casual_leave_balance,
                "SL": balance.sick_leave_balance,
            }

            return JsonResponse(
                {
                    "status": "success",
                    "validation": validation,
                    "balances": balances,
                    "requested_days": temp_leave.total_days,
                }
            )

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)


@login_required
def approve_leave(request, pk):
    from django.contrib import messages

    if request.method == "POST":
        leave_request = LeaveRequest.objects.get(pk=pk)
        approval_type = request.POST.get("approval_type", "FULL")

        # Security check: Only Manager or Admin can approve
        user = request.user
        is_admin = user.role == User.Role.COMPANY_ADMIN or user.is_superuser
        is_manager = (
            user.role == User.Role.MANAGER and leave_request.employee.manager and leave_request.employee.manager == user
        )

        if not (is_admin or is_manager):
            return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)

        # Multi-level Workflow Integration
        if leave_request.workflow:
            config = leave_request.workflow.levels_config
            curr_step = str(leave_request.current_step)
            step_config = config.get(curr_step, {})

            # Check if user role matches step config
            required_role = step_config.get("role")
            if required_role and user.role != required_role:
                # If specifically looking for MANAGER, it must be the direct manager
                if required_role == "MANAGER" and not is_manager:
                    return JsonResponse(
                        {"status": "error", "message": "Only the direct manager can approve this step."}, status=403
                    )
                elif user.role != required_role:
                    return JsonResponse(
                        {"status": "error", "message": f"This step requires {required_role} approval."}, status=403
                    )

            # Move to next step or finalize
            if leave_request.current_step < leave_request.workflow.levels:
                leave_request.current_step += 1
                leave_request.save()
                messages.success(
                    request,
                    f"Approver level {curr_step} completed. Request moved to level {leave_request.current_step}.",
                )
                return redirect("leave_requests")
            else:
                # Final level reached, proceed to actual approval
                pass

        # Use the new approval method from the model
        if leave_request.approve_leave(user, approval_type=approval_type):
            # Update Attendance Records
            from datetime import timedelta

            current_date = leave_request.start_date
            while current_date <= leave_request.end_date:
                # Get or Create Attendance
                att_record, created = Attendance.objects.get_or_create(
                    employee=leave_request.employee, date=current_date
                )

                # Update status
                if leave_request.leave_type == "OD":
                    att_record.status = "ON_DUTY"
                elif approval_type == "WITH_LOP" or leave_request.leave_type == "UL":
                    # Mark as LOP for days beyond available balance
                    validation = leave_request.validate_leave_application()
                    days_from_start = (current_date - leave_request.start_date).days
                    available_days = validation.get("available_balance", 0)

                    if days_from_start < available_days:
                        att_record.status = "LEAVE"
                    else:
                        att_record.status = "LEAVE"  # Still mark as leave, LOP tracked in balance
                else:
                    att_record.status = "LEAVE"

                att_record.save()
                current_date += timedelta(days=1)

            # Send Approval Email with approval type info asynchronously
            import threading

            def send_email_async():
                try:
                    import logging

                    from core.email_utils import send_leave_approval_notification

                    logger = logging.getLogger(__name__)
                    if not send_leave_approval_notification(leave_request):
                        logger.warning("Leave approval email notification failed")
                except Exception as e:
                    import logging

                    logger = logging.getLogger(__name__)
                    logger.error(f"Leave approval email error: {str(e)}")

            # Start email sending in background thread
            email_thread = threading.Thread(target=send_email_async)
            email_thread.daemon = True
            email_thread.start()

            # Show success message immediately
            approval_msg = {
                "FULL": "Leave approved successfully (Full Balance).",
                "WITH_LOP": "Leave approved with LOP for excess days.",
                "ONLY_AVAILABLE": "Leave approved for available days only.",
            }.get(approval_type, "Leave approved successfully.")
            messages.success(
                request,
                f"{approval_msg} Notification will be sent to {leave_request.employee.user.first_name}.",
            )

    # Redirect back to leave requests page or dashboard
    return redirect("leave_requests")


@login_required
def reject_leave(request, pk):
    if request.method == "POST":
        leave_request = LeaveRequest.objects.get(pk=pk)

        user = request.user
        is_admin = user.role == User.Role.COMPANY_ADMIN or user.is_superuser
        is_manager = (
            user.role == User.Role.MANAGER and leave_request.employee.manager and leave_request.employee.manager == user
        )

        if not (is_admin or is_manager):
            return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)

        leave_request.status = "REJECTED"
        leave_request.rejection_reason = request.POST.get("rejection_reason", "")
        leave_request.approved_by = user  # Store acts as 'processed_by'
        leave_request.approved_at = timezone.now()
        leave_request.save()

        # Send Rejection Email asynchronously
        import threading

        def send_email_async():
            try:
                import logging

                from core.email_utils import send_leave_rejection_notification

                logger = logging.getLogger(__name__)
                if not send_leave_rejection_notification(leave_request):
                    logger.warning("Leave rejection email notification failed")
            except Exception as e:
                import logging

                logger = logging.getLogger(__name__)
                logger.error(f"Error sending rejection email: {e}")

        # Start email sending in background thread
        email_thread = threading.Thread(target=send_email_async)
        email_thread.daemon = True
        email_thread.start()

        messages.success(request, "Leave rejected. Notification will be sent.")

        return redirect(request.META.get("HTTP_REFERER", "leave_requests"))
    return redirect("leave_requests")


@login_required
def attendance_map(request, pk):
    try:
        attendance = Attendance.objects.get(pk=pk)
        employee = attendance.employee

        # Permission Check
        user = request.user
        is_admin = user.role == User.Role.COMPANY_ADMIN or user.is_superuser
        is_manager = user.role == User.Role.MANAGER and employee.manager and employee.manager == user
        is_self = user == employee.user

        if not (is_admin or is_manager or is_self):
            messages.error(request, "Permission Denied")
            return redirect("dashboard")

        # Get logs within the attendance period
        logs = LocationLog.objects.filter(employee=employee)

        if attendance.clock_in:
            logs = logs.filter(timestamp__gte=attendance.clock_in)

        if attendance.clock_out:
            logs = logs.filter(timestamp__lte=attendance.clock_out)

        logs = logs.order_by("timestamp")

        # Prepare Map Data
        map_locations = []

        # Determine timezone
        target_tz = pytz.timezone("Asia/Kolkata")  # Default

        if attendance.user_timezone:
            with contextlib.suppress(BaseException):
                target_tz = pytz.timezone(attendance.user_timezone)
        elif employee.location and hasattr(employee.location, "timezone") and employee.location.timezone:
            with contextlib.suppress(BaseException):
                target_tz = pytz.timezone(employee.location.timezone)

        # Activate timezone for the template
        timezone.activate(target_tz)

        # Helper for time formatting
        def format_time(dt):
            if not dt:
                return ""
            local_dt = timezone.localtime(dt, target_tz)
            return local_dt.strftime("%I:%M %p")

        # 1. Clock In
        if attendance.location_in:
            lat, lng = safe_parse_location(attendance.location_in)
            if lat:
                map_locations.append(
                    {
                        "lat": lat,
                        "lng": lng,
                        "title": f"Clock In: {format_time(attendance.clock_in)}",
                        "type": "in",
                    }
                )

        # 2. Logs
        for log in logs:
            map_locations.append(
                {
                    "lat": float(log.latitude),
                    "lng": float(log.longitude),
                    "title": f"Log: {format_time(log.timestamp)}",
                    "type": "log",
                }
            )

        # 3. Clock Out
        if attendance.location_out:
            lat, lng = safe_parse_location(attendance.location_out)
            if lat:
                map_locations.append(
                    {
                        "lat": lat,
                        "lng": lng,
                        "title": f"Clock Out: {format_time(attendance.clock_out)}",
                        "type": "out",
                    }
                )

        return render(
            request,
            "core/attendance_map.html",
            {
                "attendance": attendance,
                "logs": logs,
                "map_data_json": map_locations,
                "title": "Location History",
            },
        )
    except Attendance.DoesNotExist:
        messages.error(request, "Attendance record not found")
        return redirect("dashboard")


@login_required
def employee_detail(request, pk):
    try:
        employee = Employee.objects.get(pk=pk)

        # Permission Check (Company Admin or Manager of the employee)
        user = request.user
        is_admin = user.role == User.Role.COMPANY_ADMIN or user.is_superuser
        is_manager = user.role == User.Role.MANAGER and employee.manager and employee.manager == user

        if not (is_admin or is_manager):
            messages.error(request, "Permission denied")
            return redirect("employee_list")

        # Handle POST requests for shift assignment
        if request.method == "POST":
            action = request.POST.get("action")

            if action == "assign_shift":
                # Only allow Company Admins to assign shifts
                if not is_admin:
                    messages.error(request, "You don't have permission to assign shifts.")
                    return redirect("employee_detail", pk=pk)

                try:
                    from companies.models import ShiftSchedule

                    shift_id = request.POST.get("assigned_shift")
                    if shift_id:
                        shift = ShiftSchedule.objects.get(id=shift_id, company=user.company)
                        old_shift_name = employee.assigned_shift.name if employee.assigned_shift else "None"
                        employee.assigned_shift = shift
                        employee.save(update_fields=["assigned_shift"])

                        messages.success(
                            request,
                            f"Shift assignment updated from '{old_shift_name}' to '{shift.name}' for {employee.user.get_full_name()}.",
                        )
                    else:
                        # Remove shift assignment
                        old_shift_name = employee.assigned_shift.name if employee.assigned_shift else "None"
                        employee.assigned_shift = None
                        employee.save(update_fields=["assigned_shift"])

                        messages.success(
                            request,
                            f"Shift assignment removed from {employee.user.get_full_name()}. Previous shift: '{old_shift_name}'.",
                        )

                    return redirect("employee_detail", pk=pk)

                except ShiftSchedule.DoesNotExist:
                    messages.error(request, "Selected shift not found.")
                    return redirect("employee_detail", pk=pk)
                except Exception as e:
                    messages.error(request, f"Error assigning shift: {str(e)}")
                    return redirect("employee_detail", pk=pk)

        today = timezone.localdate()

        # Date Filter (Default to current month)
        import calendar

        month = int(request.GET.get("month", today.month))
        year = int(request.GET.get("year", today.year))

        # Get start and end of month
        num_days = calendar.monthrange(year, month)[1]
        start_date = timezone.datetime(year, month, 1).date()
        end_date = timezone.datetime(year, month, num_days).date()

        # Fetch Attendance for the period
        attendance_records = {
            att.date: att for att in Attendance.objects.filter(employee=employee, date__range=[start_date, end_date])
        }

        # Fetch Approved Leaves
        leaves = LeaveRequest.objects.filter(
            employee=employee, status="APPROVED", start_date__lte=end_date, end_date__gte=start_date
        )
        leave_dates = {}
        for leave in leaves:
            curr = max(leave.start_date, start_date)
            while curr <= min(leave.end_date, end_date):
                leave_dates[curr] = "LEAVE"
                curr += timedelta(days=1)

        # Fetch Holidays
        from django.db.models import Q

        from companies.models import Holiday

        holiday_q = Q(location__isnull=True)
        if employee.location:
            holiday_q |= Q(location=employee.location)

        holidays = Holiday.objects.filter(
            company=employee.company, date__range=[start_date, end_date], is_active=True
        ).filter(holiday_q)
        holiday_dates = {h.date: h.name for h in holidays}

        # Build Full Attendance History
        full_history = []
        curr_date = min(end_date, today)
        while curr_date >= start_date:
            if curr_date in attendance_records:
                full_history.append(attendance_records[curr_date])
            else:
                # Determine status for missing record
                if employee.is_week_off(curr_date):
                    status = "WEEKLY_OFF"
                elif curr_date in holiday_dates:
                    status = "HOLIDAY"
                elif curr_date in leave_dates:
                    status = "LEAVE"
                elif curr_date == today:
                    status = "NOT_LOGGED_IN"
                elif curr_date > today:
                    status = "UPCOMING"
                else:
                    status = "MISSED"

                # Mock attendance object for template
                full_history.append(
                    {
                        "date": curr_date,
                        "status": status,
                        "id": None,
                        "effective_hours": "-",
                        "clock_in": None,
                        "clock_out": None,
                        "holiday_name": holiday_dates.get(curr_date),
                    }
                )
            curr_date -= timedelta(days=1)

        # QuerySet for map logic and stats (original records)
        attendance_qs = Attendance.objects.filter(employee=employee, date__range=[start_date, end_date]).order_by(
            "-date"
        )

        # Calculate Stats (using the full history)
        total_days = len(full_history)
        present = 0
        wfh = 0
        leave = 0
        absent = 0

        for item in full_history:
            status = item.status if isinstance(item, Attendance) else item.get("status")

            if status == "PRESENT":
                present += 1
            elif status == "WFH":
                wfh += 1
            elif status == "LEAVE":
                leave += 1
            elif status in ["ABSENT", "MISSED"]:
                absent += 1

        # Location Data for Map (Today's path or last active day)
        # We try to get today's attendance first
        map_date = today
        map_attendance = attendance_qs.filter(date=today).first()

        # If no attendance today, grab the last one with location data
        if not map_attendance or (not map_attendance.location_in and not map_attendance.location_out):
            # Find last record with location
            last_loc_att = attendance_qs.exclude(location_in__isnull=True).first()
            if last_loc_att:
                map_attendance = last_loc_att
                map_date = last_loc_att.date

        map_data = []
        if map_attendance:
            # Determine timezone
            tz_name = normalize_timezone(employee.location.timezone if employee.location else None)
            local_tz = pytz.timezone(tz_name)

            def format_local_time(dt):
                if not dt:
                    return ""
                return timezone.localtime(dt, local_tz).strftime("%I:%M %p")

            # Clock In Marker
            if map_attendance.location_in:
                lat, lng = safe_parse_location(map_attendance.location_in)
                if lat:
                    map_data.append(
                        {
                            "lat": lat,
                            "lng": lng,
                            "title": f"Clock In ({format_local_time(map_attendance.clock_in)})",
                            "type": "start",
                            "time_display": format_local_time(map_attendance.clock_in),
                        }
                    )

            # Location Logs (Hourly Tracking)
            day_start = timezone.make_aware(timezone.datetime.combine(map_date, timezone.datetime.min.time()))
            day_end = timezone.make_aware(timezone.datetime.combine(map_date, timezone.datetime.max.time()))

            logs = LocationLog.objects.filter(employee=employee, timestamp__range=[day_start, day_end]).order_by(
                "timestamp"
            )

            for log in logs:
                try:
                    lat = float(log.latitude)
                    lng = float(log.longitude)
                except (TypeError, ValueError):
                    logger.warning(
                        "Skipping invalid location log coordinates",
                        employee_id=employee.id,
                        log_id=log.id,
                        latitude=log.latitude,
                        longitude=log.longitude,
                    )
                    continue

                title = "Location Punch" if log.log_type == "HOURLY" else "Movement Log"
                map_data.append(
                    {
                        "lat": lat,
                        "lng": lng,
                        "title": f"{title} ({format_local_time(log.timestamp)})",
                        "type": "log",
                        "log_type": log.log_type,
                        "time_display": format_local_time(log.timestamp),
                    }
                )

            # Clock Out Marker
            if map_attendance.location_out:
                lat, lng = safe_parse_location(map_attendance.location_out)
                if lat:
                    map_data.append(
                        {
                            "lat": lat,
                            "lng": lng,
                            "title": f"Clock Out ({format_local_time(map_attendance.clock_out)})",
                            "type": "end",
                            "time_display": format_local_time(map_attendance.clock_out),
                        }
                    )

        # Calculate Probation Date (3 months from joining)
        probation_date = None
        probation_status = None
        if employee.date_of_joining:
            probation_date = employee.get_probation_end_date()
            probation_status = employee.get_probation_status()

        # Get available shifts for the company (for shift assignment)
        available_shifts = []
        if is_admin:
            from companies.models import ShiftSchedule

            available_shifts = ShiftSchedule.objects.filter(company=user.company, is_active=True).order_by("name")

        context = {
            "employee": employee,
            "attendance_history": full_history,
            "current_month": start_date.strftime("%B"),
            "current_year": year,
            "stats": {
                "total": total_days,
                "present": present,
                "wfh": wfh,
                "leave": leave,
                "absent": absent,
            },
            "map_data": json.dumps(map_data),
            "map_date": map_date,
            "probation_date": probation_date,
            "probation_status": probation_status,
            "available_shifts": available_shifts,
            "is_admin": is_admin,
        }
        return render(request, "employees/employee_detail.html", context)

    except Employee.DoesNotExist:
        messages.error(request, "Employee not found")
        return redirect("employee_list")


@csrf_exempt
@login_required
def employee_exit_action(request, pk):
    """
    Handle employee exit actions (Resignation, Absconding, Termination)
    - Resignation: Creates pending exit initiative, sends email to admin/manager
    - Absconding/Termination: Calculates last working day, disables login immediately
    Only accessible by Company Admin or Super Admin
    """
    from .models import ExitInitiative

    # Permission check
    if request.user.role not in [User.Role.COMPANY_ADMIN, User.Role.SUPERADMIN]:
        return JsonResponse(
            {
                "status": "error",
                "message": "Permission denied. Only admins can perform exit actions.",
            },
            status=403,
        )

    try:
        employee = Employee.objects.get(pk=pk)

        # Prevent exiting already exited employees
        if not employee.is_active:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "This employee has already been exited from the organization.",
                },
                status=400,
            )

        # Get form data
        exit_type = request.POST.get("exit_type")  # RESIGNATION, ABSCONDED, TERMINATED
        submission_date_str = request.POST.get("submission_date")
        exit_note = request.POST.get("exit_note", "").strip()
        notice_period_days = request.POST.get("notice_period_days", "").strip()

        # Validation
        if not exit_type or exit_type not in ["RESIGNATION", "ABSCONDED", "TERMINATED"]:
            return JsonResponse(
                {"status": "error", "message": "Invalid exit type selected."},
                status=400,
            )

        if not submission_date_str:
            return JsonResponse(
                {"status": "error", "message": "Submission date is required."},
                status=400,
            )

        if not exit_note:
            return JsonResponse(
                {"status": "error", "message": "Exit reason/note is required."},
                status=400,
            )

        # Parse and validate date
        try:
            submission_date = timezone.datetime.strptime(submission_date_str, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse({"status": "error", "message": "Invalid date format."}, status=400)

        # Handle different exit types
        if exit_type == "RESIGNATION":
            # Create ExitInitiative with PENDING status
            exit_initiative = ExitInitiative.objects.create(
                employee=employee,
                exit_type=exit_type,
                submission_date=submission_date,
                exit_note=exit_note,
                status="PENDING",
            )

            # Update employee status but keep them active
            # Update employee status but keep them active
            employee.employment_status = exit_type
            employee.exit_note = exit_note
            # Don't set exit_date yet - will be set upon approval
            # Don't disable login - employee can work until last working day
            employee.save()

            # --- Email Notification ---
            try:
                from django.core.mail import send_mail

                # 1. Recipients
                recipients = []

                # Reporting Manager
                if employee.manager and employee.manager.email:
                    recipients.append(employee.manager.email)

                # HR/Admins
                company_admins = User.objects.filter(company=employee.company, role="COMPANY_ADMIN").values_list(
                    "email", flat=True
                )
                recipients.extend(list(company_admins))

                # Company HR Email
                if employee.company.hr_email:
                    recipients.append(employee.company.hr_email)

                # Deduplicate
                recipients = list(set(filter(None, recipients)))

                if recipients:
                    subject = f"Resignation Submitted: {employee.user.get_full_name()} ({employee.designation})"
                    message = f"""
                    Dear Team,

                    This is to inform you that {employee.user.get_full_name()} ({employee.designation}) has submitted their resignation on {submission_date.strftime("%d %b %Y")}.

                    Reason:
                    {exit_note}

                    Current Status: Pending Approval

                    Please login to the HRMS portal to review and take necessary action.

                    Regards,
                    HRMS System
                    """
                    send_mail(
                        subject,
                        message,
                        settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL,
                        recipients,
                        fail_silently=True,
                    )

            except Exception as e:
                logger.error(f"Error in resignation notification: {e}")

            messages.success(
                request,
                f"Resignation request from {employee.user.get_full_name()} has been submitted for approval.",
            )

            return JsonResponse(
                {
                    "status": "success",
                    "message": "Resignation request submitted successfully. Awaiting approval.",
                    "redirect_url": reverse("employee_list"),
                }
            )

        elif exit_type in ["ABSCONDED", "TERMINATED"]:
            # Validate notice period
            if not notice_period_days:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "Notice period (in days) is required for absconding/termination.",
                    },
                    status=400,
                )

            try:
                notice_period = int(notice_period_days)
                if notice_period < 0:
                    raise ValueError("Notice period cannot be negative")
            except ValueError:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "Invalid notice period. Please enter a valid number of days.",
                    },
                    status=400,
                )

            # Calculate last working day
            last_working_day = submission_date + timedelta(days=notice_period)

            # Create ExitInitiative with PENDING status (requires approval)
            exit_initiative = ExitInitiative.objects.create(
                employee=employee,
                exit_type=exit_type,
                submission_date=submission_date,
                exit_note=exit_note,
                notice_period_days=notice_period,
                last_working_day=last_working_day,
                status="PENDING",  # Changed to PENDING
            )

            # Update employee status but keep them active until approval
            employee.employment_status = exit_type
            employee.exit_note = exit_note
            # Don't set exit_date yet - will be set upon approval
            # Don't disable login - employee can work until approved
            employee.save()

            # --- Email Notification to HR/Admin/Manager ---
            try:
                from django.core.mail import send_mail

                # 1. Recipients
                recipients = []

                # Reporting Manager
                if employee.manager and employee.manager.email:
                    recipients.append(employee.manager.email)

                # HR/Admins
                company_admins = User.objects.filter(company=employee.company, role="COMPANY_ADMIN").values_list(
                    "email", flat=True
                )
                recipients.extend(list(company_admins))

                # Company HR Email
                if employee.company.hr_email:
                    recipients.append(employee.company.hr_email)

                # Deduplicate
                recipients = list(set(filter(None, recipients)))

                if recipients:
                    subject = f"{exit_initiative.get_exit_type_display()} Submitted: {employee.user.get_full_name()} ({employee.designation})"
                    message = f"""
Dear Team,

This is to inform you that a {exit_initiative.get_exit_type_display().lower()} request has been submitted for {employee.user.get_full_name()} ({employee.designation}) on {submission_date.strftime("%d %b %Y")}.

Reason:
{exit_note}

Notice Period: {notice_period} days
Proposed Last Working Day: {last_working_day.strftime("%d %b %Y")}

Current Status: Pending Approval

Please login to the HRMS portal to review and take necessary action.

Regards,
HRMS System
                    """
                    send_mail(
                        subject,
                        message,
                        settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL,
                        recipients,
                        fail_silently=True,
                    )

            except Exception as e:
                logger.error(f"Error in {exit_type.lower()} notification: {e}")

            messages.success(
                request,
                f"{exit_initiative.get_exit_type_display()} request for {employee.user.get_full_name()} has been submitted for approval.",
            )

            return JsonResponse(
                {
                    "status": "success",
                    "message": f"{exit_initiative.get_exit_type_display()} request submitted successfully. Awaiting approval.",
                    "redirect_url": reverse("employee_list"),
                }
            )

    except Employee.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Employee not found."}, status=404)
    except Exception as e:
        import traceback

        traceback.print_exc()
        return JsonResponse({"status": "error", "message": f"An error occurred: {str(e)}"}, status=500)


@login_required
def approve_exit_initiative(request, pk):
    """
    Approve an exit initiative (resignation, termination, or absconding)
    Accessible by Admin and Manager
    """
    from datetime import datetime, timedelta

    from dateutil.relativedelta import relativedelta

    from .models import ExitInitiative

    try:
        exit_initiative = ExitInitiative.objects.get(pk=pk)
        employee = exit_initiative.employee

        # Permission check: Admin or Manager
        user = request.user
        is_admin = user.role in [User.Role.COMPANY_ADMIN, User.Role.SUPERADMIN]
        is_manager = user.role == User.Role.MANAGER and employee.manager and employee.manager == user

        if not (is_admin or is_manager):
            messages.error(request, "Permission denied. Only Admin or Manager can approve exit initiatives.")
            return redirect("exit_initiatives_list")

        # Check if already processed
        if exit_initiative.status != "PENDING":
            messages.warning(request, f"This exit initiative has already been {exit_initiative.status.lower()}.")
            return redirect("exit_initiatives_list")

        # Get last working day from POST request
        if request.method == "POST":
            last_working_day_str = request.POST.get("last_working_day")
            if not last_working_day_str:
                messages.error(request, "Last working day is required.")
                return redirect("exit_initiatives_list")

            try:
                last_working_day = datetime.strptime(last_working_day_str, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format for last working day.")
                return redirect("exit_initiatives_list")

            # Validate that last working day is not before submission date
            if last_working_day < exit_initiative.submission_date:
                messages.error(request, "Last working day cannot be before submission date.")
                return redirect("exit_initiatives_list")
        else:
            # For backward compatibility with GET requests (if any direct links exist)
            # Calculate last working day based on exit type
            if exit_initiative.exit_type == "RESIGNATION":
                # 2 months from approval date
                last_working_day = timezone.now().date() + relativedelta(months=2)
            elif exit_initiative.exit_type in ["ABSCONDED", "TERMINATED"]:
                # Already calculated during submission or use submission date
                if exit_initiative.last_working_day:
                    last_working_day = exit_initiative.last_working_day
                else:
                    last_working_day = exit_initiative.submission_date + timedelta(
                        days=exit_initiative.notice_period_days or 0
                    )
            else:
                last_working_day = timezone.now().date()

        # Approve the exit initiative
        exit_initiative.status = "APPROVED"
        exit_initiative.approved_by = user
        exit_initiative.approved_at = timezone.now()
        exit_initiative.last_working_day = last_working_day
        exit_initiative.save()

        # Update employee record
        employee.exit_date = last_working_day
        employee.exit_note = exit_initiative.exit_note

        # Check if exit is effective today/past or future
        today = timezone.localdate()
        is_immediate = last_working_day <= today

        if is_immediate:
            # Immediate Exit - change to Ex-Employee type and disable employee
            employee.employment_status = "EX_EMPLOYEE"
            employee.is_active = False
            employee.save()

            # Disable user login
            emp_user = employee.user
            emp_user.is_active = False
            emp_user.save()

            status_msg = f"Exit initiative approved. {employee.user.get_full_name()}'s account has been changed to Ex-Employee type and access has been disabled immediately."
        else:
            # Future Exit - keep active until last working day
            # Keep current employment status until last working day
            employee.employment_status = exit_initiative.exit_type
            employee.is_active = True
            employee.save()

            # Ensure user is active
            emp_user = employee.user
            emp_user.is_active = True
            emp_user.save()

            status_msg = f"Exit initiative approved. {employee.user.get_full_name()}'s last working day is {last_working_day.strftime('%d %b %Y')}. Account will be changed to Ex-Employee type on that date."

        # Send email notification to employee
        try:
            from django.core.mail import EmailMultiAlternatives

            # Prepare recipient list (official email + personal email if available)
            recipients = [employee.user.email]
            if employee.personal_email:
                recipients.append(employee.personal_email)

            subject = f"Exit Initiative Approved - {exit_initiative.get_exit_type_display()}"

            # Plain text version
            text_message = f"""
Dear {employee.user.get_full_name()},

Your {exit_initiative.get_exit_type_display().lower()} request has been approved.

Submission Date: {exit_initiative.submission_date.strftime("%d %b %Y")}
Last Working Day: {last_working_day.strftime("%d %b %Y")}
Approved By: {user.get_full_name()}

Reason:
{exit_initiative.exit_note}

On your last working day, your account will be changed to Ex-Employee type and access will be disabled.

Please complete all exit formalities before your last working day.

Regards,
{employee.company.name} HR Team
            """

            # HTML version
            html_message = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f7fa;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f7fa; padding: 40px 20px;">
        <tr>
            <td align="center">
                <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); overflow: hidden;">
                    <!-- Header -->
                    <tr>
                        <td style="background: linear-gradient(135deg, #22c55e 0%, #16a34a 100%); padding: 40px 30px; text-align: center;">
                            <h1 style="margin: 0; color: #ffffff; font-size: 28px; font-weight: 600; letter-spacing: -0.5px;">✓ Exit Request Approved</h1>
                            <p style="margin: 10px 0 0 0; color: #dcfce7; font-size: 14px;">{exit_initiative.get_exit_type_display()}</p>
                        </td>
                    </tr>

                    <!-- Body -->
                    <tr>
                        <td style="padding: 40px 30px;">
                            <p style="margin: 0 0 24px 0; color: #1f2937; font-size: 16px; line-height: 1.6;">
                                Dear <strong>{employee.user.get_full_name()}</strong>,
                            </p>

                            <p style="margin: 0 0 30px 0; color: #4b5563; font-size: 15px; line-height: 1.6;">
                                Your <strong>{exit_initiative.get_exit_type_display().lower()}</strong> request has been <span style="color: #16a34a; font-weight: 600;">approved</span>.
                            </p>

                            <!-- Details Box -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f9fafb; border-radius: 8px; border-left: 4px solid #22c55e; margin-bottom: 30px;">
                                <tr>
                                    <td style="padding: 24px;">
                                        <table width="100%" cellpadding="8" cellspacing="0">
                                            <tr>
                                                <td style="color: #6b7280; font-size: 14px; padding: 8px 0;">📅 Submission Date:</td>
                                                <td style="color: #1f2937; font-size: 14px; font-weight: 600; text-align: right; padding: 8px 0;">{exit_initiative.submission_date.strftime("%d %b %Y")}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #6b7280; font-size: 14px; padding: 8px 0;">🗓️ Last Working Day:</td>
                                                <td style="color: #dc2626; font-size: 14px; font-weight: 600; text-align: right; padding: 8px 0;">{last_working_day.strftime("%d %b %Y")}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #6b7280; font-size: 14px; padding: 8px 0;">✅ Approved By:</td>
                                                <td style="color: #1f2937; font-size: 14px; font-weight: 600; text-align: right; padding: 8px 0;">{user.get_full_name()}</td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>

                            <!-- Reason Section -->
                            <div style="margin-bottom: 30px;">
                                <p style="margin: 0 0 12px 0; color: #6b7280; font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Reason for Exit:</p>
                                <div style="background-color: #fef3c7; border-left: 4px solid #f59e0b; padding: 16px; border-radius: 6px;">
                                    <p style="margin: 0; color: #92400e; font-size: 14px; line-height: 1.6; font-style: italic;">{exit_initiative.exit_note}</p>
                                </div>
                            </div>

                            <!-- Important Notice -->
                            <div style="background-color: #fef2f2; border: 1px solid #fecaca; border-radius: 8px; padding: 20px; margin-bottom: 30px;">
                                <p style="margin: 0 0 8px 0; color: #991b1b; font-size: 14px; font-weight: 600;">⚠️ Important Notice:</p>
                                <p style="margin: 0; color: #7f1d1d; font-size: 13px; line-height: 1.6;">
                                    On your last working day, your account will be changed to <strong>Ex-Employee</strong> type and access will be disabled.
                                </p>
                            </div>

                            <!-- Action Items -->
                            <div style="background-color: #eff6ff; border-radius: 8px; padding: 20px; margin-bottom: 30px;">
                                <p style="margin: 0 0 12px 0; color: #1e40af; font-size: 14px; font-weight: 600;">📋 Next Steps:</p>
                                <ul style="margin: 0; padding-left: 20px; color: #1e3a8a; font-size: 13px; line-height: 1.8;">
                                    <li>Complete all pending tasks and handover documentation</li>
                                    <li>Return company assets (laptop, ID card, access cards, etc.)</li>
                                    <li>Complete exit interview with HR</li>
                                    <li>Settle any pending dues or reimbursements</li>
                                    <li>Update your personal contact information for future correspondence</li>
                                </ul>
                            </div>

                            <p style="margin: 0 0 8px 0; color: #4b5563; font-size: 14px; line-height: 1.6;">
                                Please complete all exit formalities before your last working day.
                            </p>

                            <p style="margin: 24px 0 0 0; color: #6b7280; font-size: 13px; line-height: 1.6;">
                                If you have any questions, please contact the HR department.
                            </p>
                        </td>
                    </tr>

                    <!-- Footer -->
                    <tr>
                        <td style="background-color: #f9fafb; padding: 30px; text-align: center; border-top: 1px solid #e5e7eb;">
                            <p style="margin: 0 0 8px 0; color: #1f2937; font-size: 14px; font-weight: 600;">
                                Best Regards,
                            </p>
                            <p style="margin: 0; color: #6b7280; font-size: 14px;">
                                {employee.company.name} HR Team
                            </p>
                            <p style="margin: 16px 0 0 0; color: #9ca3af; font-size: 12px;">
                                This is an automated notification from the HRMS system.
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
            """

            # Create email with both plain text and HTML
            email = EmailMultiAlternatives(
                subject,
                text_message,
                settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL,
                recipients,
            )
            email.attach_alternative(html_message, "text/html")
            email.send(fail_silently=True)
        except Exception as e:
            logger.error(f"Error sending approval email: {e}")

        messages.success(request, status_msg)
        return redirect("exit_initiatives_list")

    except ExitInitiative.DoesNotExist:
        messages.error(request, "Exit initiative not found.")
        return redirect("exit_initiatives_list")
    except Exception as e:
        import traceback

        traceback.print_exc()
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("exit_initiatives_list")


@login_required
def reject_exit_initiative(request, pk):
    """
    Reject an exit initiative
    Accessible by Admin and Manager
    """
    from .models import ExitInitiative

    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("exit_initiatives_list")

    try:
        exit_initiative = ExitInitiative.objects.get(pk=pk)
        employee = exit_initiative.employee

        # Permission check: Admin or Manager
        user = request.user
        is_admin = user.role in [User.Role.COMPANY_ADMIN, User.Role.SUPERADMIN]
        is_manager = user.role == User.Role.MANAGER and employee.manager and employee.manager == user

        if not (is_admin or is_manager):
            messages.error(request, "Permission denied. Only Admin or Manager can reject exit initiatives.")
            return redirect("exit_initiatives_list")

        # Check if already processed
        if exit_initiative.status != "PENDING":
            messages.warning(request, f"This exit initiative has already been {exit_initiative.status.lower()}.")
            return redirect("exit_initiatives_list")

        # Get rejection reason
        rejection_reason = request.POST.get("rejection_reason", "").strip()
        if not rejection_reason:
            messages.error(request, "Please provide a reason for rejection.")
            return redirect("exit_initiatives_list")

        # Reject the exit initiative
        exit_initiative.status = "REJECTED"
        exit_initiative.rejection_reason = rejection_reason
        exit_initiative.approved_by = user  # Store who rejected it
        exit_initiative.approved_at = timezone.now()
        exit_initiative.save()

        # Reset employee status
        employee.employment_status = "ACTIVE"
        employee.exit_date = None
        employee.exit_note = ""
        employee.is_active = True
        employee.save()

        # Ensure user is active
        emp_user = employee.user
        emp_user.is_active = True
        emp_user.save()

        # Send email notification to employee
        try:
            from django.core.mail import EmailMultiAlternatives

            # Prepare recipient list (official email + personal email if available)
            recipients = [employee.user.email]
            if employee.personal_email:
                recipients.append(employee.personal_email)

            subject = f"Exit Initiative Rejected - {exit_initiative.get_exit_type_display()}"

            # Plain text version
            text_message = f"""
Dear {employee.user.get_full_name()},

Your {exit_initiative.get_exit_type_display().lower()} request has been rejected.

Submission Date: {exit_initiative.submission_date.strftime("%d %b %Y")}
Rejected By: {user.get_full_name()}

Reason for Rejection:
{rejection_reason}

If you have any questions, please contact HR.

Regards,
{employee.company.name} HR Team
            """

            # HTML version
            html_message = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f7fa;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f7fa; padding: 40px 20px;">
        <tr>
            <td align="center">
                <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); overflow: hidden;">
                    <!-- Header -->
                    <tr>
                        <td style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); padding: 40px 30px; text-align: center;">
                            <h1 style="margin: 0; color: #ffffff; font-size: 28px; font-weight: 600; letter-spacing: -0.5px;">✕ Exit Request Rejected</h1>
                            <p style="margin: 10px 0 0 0; color: #fecaca; font-size: 14px;">{exit_initiative.get_exit_type_display()}</p>
                        </td>
                    </tr>

                    <!-- Body -->
                    <tr>
                        <td style="padding: 40px 30px;">
                            <p style="margin: 0 0 24px 0; color: #1f2937; font-size: 16px; line-height: 1.6;">
                                Dear <strong>{employee.user.get_full_name()}</strong>,
                            </p>

                            <p style="margin: 0 0 30px 0; color: #4b5563; font-size: 15px; line-height: 1.6;">
                                Your <strong>{exit_initiative.get_exit_type_display().lower()}</strong> request has been <span style="color: #dc2626; font-weight: 600;">rejected</span>.
                            </p>

                            <!-- Details Box -->
                            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f9fafb; border-radius: 8px; border-left: 4px solid #ef4444; margin-bottom: 30px;">
                                <tr>
                                    <td style="padding: 24px;">
                                        <table width="100%" cellpadding="8" cellspacing="0">
                                            <tr>
                                                <td style="color: #6b7280; font-size: 14px; padding: 8px 0;">📅 Submission Date:</td>
                                                <td style="color: #1f2937; font-size: 14px; font-weight: 600; text-align: right; padding: 8px 0;">{exit_initiative.submission_date.strftime("%d %b %Y")}</td>
                                            </tr>
                                            <tr>
                                                <td style="color: #6b7280; font-size: 14px; padding: 8px 0;">❌ Rejected By:</td>
                                                <td style="color: #1f2937; font-size: 14px; font-weight: 600; text-align: right; padding: 8px 0;">{user.get_full_name()}</td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>

                            <!-- Rejection Reason Section -->
                            <div style="margin-bottom: 30px;">
                                <p style="margin: 0 0 12px 0; color: #6b7280; font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Reason for Rejection:</p>
                                <div style="background-color: #fef2f2; border-left: 4px solid #ef4444; padding: 16px; border-radius: 6px;">
                                    <p style="margin: 0; color: #7f1d1d; font-size: 14px; line-height: 1.6;">{rejection_reason}</p>
                                </div>
                            </div>

                            <!-- Status Notice -->
                            <div style="background-color: #ecfdf5; border: 1px solid #a7f3d0; border-radius: 8px; padding: 20px; margin-bottom: 30px;">
                                <p style="margin: 0 0 8px 0; color: #065f46; font-size: 14px; font-weight: 600;">✓ Your Employment Status:</p>
                                <p style="margin: 0; color: #047857; font-size: 13px; line-height: 1.6;">
                                    Your employment status remains <strong>Active</strong>. You can continue working as usual.
                                </p>
                            </div>

                            <!-- Next Steps -->
                            <div style="background-color: #eff6ff; border-radius: 8px; padding: 20px; margin-bottom: 30px;">
                                <p style="margin: 0 0 12px 0; color: #1e40af; font-size: 14px; font-weight: 600;">💬 Need to Discuss?</p>
                                <p style="margin: 0; color: #1e3a8a; font-size: 13px; line-height: 1.6;">
                                    If you have any questions or concerns regarding this decision, please feel free to:
                                </p>
                                <ul style="margin: 12px 0 0 0; padding-left: 20px; color: #1e3a8a; font-size: 13px; line-height: 1.8;">
                                    <li>Schedule a meeting with your reporting manager</li>
                                    <li>Contact the HR department for clarification</li>
                                    <li>Submit a new request if circumstances change</li>
                                </ul>
                            </div>

                            <p style="margin: 0; color: #6b7280; font-size: 13px; line-height: 1.6;">
                                We appreciate your understanding and continued contribution to the organization.
                            </p>
                        </td>
                    </tr>

                    <!-- Footer -->
                    <tr>
                        <td style="background-color: #f9fafb; padding: 30px; text-align: center; border-top: 1px solid #e5e7eb;">
                            <p style="margin: 0 0 8px 0; color: #1f2937; font-size: 14px; font-weight: 600;">
                                Best Regards,
                            </p>
                            <p style="margin: 0; color: #6b7280; font-size: 14px;">
                                {employee.company.name} HR Team
                            </p>
                            <p style="margin: 16px 0 0 0; color: #9ca3af; font-size: 12px;">
                                This is an automated notification from the HRMS system.
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
            """

            # Create email with both plain text and HTML
            email = EmailMultiAlternatives(
                subject,
                text_message,
                settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL,
                recipients,
            )
            email.attach_alternative(html_message, "text/html")
            email.send(fail_silently=True)
        except Exception as e:
            logger.error(f"Error sending rejection email: {e}")

        messages.success(request, f"Exit initiative rejected. {employee.user.get_full_name()} has been notified.")
        return redirect("exit_initiatives_list")

    except ExitInitiative.DoesNotExist:
        messages.error(request, "Exit initiative not found.")
        return redirect("exit_initiatives_list")
    except Exception as e:
        import traceback

        traceback.print_exc()
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("exit_initiatives_list")


@login_required
def exit_initiatives_list(request):
    """
    Display all pending exit initiatives for approval
    Accessible by Admin and Managers (for their team members)
    """
    from .models import ExitInitiative

    user = request.user
    is_admin = user.role in [User.Role.COMPANY_ADMIN, User.Role.SUPERADMIN]

    # Get exit initiatives based on role
    if is_admin:
        # Admin can see all exit initiatives in their company
        exit_initiatives = (
            ExitInitiative.objects.filter(employee__company=user.company)
            .select_related("employee", "employee__user", "approved_by")
            .order_by("-created_at")
        )
    elif user.role == User.Role.MANAGER:
        # Managers can only see their team members' exit initiatives
        try:
            exit_initiatives = (
                ExitInitiative.objects.filter(employee__manager=user)
                .select_related("employee", "employee__user", "approved_by")
                .order_by("-created_at")
            )
        except Exception:
            exit_initiatives = ExitInitiative.objects.none()
    else:
        # Regular employees cannot access this page
        messages.error(request, "Permission denied. Only Admin or Managers can access this page.")
        return redirect("dashboard")

    # Filter by status
    status_filter = request.GET.get("status", "pending")
    if status_filter == "pending":
        exit_initiatives = exit_initiatives.filter(status="PENDING")
    elif status_filter == "approved":
        exit_initiatives = exit_initiatives.filter(status="APPROVED")
    elif status_filter == "rejected":
        exit_initiatives = exit_initiatives.filter(status="REJECTED")
    # 'all' shows everything

    context = {
        "exit_initiatives": exit_initiatives,
        "status_filter": status_filter,
        "pending_count": ExitInitiative.objects.filter(employee__company=user.company, status="PENDING").count()
        if is_admin
        else ExitInitiative.objects.filter(employee__manager=user, status="PENDING").count()
        if user.role == User.Role.MANAGER
        else 0,
    }

    return render(request, "employees/exit_initiatives_list.html", context)


@csrf_exempt
@login_required
def get_attendance_map_data(request, pk):
    try:
        attendance = Attendance.objects.get(pk=pk)

        # Permission check
        is_admin = request.user.role == User.Role.COMPANY_ADMIN or request.user.is_superuser

        # Safe manager check
        is_manager = False
        if (
            request.user.role == User.Role.MANAGER
            and hasattr(request.user, "employee_profile")
            and attendance.employee.manager
        ):
            # Compare manager's user object with request user
            is_manager = attendance.employee.manager == request.user

        is_self = attendance.employee.user == request.user

        if not (is_admin or is_manager or is_self):
            return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)

        map_locations = []

        # 1. Clock In
        if attendance.location_in:
            lat, lng = safe_parse_location(attendance.location_in)
            if lat:
                map_locations.append(
                    {
                        "lat": lat,
                        "lng": lng,
                        "title": f"Clock In: {attendance.clock_in.strftime('%I:%M %p')}",
                        "time_display": attendance.clock_in.strftime("%I:%M %p"),
                        "type": "start",
                        "type_display": "Clock In",
                    }
                )

        # 2. Logs
        if attendance.clock_in:
            # Determine end time for log query (clock_out or now)
            end_time = attendance.clock_out if attendance.clock_out else timezone.now()

            logs = LocationLog.objects.filter(
                employee=attendance.employee,
                timestamp__gte=attendance.clock_in,
                timestamp__lte=end_time,
            ).order_by("timestamp")

            for log in logs:
                title = "Location Punch" if log.log_type == "HOURLY" else "Movement Log"
                map_locations.append(
                    {
                        "lat": float(log.latitude),
                        "lng": float(log.longitude),
                        "title": f"{title}: {log.timestamp.strftime('%I:%M %p')}",
                        "time_display": log.timestamp.strftime("%I:%M %p"),
                        "type": "log",
                        "type_display": title,
                    }
                )

        # 3. Clock Out
        if attendance.location_out:
            lat, lng = safe_parse_location(attendance.location_out)
            if lat:
                map_locations.append(
                    {
                        "lat": lat,
                        "lng": lng,
                        "title": f"Clock Out: {attendance.clock_out.strftime('%I:%M %p')}",
                        "time_display": attendance.clock_out.strftime("%I:%M %p"),
                        "type": "end",
                        "type_display": "Clock Out",
                    }
                )

        # Sort by type to ensure logical connecting line (Start -> Logs -> End) is somewhat respected,
        # OR better: if we have timestamps for all events, sort by them.
        # But `map_locations` does not store raw datetime objects right now.
        # Let's rely on the current append order: Start -> Logs -> End. This assumes Logs are between Start and End.
        # This is generally true.

        return JsonResponse({"status": "success", "data": map_locations})

    except Attendance.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Attendance not found"}, status=404)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


class BulkEmployeeImportView(LoginRequiredMixin, CompanyAdminRequiredMixin, FormView):
    template_name = "employees/bulk_import.html"
    form_class = EmployeeBulkImportForm
    success_url = reverse_lazy("employee_list")

    def form_valid(self, form):
        import pandas as pd
        from django.contrib import messages

        from .utils import send_activation_email

        file = form.cleaned_data["import_file"]
        try:
            # Read Excel
            try:
                df = pd.read_excel(file, engine="openpyxl")
            except Exception as e:
                messages.error(
                    self.request, f"Error reading Excel file: {str(e)}. Please ensure the file is a valid .xlsx file."
                )
                return self.form_invalid(form)

            # Check if file is empty
            if df.empty:
                messages.error(self.request, "The uploaded file is empty. Please add employee data and try again.")
                return self.form_invalid(form)

            # Normalize columns
            df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

            # DOJ and DOB are now mandatory
            required_columns = [
                "first_name",
                "last_name",
                "email",
                "designation",
                "department",
                "date_of_joining",
                "date_of_birth",
                "reporting_manager_name",
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(
                    self.request,
                    f"Missing required columns: {', '.join([c.replace('_', ' ').title() for c in missing_columns])}. Please download the sample file for reference.",
                )
                return self.form_invalid(form)

            success_count = 0
            skipped_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    with transaction.atomic():
                        # 1. Basic Data
                        email = str(row.get("email", "")).strip().lower()
                        if not email or email == "nan":
                            raise ValueError("Email is required")

                        # Validate email format
                        import re

                        email_pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
                        if not re.match(email_pattern, email):
                            raise ValueError(f"Invalid email format: {email}")

                        # Check if user already exists
                        user = User.objects.filter(email=email).first()
                        if user:
                            # Check if the user already has an employee profile (is a real employee)
                            if hasattr(user, "employee_profile"):
                                skipped_count += 1
                                errors.append(f"Row {index + 2} ({email}): Employee already exists - SKIPPED")
                                continue
                            else:
                                # User exists but no employee profile (Zombie record) -> Reuse it
                                user.first_name = str(row.get("first_name", "")).strip()
                                user.last_name = str(row.get("last_name", "")).strip()
                                user.role = "EMPLOYEE"
                                user.company = self.request.user.company
                                user.save()
                        else:
                            # 2. Create New User
                            first_name = str(row.get("first_name", "")).strip()
                            last_name = str(row.get("last_name", "")).strip()

                            if not first_name:
                                raise ValueError("First Name is required")
                            if not last_name:
                                raise ValueError("Last Name is required")

                            user = User.objects.create_user(
                                username=email,
                                email=email,
                                password=None,
                                first_name=first_name,
                                last_name=last_name,
                                role="EMPLOYEE",
                                company=self.request.user.company,
                            )
                            user.set_unusable_password()
                            user.save()

                        # 3. Parse Dates
                        doj = row.get("date_of_joining")
                        if pd.isna(doj) or str(doj).strip() == "":
                            raise ValueError("Date of Joining is mandatory")
                        elif hasattr(doj, "date"):
                            doj = doj.date()
                        else:
                            # Try parsing string
                            try:
                                doj = pd.to_datetime(doj).date()
                            except Exception:
                                raise ValueError("Invalid Date of Joining format. Use YYYY-MM-DD format.")

                        dob = row.get("date_of_birth")
                        if pd.isna(dob) or str(dob).strip() == "":
                            raise ValueError("Date of Birth is mandatory")
                        elif hasattr(dob, "date"):
                            dob = dob.date()
                        else:
                            # Try parsing string
                            try:
                                dob = pd.to_datetime(dob).date()
                            except Exception:
                                raise ValueError("Invalid Date of Birth format. Use YYYY-MM-DD format.")

                        # 4. Parse Location
                        from companies.models import Location

                        location_name = str(row.get("location", "")).strip()
                        location = None
                        if location_name and location_name != "nan":
                            location = Location.objects.filter(
                                company=self.request.user.company,
                                name__iexact=location_name,
                            ).first()

                        # Fallback location
                        if not location:
                            location = Location.objects.filter(company=self.request.user.company).first()

                        if not location:
                            raise ValueError(
                                "No location found for this company. Please create at least one location first."
                            )

                        # 4.5 Reporting Manager
                        manager_name = str(row.get("reporting_manager_name", "")).strip()
                        if not manager_name or manager_name == "nan":
                            raise ValueError("Reporting Manager Name is mandatory")

                        # Find manager user
                        from django.db.models import Value
                        from django.db.models.functions import Concat

                        manager_user = (
                            User.objects.annotate(full_name=Concat("first_name", Value(" "), "last_name"))
                            .filter(
                                company=self.request.user.company,
                                full_name__iexact=manager_name,
                            )
                            .first()
                        )

                        if not manager_user:
                            # Try approximate match (split first/last)
                            parts = manager_name.split()
                            if len(parts) >= 2:
                                manager_user = User.objects.filter(
                                    company=self.request.user.company,
                                    first_name__iexact=parts[0],
                                    last_name__iexact=" ".join(parts[1:]),
                                ).first()

                        if not manager_user:
                            raise ValueError(
                                f"Reporting Manager '{manager_name}' not found. Please ensure the manager exists and the name matches exactly."
                            )

                        # 5. Create Employee
                        badge_id = str(row.get("badge_id", ""))
                        if badge_id == "nan" or badge_id == "":
                            badge_id = None
                        elif badge_id.endswith(".0"):
                            badge_id = badge_id[:-2]  # 101.0 -> 101

                        # Check for duplicate badge_id
                        if (
                            badge_id
                            and Employee.objects.filter(badge_id=badge_id, company=self.request.user.company).exists()
                        ):
                            raise ValueError(f"Badge ID '{badge_id}' already exists. Please use a unique Badge ID.")

                        # Sync Department & Designation with Role Configuration
                        from companies.models import Department, Designation

                        dept_name = str(row.get("department", "General")).strip().title()
                        desig_name = str(row.get("designation", "Employee")).strip().title()

                        if dept_name == "Nan":
                            dept_name = "General"
                        if desig_name == "Nan":
                            desig_name = "Employee"

                        # Ensure they exist in Role Config (prevents duplicates)
                        Department.objects.get_or_create(company=self.request.user.company, name=dept_name)
                        Designation.objects.get_or_create(company=self.request.user.company, name=desig_name)

                        # Parse gender and marital status
                        gender = str(row.get("gender", "M")).strip().upper()
                        if gender not in ["M", "F", "O"]:
                            gender = "M"

                        marital_status = str(row.get("marital_status", "S")).strip().upper()
                        if marital_status not in ["S", "M", "D", "W"]:
                            marital_status = "S"

                        # Parse mobile number
                        mobile = str(row.get("mobile", "")).strip()
                        if mobile == "nan" or mobile == "":
                            mobile = None

                        # Parse pseudo name
                        pseudo_name = str(row.get("pseudo_name", "")).strip()
                        if pseudo_name == "nan" or pseudo_name == "":
                            pseudo_name = None

                        # Parse annual CTC
                        annual_ctc = row.get("annual_ctc")
                        if pd.isna(annual_ctc) or annual_ctc == "":
                            annual_ctc = 0
                        else:
                            try:
                                annual_ctc = float(annual_ctc)
                            except (ValueError, TypeError):
                                annual_ctc = 0

                        Employee.objects.create(
                            user=user,
                            company=self.request.user.company,
                            manager=manager_user,
                            designation=desig_name,
                            department=dept_name,
                            location=location,
                            badge_id=badge_id,
                            mobile_number=mobile,
                            gender=gender[0] if gender else "M",
                            marital_status=marital_status[0] if marital_status else "S",
                            date_of_joining=doj,
                            dob=dob,
                            pseudo_name=pseudo_name,
                            annual_ctc=annual_ctc,
                        )

                        # 6. Create Leave Balance (handled by signal)
                        # LeaveBalance.objects.create(...) - REMOVED to avoid duplicate key error

                        # 7. Send Activation Email
                        try:
                            send_activation_email(user, self.request)
                        except Exception as email_error:
                            # Don't fail the import if email fails
                            errors.append(
                                f"Row {index + 2} ({email}): Employee created but activation email failed - {str(email_error)}"
                            )

                        success_count += 1

                except Exception as e:
                    errors.append(f"Row {index + 2} ({row.get('email', 'Unknown')}): {str(e)}")

            # Prepare success message
            if success_count > 0:
                success_msg = f"Successfully created {success_count} employee(s)."
                if skipped_count > 0:
                    success_msg += f" Skipped {skipped_count} existing employee(s)."
                messages.success(self.request, success_msg)

            if errors:
                # Show first 10 errors
                error_msg = "<strong>Import completed with errors:</strong><br><ul>"
                for error in errors[:10]:
                    error_msg += f"<li>{error}</li>"
                error_msg += "</ul>"
                if len(errors) > 10:
                    error_msg += f"<p><em>...and {len(errors) - 10} more error(s).</em></p>"
                messages.warning(self.request, error_msg)

            if success_count == 0 and not errors:
                messages.warning(self.request, "No employees were imported. Please check your file and try again.")

            return super().form_valid(form)

        except Exception as e:
            import traceback

            traceback.print_exc()
            messages.error(self.request, f"System Error: {str(e)}. Please contact support if the issue persists.")
            return self.form_invalid(form)


@login_required
def download_sample_import_file(request):
    """
    Downloads a sample Excel file for bulk employee import.
    """
    import pandas as pd

    # Define columns
    columns = [
        "First Name",
        "Last Name",
        "Pseudo Name",
        "Reporting Manager Name",
        "Email",
        "Designation",
        "Department",
        "Mobile",
        "Date of Joining",
        "Date of Birth",
        "Badge ID",
        "Gender",
        "Marital Status",
        "Annual CTC",
        "Location",
    ]

    # Create empty DataFrame
    df = pd.DataFrame(columns=columns)

    # Add a dummy row to guide user
    dummy_data = {
        "First Name": "John",
        "Last Name": "Doe",
        "Pseudo Name": "Johnny",
        "Reporting Manager Name": "Petabytz Admin",
        "Email": "john.doe@example.com",
        "Designation": "Software Engineer",
        "Department": "IT",
        "Mobile": "9876543210",
        "Date of Joining": "2025-01-01",
        "Date of Birth": "1990-05-15",
        "Badge ID": "E001",
        "Gender": "M",
        "Marital Status": "S",
        "Annual CTC": 1200000,
        "Location": "Head Office",
    }
    df = pd.concat([df, pd.DataFrame([dummy_data])], ignore_index=True)

    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=employee_bulk_import_sample.xlsx"

    try:
        # Use openpyxl engine
        df.to_excel(response, index=False, engine="openpyxl")
    except ImportError:
        # Fallback if openpyxl missing
        return HttpResponse(
            "Error: openpyxl library not found. Please contact admin to install it.",
            status=500,
        )
    except Exception as e:
        return HttpResponse(f"Error generating file: {str(e)}", status=500)

    return response


# --- Regularization Views ---


class RegularizationCreateView(LoginRequiredMixin, CreateView):
    model = RegularizationRequest
    form_class = RegularizationRequestForm
    template_name = "employees/regularization_form.html"
    success_url = reverse_lazy("regularization_list")  # Redirect to list or profile

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass user's regularizations to show history on same page if needed
        if hasattr(self.request.user, "employee_profile"):
            context["history"] = RegularizationRequest.objects.filter(employee=self.request.user.employee_profile)
        return context

    def form_valid(self, form):
        if not hasattr(self.request.user, "employee_profile"):
            form.add_error(None, "You do not have an employee profile.")
            return self.form_invalid(form)

        employee = self.request.user.employee_profile

        # Server-side duplicate prevention: Check for recent duplicate submissions
        from datetime import timedelta

        from django.utils import timezone

        recent_duplicate = RegularizationRequest.objects.filter(
            employee=employee,
            date=form.cleaned_data["date"],
            created_at__gte=timezone.now() - timedelta(seconds=10),
        ).exists()

        if recent_duplicate:
            from django.contrib import messages

            messages.warning(
                self.request,
                "You just submitted a regularization request for this date. Please wait before submitting again.",
            )
            return self.form_invalid(form)

        form.instance.employee = employee
        response = super().form_valid(form)

        # Send Email Notification asynchronously to avoid blocking the response
        import threading

        def send_email_async():
            try:
                import logging

                from core.email_utils import send_regularization_request_notification

                logger = logging.getLogger(__name__)
                result = send_regularization_request_notification(self.object)
                if not result.get("hr", False):
                    logger.error(f"Failed to send regularization request email to HR for {self.object.id}")
            except Exception as e:
                import logging

                logger = logging.getLogger(__name__)
                logger.error(f"Error calling regularization email utility: {e}")

        # Start email sending in background thread
        email_thread = threading.Thread(target=send_email_async)
        email_thread.daemon = True
        email_thread.start()

        return response

    def get_success_url(self):
        # If admin/manager, maybe go to list, if employee go to profile/home
        return reverse_lazy("personal_home")  # Or wherever "My Regularizations" are shown


class RegularizationListView(LoginRequiredMixin, ListView):
    model = RegularizationRequest
    template_name = "employees/regularization_list.html"
    context_object_name = "requests"
    ordering = ["-created_at"]

    def get_queryset(self):
        user = self.request.user
        qs = RegularizationRequest.objects.all()

        if user.role == User.Role.COMPANY_ADMIN:
            return qs.filter(employee__company=user.company)
        elif user.role == User.Role.MANAGER:
            # Show requests from subordinates
            return qs.filter(employee__manager=user)
        else:
            # Employees see their own
            employee = safe_get_employee_profile(user)
            if employee:
                return qs.filter(employee=employee)
            return qs.none()


@login_required
def approve_regularization(request, pk):
    if request.method == "POST":
        reg_request = RegularizationRequest.objects.get(pk=pk)

        user = request.user
        is_admin = user.role == User.Role.COMPANY_ADMIN or user.is_superuser
        is_manager = (
            user.role == User.Role.MANAGER and reg_request.employee.manager and reg_request.employee.manager == user
        )

        if not (is_admin or is_manager):
            return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)

        # Approve
        reg_request.status = "APPROVED"
        reg_request.approved_by = user
        reg_request.approved_at = timezone.now()
        reg_request.manager_comment = request.POST.get("manager_comment", "")
        reg_request.save()

        # Update Attendance
        # We need to find or create the attendance record for that date
        attendance, created = Attendance.objects.get_or_create(employee=reg_request.employee, date=reg_request.date)

        # Get employee location timezone
        from core.utils import get_user_timezone

        tz_name = get_user_timezone(reg_request.employee.user, reg_request.employee.company)
        attendance.user_timezone = tz_name

        import pytz

        local_tz = pytz.timezone(tz_name)

        if reg_request.check_in:
            # Combine date and time, then localize to employee's timezone
            local_dt = timezone.datetime.combine(reg_request.date, reg_request.check_in)
            attendance.clock_in = local_tz.localize(local_dt)

        if reg_request.check_out:
            local_dt = timezone.datetime.combine(reg_request.date, reg_request.check_out)
            # Handle possible overnight shift if check_out < check_in (though form currently validates against this)
            if reg_request.check_in and reg_request.check_out < reg_request.check_in:
                local_dt += timedelta(days=1)
            attendance.clock_out = local_tz.localize(local_dt)

        # Update status to Present if not already (or whatever logic user wants, implicitly if regulating, they were present)
        attendance.status = "PRESENT"

        # Create or update AttendanceSession for consistency
        if reg_request.check_in and reg_request.check_out:
            # Check if there's already a session for this date
            existing_session = AttendanceSession.objects.filter(
                employee=reg_request.employee, date=reg_request.date
            ).first()

            if existing_session:
                # Update existing session
                existing_session.clock_in = attendance.clock_in
                existing_session.clock_out = attendance.clock_out
                existing_session.session_type = "WEB"  # Default to web for regularized entries
                existing_session.save()
            else:
                # Create new session
                AttendanceSession.objects.create(
                    employee=reg_request.employee,
                    date=reg_request.date,
                    session_number=1,
                    clock_in=attendance.clock_in,
                    clock_out=attendance.clock_out,
                    session_type="WEB",
                    is_active=False,
                    location_validated=True,  # Assume validated for regularized entries
                )

        # Re-calc late/early
        attendance.calculate_late_arrival()
        attendance.calculate_early_departure()

        # Recalculate working hours after regularization
        attendance.calculate_total_working_hours()

        attendance.save()

        # Send Approval Email asynchronously
        import threading

        def send_email_async():
            try:
                import logging

                from core.email_utils import send_regularization_approval_notification

                logger = logging.getLogger(__name__)
                if not send_regularization_approval_notification(reg_request):
                    logger.warning("Regularization approval email notification failed")
            except Exception as e:
                import logging

                logger = logging.getLogger(__name__)
                logger.error(f"Error sending regularization approval email: {e}")

        # Start email sending in background thread
        email_thread = threading.Thread(target=send_email_async)
        email_thread.daemon = True
        email_thread.start()

        messages.success(request, "Regularization approved. Notification will be sent.")

        return redirect(request.META.get("HTTP_REFERER", "regularization_list"))

    return redirect("regularization_list")


@login_required
def reject_regularization(request, pk):
    if request.method == "POST":
        reg_request = RegularizationRequest.objects.get(pk=pk)

        user = request.user
        is_admin = user.role == User.Role.COMPANY_ADMIN or user.is_superuser
        is_manager = (
            user.role == User.Role.MANAGER and reg_request.employee.manager and reg_request.employee.manager == user
        )

        if not (is_admin or is_manager):
            return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)

        reg_request.status = "REJECTED"
        reg_request.manager_comment = request.POST.get(
            "rejection_reason", ""
        )  # Use manager_comment for rejection reason
        reg_request.save()

        # Send Rejection Email asynchronously
        import threading

        def send_email_async():
            try:
                import logging

                from core.email_utils import send_regularization_rejection_notification

                logger = logging.getLogger(__name__)
                if not send_regularization_rejection_notification(reg_request):
                    logger.warning("Regularization rejection email notification failed")
            except Exception as e:
                import logging

                logger = logging.getLogger(__name__)
                logger.error(f"Error sending regularization rejection email: {e}")

        # Start email sending in background thread
        email_thread = threading.Thread(target=send_email_async)
        email_thread.daemon = True
        email_thread.start()

        messages.success(request, "Regularization rejected. Notification will be sent.")

        return redirect(request.META.get("HTTP_REFERER", "regularization_list"))

    return redirect("regularization_list")


# --- Leave Configuration ---


@login_required
def leave_configuration(request):
    from django.contrib import messages

    user = request.user
    if user.role not in [User.Role.COMPANY_ADMIN, User.Role.MANAGER]:
        messages.error(request, "Permission Denied")
        return redirect("dashboard")

    company = user.company

    # Manager sees team, Admin sees all
    if user.role == User.Role.MANAGER:
        all_employees = Employee.objects.filter(manager=user)
    else:
        all_employees = Employee.objects.filter(company=company)

    # Apply status filter
    status_filter = request.GET.get("status", "active")
    if status_filter == "active":
        employees = all_employees.filter(is_active=True)
    elif status_filter == "inactive":
        employees = all_employees.filter(is_active=False)
    else:
        employees = all_employees

    # Prefetch leave balances to avoid N+1 queries and ensure fresh data
    employees = employees.select_related("user", "leave_balance").order_by("user__first_name")

    # Ensure all employees have leave balance records and refresh from DB
    from django.core.exceptions import ObjectDoesNotExist

    for employee in employees:
        try:
            # Force refresh from database to get latest data
            employee.refresh_from_db()
            leave_balance = employee.leave_balance
            leave_balance.refresh_from_db()
        except ObjectDoesNotExist:
            # Create with 0 leaves for new employees (probation period)
            LeaveBalance.objects.create(employee=employee, casual_leave_allocated=0.0, sick_leave_allocated=0.0)

    # Context for Accrual Modal (Safe from template syntax errors)
    import calendar

    from django.utils import timezone

    now = timezone.now()
    current_month = now.month
    current_year = now.year

    months_ctx = []
    for i in range(1, 13):
        months_ctx.append(
            {
                "value": i,
                "name": calendar.month_name[i],
                "selected": "selected" if i == current_month else "",
            }
        )

    years_ctx = []
    # Show current year and next 2 years, or surrounding
    for y in [2024, 2025, 2026]:
        years_ctx.append({"value": y, "selected": "selected" if y == current_year else ""})

    return render(
        request,
        "employees/leave_configuration.html",
        {
            "employees": employees,
            "months_ctx": months_ctx,
            "years_ctx": years_ctx,
            "active_count": all_employees.filter(is_active=True).count(),
            "inactive_count": all_employees.filter(is_active=False).count(),
            "selected_status": status_filter,
            "is_bluebix": company.name.lower() in ["bluebix", "softstandard", "softstandard solutions"],
        },
    )


@csrf_exempt
@login_required
def update_leave_balance(request, pk):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)

    user = request.user
    if user.role not in [User.Role.COMPANY_ADMIN, User.Role.MANAGER]:
        return JsonResponse({"status": "error", "message": "Permission Denied"}, status=403)

    try:
        employee = Employee.objects.get(pk=pk)
        # Verify access rights
        if user.role == User.Role.MANAGER and employee.manager != user:
            return JsonResponse({"status": "error", "message": "Permission Denied"}, status=403)

        if user.role == User.Role.COMPANY_ADMIN and employee.company != user.company:
            return JsonResponse({"status": "error", "message": "Permission Denied"}, status=403)

        balance = employee.leave_balance

        # Get data
        data = json.loads(request.body)
        sick_balance_desired = data.get("sick_leave_allocated")  # This is actually the desired balance
        casual_balance_desired = data.get("casual_leave_allocated")  # This is actually the desired balance
        combined_balance_desired = data.get("combined_sick_casual_allocated")  # For Bluebix combined leave

        # For Bluebix, handle combined sick/casual leave
        if employee.company.name.lower() in ["bluebix", "softstandard", "softstandard solutions"]:
            # For Bluebix, use the combined field
            if combined_balance_desired is not None:
                try:
                    desired_balance = float(combined_balance_desired)
                    # Calculate new allocation: desired_balance + current_used
                    new_allocation = desired_balance + balance.combined_sick_casual_used
                    balance.combined_sick_casual_allocated = new_allocation
                except ValueError:
                    pass  # Ignore invalid
            # Also handle legacy sick/casual updates (for backward compatibility)
            elif sick_balance_desired is not None or casual_balance_desired is not None:
                try:
                    # Use whichever value is provided (they should be the same for Bluebix)
                    desired_balance = float(sick_balance_desired or casual_balance_desired or 0)
                    # Calculate new allocation: desired_balance + current_used
                    new_allocation = desired_balance + balance.combined_sick_casual_used
                    balance.combined_sick_casual_allocated = new_allocation
                except ValueError:
                    pass  # Ignore invalid
        else:
            # For other companies, handle separate pools
            if sick_balance_desired is not None:
                # Handle empty strings or invalid input gracefully
                try:
                    desired_balance = float(sick_balance_desired)
                    # Calculate new allocation: desired_balance + current_used
                    new_allocation = desired_balance + balance.sick_leave_used
                    balance.sick_leave_allocated = new_allocation
                except ValueError:
                    pass  # Ignore invalid

            if casual_balance_desired is not None:
                try:
                    desired_balance = float(casual_balance_desired)
                    # Calculate new allocation: desired_balance + current_used
                    new_allocation = desired_balance + balance.casual_leave_used
                    balance.casual_leave_allocated = new_allocation
                except ValueError:
                    pass

        balance.save()

        return JsonResponse({"status": "success", "message": "Balance updated successfully"})

    except Employee.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Employee not found"}, status=404)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
def run_monthly_accrual(request):
    import calendar

    from django.contrib import messages

    user = request.user
    if user.role not in [User.Role.COMPANY_ADMIN, User.Role.MANAGER]:
        messages.error(request, "Permission Denied")
        return redirect("leave_configuration")

    from django.core.management import call_command

    try:
        month = request.POST.get("month")
        year = request.POST.get("year")

        period_msg = ""
        if month and year:
            try:
                month_name = calendar.month_name[int(month)]
                period_msg = f"for {month_name} {year}"
            except (ValueError, IndexError) as e:
                logger.debug(
                    "Failed to parse month/year for accrual message",
                    month=month,
                    year=year,
                    error=str(e),
                )

        # Run the command
        call_command("accrue_monthly_leaves")

        success_msg = (
            f"Monthly accrual processed {period_msg}: +1 Sick and +1 Casual leave added to all employees."
            if period_msg
            else "Monthly accrual processed: +1 Sick and +1 Casual leave added to all employees."
        )

        messages.success(request, success_msg)

    except Exception as e:
        logger.exception("Error running monthly leave accrual", error=str(e))
        capture_exception(e, properties={"action": "manual_accrual"})
        messages.error(request, f"Error running accrual: {str(e)}")

    return redirect("leave_configuration")


@login_required
def bulk_leave_upload(request):
    """Handle bulk leave balance upload"""
    from django.contrib import messages

    from .forms import PANDAS_AVAILABLE, BulkLeaveUploadForm

    user = request.user
    if user.role not in [User.Role.COMPANY_ADMIN]:
        messages.error(request, "Permission Denied. Only Company Admins can upload bulk leave data.")
        return redirect("leave_configuration")

    if not PANDAS_AVAILABLE:
        messages.error(request, "Bulk upload feature is not available. Please install pandas and openpyxl packages.")
        return redirect("leave_configuration")

    if request.method == "POST":
        form = BulkLeaveUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Validate file content
                processed_data, errors = form.validate_file_content(user.company)

                if errors:
                    for error in errors:
                        messages.error(request, error)
                    return render(request, "employees/bulk_leave_upload.html", {"form": form})

                if not processed_data:
                    messages.warning(request, "No valid data found in the uploaded file.")
                    return render(request, "employees/bulk_leave_upload.html", {"form": form})

                # Process the data
                update_mode = form.cleaned_data["update_mode"]
                success_count = 0
                error_count = 0

                with transaction.atomic():
                    for data in processed_data:
                        try:
                            employee = data["employee"]

                            # Get or create leave balance
                            leave_balance, created = LeaveBalance.objects.get_or_create(
                                employee=employee,
                                defaults={
                                    "casual_leave_allocated": 0.0,
                                    "sick_leave_allocated": 0.0,
                                    "casual_leave_used": 0.0,
                                    "sick_leave_used": 0.0,
                                    "carry_forward_leave": 0.0,
                                },
                            )

                            # Log the data being processed for debugging
                            logger.info(
                                f"Processing bulk upload for {employee.user.get_full_name()}: "
                                f"CL_allocated: {data['casual_leave_allocated']}, "
                                f"SL_allocated: {data['sick_leave_allocated']}, "
                                f"CL_used: {data['casual_leave_used']}, "
                                f"SL_used: {data['sick_leave_used']}, "
                                f"CF: {data['carry_forward_leave']}"
                            )

                            # Update based on mode and company type
                            if data.get("is_bluebix", False):
                                # Bluebix: Handle combined sick/casual leave
                                if update_mode == "REPLACE":
                                    leave_balance.combined_sick_casual_allocated = data[
                                        "combined_sick_casual_allocated"
                                    ]
                                    leave_balance.combined_sick_casual_used = data["combined_sick_casual_used"]
                                    leave_balance.carry_forward_leave = data["carry_forward_leave"]

                                elif update_mode == "ADD":
                                    leave_balance.combined_sick_casual_allocated += data[
                                        "combined_sick_casual_allocated"
                                    ]
                                    leave_balance.combined_sick_casual_used += data["combined_sick_casual_used"]
                                    leave_balance.carry_forward_leave += data["carry_forward_leave"]

                                elif update_mode == "UPDATE_ONLY":
                                    # Only update non-zero values
                                    if data["combined_sick_casual_allocated"] > 0:
                                        leave_balance.combined_sick_casual_allocated = data[
                                            "combined_sick_casual_allocated"
                                        ]
                                    if data["combined_sick_casual_used"] > 0:
                                        leave_balance.combined_sick_casual_used = data["combined_sick_casual_used"]
                                    if data["carry_forward_leave"] > 0:
                                        leave_balance.carry_forward_leave = data["carry_forward_leave"]

                                logger.info(
                                    f"Bluebix bulk upload - {employee.user.get_full_name()}: "
                                    f"Combined allocated: {leave_balance.combined_sick_casual_allocated}, "
                                    f"Combined used: {leave_balance.combined_sick_casual_used}, "
                                    f"CF: {leave_balance.carry_forward_leave}"
                                )
                            else:
                                # Other companies: Handle separate casual and sick leave
                                if update_mode == "REPLACE":
                                    leave_balance.casual_leave_allocated = data["casual_leave_allocated"]
                                    leave_balance.sick_leave_allocated = data["sick_leave_allocated"]
                                    leave_balance.casual_leave_used = data["casual_leave_used"]
                                    leave_balance.sick_leave_used = data["sick_leave_used"]
                                    leave_balance.carry_forward_leave = data["carry_forward_leave"]

                                elif update_mode == "ADD":
                                    leave_balance.casual_leave_allocated += data["casual_leave_allocated"]
                                    leave_balance.sick_leave_allocated += data["sick_leave_allocated"]
                                    leave_balance.casual_leave_used += data["casual_leave_used"]
                                    leave_balance.sick_leave_used += data["sick_leave_used"]
                                    leave_balance.carry_forward_leave += data["carry_forward_leave"]

                                elif update_mode == "UPDATE_ONLY":
                                    # Only update non-zero values
                                    if data["casual_leave_allocated"] > 0:
                                        leave_balance.casual_leave_allocated = data["casual_leave_allocated"]
                                    if data["sick_leave_allocated"] > 0:
                                        leave_balance.sick_leave_allocated = data["sick_leave_allocated"]
                                    if data["casual_leave_used"] > 0:
                                        leave_balance.casual_leave_used = data["casual_leave_used"]
                                    if data["sick_leave_used"] > 0:
                                        leave_balance.sick_leave_used = data["sick_leave_used"]
                                    if data["carry_forward_leave"] > 0:
                                        leave_balance.carry_forward_leave = data["carry_forward_leave"]

                                logger.info(
                                    f"Standard bulk upload - {employee.user.get_full_name()}: "
                                    f"CL: {leave_balance.casual_leave_allocated}/{leave_balance.casual_leave_used}, "
                                    f"SL: {leave_balance.sick_leave_allocated}/{leave_balance.sick_leave_used}, "
                                    f"CF: {leave_balance.carry_forward_leave}"
                                )

                            # Save the updated leave balance with validation
                            leave_balance.validate_and_save()

                            # Force refresh from database to ensure consistency
                            leave_balance.refresh_from_db()

                            # Log the values after saving and refresh
                            logger.info(
                                f"After save - {employee.user.get_full_name()}: "
                                f"CL: {leave_balance.casual_leave_allocated}/{leave_balance.casual_leave_used}, "
                                f"SL: {leave_balance.sick_leave_allocated}/{leave_balance.sick_leave_used}, "
                                f"CF: {leave_balance.carry_forward_leave}"
                            )

                            # Clear any cached data related to this employee
                            from django.core.cache import cache

                            cache_keys_to_clear = [
                                f"employee_leave_balance_{employee.id}",
                                f"employee_dashboard_data_{employee.id}",
                                f"employee_personal_home_{employee.id}",
                                f"leave_config_data_{user.company.id}",
                            ]
                            for cache_key in cache_keys_to_clear:
                                cache.delete(cache_key)

                            # Trigger any related model updates
                            # This ensures that any dependent calculations are updated
                            employee.save(update_fields=["updated_at"])

                            # Log the change for audit trail
                            logger.info(
                                f"Bulk upload: Updated leave balance for {employee.user.get_full_name()} - "
                                f"CL: {leave_balance.casual_leave_allocated}/{leave_balance.casual_leave_used}, "
                                f"SL: {leave_balance.sick_leave_allocated}/{leave_balance.sick_leave_used}, "
                                f"CF: {leave_balance.carry_forward_leave}"
                            )

                            success_count += 1

                        except Exception as e:
                            error_count += 1
                            logger.error(f"Error updating leave balance for {data['employee_name']}: {str(e)}")

                # Show results with detailed feedback
                if success_count > 0:
                    messages.success(
                        request,
                        f"Successfully updated leave balances for {success_count} employees. "
                        f"All changes are now reflected across the system.",
                    )
                    # Add detailed success message for debugging
                    messages.info(
                        request,
                        f"Updated employees: {', '.join([data['employee_name'] for data in processed_data[:5]])}"
                        f"{'...' if len(processed_data) > 5 else ''}",
                    )

                if error_count > 0:
                    messages.warning(
                        request, f"Failed to update {error_count} employee records. Check logs for details."
                    )

                # Clear company-wide cache to ensure immediate reflection
                from django.core.cache import cache

                cache.delete(f"leave_config_data_{user.company.id}")
                cache.delete(f"company_leave_summary_{user.company.id}")

                # Force clear all employee-related cache for this company
                company_employees = Employee.objects.filter(company=user.company)
                for emp in company_employees:
                    cache.delete(f"employee_leave_balance_{emp.id}")
                    cache.delete(f"employee_dashboard_data_{emp.id}")
                    cache.delete(f"employee_profile_data_{emp.id}")
                    cache.delete(f"employee_personal_home_{emp.id}")

                # Add a flag to indicate successful bulk upload for frontend handling
                messages.info(request, "BULK_UPLOAD_SUCCESS")  # Special flag for frontend

                return redirect("leave_configuration")

            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                logger.error(f"Bulk leave upload error: {str(e)}")

        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = BulkLeaveUploadForm()

    return render(request, "employees/bulk_leave_upload.html", {"form": form})


@login_required
def download_leave_template(request):
    """Download Excel template for bulk leave upload"""
    from django.contrib import messages

    user = request.user
    if user.role not in [User.Role.COMPANY_ADMIN]:
        messages.error(request, "Permission Denied")
        return redirect("leave_configuration")

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        messages.error(request, "Excel template generation is not available. Please install openpyxl package.")
        return redirect("leave_configuration")

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Balance Template"

    # Check if this is Bluebix or Softstandard company (combined leave system)
    is_bluebix = user.company.name.lower() in ["bluebix", "softstandard", "softstandard solutions"]

    # Define headers based on company type
    if is_bluebix:
        headers = [
            "employee_id",
            "employee_name",
            "sick_casual_leave_allocated",
            "sick_casual_leave_used",
            "carry_forward_leave",
        ]
    else:
        headers = [
            "employee_id",
            "employee_name",
            "casual_leave_allocated",
            "sick_leave_allocated",
            "casual_leave_used",
            "sick_leave_used",
            "carry_forward_leave",
        ]

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add sample data with company employees
    employees = Employee.objects.filter(company=user.company).select_related("user", "leave_balance")[:5]

    for row, employee in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=employee.badge_id or f"EMP{employee.id:03d}")
        ws.cell(row=row, column=2, value=employee.user.get_full_name())

        # Get current leave balance or defaults
        try:
            balance = employee.leave_balance
            if is_bluebix:
                # Bluebix: Combined sick/casual leave
                ws.cell(row=row, column=3, value=balance.combined_sick_casual_allocated)
                ws.cell(row=row, column=4, value=balance.combined_sick_casual_used)
                ws.cell(row=row, column=5, value=balance.carry_forward_leave)
            else:
                # Other companies: Separate casual and sick leave
                ws.cell(row=row, column=3, value=balance.casual_leave_allocated)
                ws.cell(row=row, column=4, value=balance.sick_leave_allocated)
                ws.cell(row=row, column=5, value=balance.casual_leave_used)
                ws.cell(row=row, column=6, value=balance.sick_leave_used)
                ws.cell(row=row, column=7, value=balance.carry_forward_leave)
        except Exception:
            # Default values for new employees
            if is_bluebix:
                # Bluebix defaults: 1 combined sick/casual leave
                ws.cell(row=row, column=3, value=1.0)  # Combined allocated
                ws.cell(row=row, column=4, value=0.0)  # Combined used
                ws.cell(row=row, column=5, value=0.0)  # Carry forward
            else:
                # Other companies defaults
                ws.cell(row=row, column=3, value=12.0)  # Default CL
                ws.cell(row=row, column=4, value=12.0)  # Default SL
                ws.cell(row=row, column=5, value=0.0)  # Used CL
                ws.cell(row=row, column=6, value=0.0)  # Used SL
                ws.cell(row=row, column=7, value=0.0)  # Carry forward

    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")

    if is_bluebix:
        instructions = [
            "BULK LEAVE UPLOAD INSTRUCTIONS - BLUEBIX / SOFTSTANDARD (COMBINED SICK/CASUAL LEAVE)",
            "",
            "Required Columns:",
            "- employee_id: Employee badge ID (e.g., PBTHYD001)",
            "- employee_name: Full name of employee",
            "- sick_casual_leave_allocated: Total sick/casual leaves allocated for the year",
            "",
            "Optional Columns:",
            "- sick_casual_leave_used: Sick/casual leaves already used (default: 0)",
            "- carry_forward_leave: Leaves carried forward from previous year (default: 0)",
            "",
            "Important Notes for Bluebix / Softstandard:",
            "1. This company uses a COMBINED sick/casual leave system",
            "2. Employees get 1 combined leave per month that can be used for either sick or casual leave",
            "3. Either employee_id OR employee_name must be provided",
            "4. All numeric values must be positive numbers",
            "5. Used leaves cannot exceed allocated leaves",
            "6. File formats supported: .xlsx, .xls, .csv",
            "7. Maximum file size: 5MB",
            "",
            "Update Modes:",
            "- REPLACE: Completely replace existing leave balances",
            "- ADD: Add values to existing balances",
            "- UPDATE_ONLY: Update only non-zero values in the file",
        ]
    else:
        instructions = [
            "BULK LEAVE UPLOAD INSTRUCTIONS - SEPARATE CASUAL & SICK LEAVE",
            "",
            "Required Columns:",
            "- employee_id: Employee badge ID (e.g., PBTHYD001)",
            "- employee_name: Full name of employee",
            "- casual_leave_allocated: Total casual leaves allocated for the year",
            "- sick_leave_allocated: Total sick leaves allocated for the year",
            "",
            "Optional Columns:",
            "- casual_leave_used: Casual leaves already used (default: 0)",
            "- sick_leave_used: Sick leaves already used (default: 0)",
            "- carry_forward_leave: Leaves carried forward from previous year (default: 0)",
            "",
            "Important Notes:",
            "1. Either employee_id OR employee_name must be provided",
            "2. All numeric values must be positive numbers",
            "3. Used leaves cannot exceed allocated leaves",
            "4. File formats supported: .xlsx, .xls, .csv",
            "5. Maximum file size: 5MB",
            "",
            "Update Modes:",
            "- REPLACE: Completely replace existing leave balances",
            "- ADD: Add values to existing balances",
            "- UPDATE_ONLY: Update only non-zero values in the file",
        ]

    for row, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row, column=1, value=instruction)
        if row == 1:  # Title
            cell.font = Font(bold=True, size=14)
        elif instruction.endswith(":"):  # Section headers
            cell.font = Font(bold=True)

    # Adjust column widths
    for ws_sheet in [ws, instructions_ws]:
        for column in ws_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sheet.column_dimensions[column_letter].width = adjusted_width

    # Create response with company-specific filename
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    company_name = user.company.name.replace(" ", "_")
    if is_bluebix:
        filename = f"leave_balance_template_bluebix_combined_{company_name}.xlsx"
    else:
        filename = f"leave_balance_template_separate_{company_name}.xlsx"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# --- ID Proof Management Views ---


@login_required
def employee_id_proofs(request, pk):
    """
    View and manage employee ID proofs.
    Admin has full access (view, edit, delete, upload).
    Employee can only upload if not already uploaded.
    """
    try:
        employee = Employee.objects.get(pk=pk)

        # Permission Check
        user = request.user
        is_admin = user.role == User.Role.COMPANY_ADMIN
        is_self = hasattr(user, "employee_profile") and user.employee_profile == employee

        if not (is_admin or is_self):
            messages.error(request, "Permission denied")
            return redirect("employee_list")

        # Get or create ID proofs
        id_proofs, created = EmployeeIDProof.objects.get_or_create(employee=employee)

        if request.method == "POST":
            action = request.POST.get("action")

            # Upload functionality
            if action == "upload":
                # Helper to check if upload allowed
                def can_upload(current_file):
                    return not current_file or is_admin

                uploaded = False

                if "aadhar_front" in request.FILES:
                    if can_upload(id_proofs.aadhar_front):
                        id_proofs.aadhar_front = request.FILES["aadhar_front"]
                        uploaded = True
                    else:
                        messages.warning(
                            request,
                            "Aadhar Front already uploaded. Only admin can replace.",
                        )

                if "aadhar_back" in request.FILES:
                    if can_upload(id_proofs.aadhar_back):
                        id_proofs.aadhar_back = request.FILES["aadhar_back"]
                        uploaded = True
                    else:
                        messages.warning(
                            request,
                            "Aadhar Back already uploaded. Only admin can replace.",
                        )

                if "pan_card" in request.FILES:
                    if can_upload(id_proofs.pan_card):
                        id_proofs.pan_card = request.FILES["pan_card"]
                        uploaded = True
                    else:
                        messages.warning(
                            request,
                            "PAN Card already uploaded. Only admin can replace.",
                        )

                if uploaded:
                    id_proofs.save()
                    messages.success(request, "ID proof(s) uploaded successfully!")

            # Delete functionality (Admin only)
            elif action == "delete" and is_admin:
                delete_type = request.POST.get("delete_type")

                if delete_type == "aadhar_front" and id_proofs.aadhar_front:
                    id_proofs.aadhar_front.delete(save=False)
                    id_proofs.aadhar_front = None
                    messages.success(request, "Aadhar Front deleted successfully!")

                elif delete_type == "aadhar_back" and id_proofs.aadhar_back:
                    id_proofs.aadhar_back.delete(save=False)
                    id_proofs.aadhar_back = None
                    messages.success(request, "Aadhar Back deleted successfully!")

                elif delete_type == "pan_card" and id_proofs.pan_card:
                    id_proofs.pan_card.delete(save=False)
                    id_proofs.pan_card = None
                    messages.success(request, "PAN Card deleted successfully!")

                id_proofs.save()

            return redirect("employee_id_proofs", pk=pk)

        context = {
            "employee": employee,
            "id_proofs": id_proofs,
            "is_admin": is_admin,
            "is_self": is_self,
        }

        return render(request, "employees/employee_id_proofs.html", context)

    except Employee.DoesNotExist:
        messages.error(request, "Employee not found")
        return redirect("employee_list")


# Emergency Contact Management Views


@login_required
@require_http_methods(["POST"])
def add_emergency_contact(request):
    """
    AJAX endpoint to add a new emergency contact
    """
    try:
        employee = request.user.employee_profile

        from .forms import EmergencyContactForm

        form = EmergencyContactForm(request.POST)
        if form.is_valid():
            contact = form.save(commit=False)
            contact.employee = employee
            contact.save()

            return JsonResponse(
                {
                    "status": "success",
                    "message": "Emergency contact added successfully",
                    "contact": {
                        "id": contact.id,
                        "name": contact.name,
                        "phone_number": contact.phone_number,
                        "relationship": contact.relationship,
                        "is_primary": contact.is_primary,
                    },
                }
            )
        else:
            return JsonResponse(
                {"status": "error", "message": "Invalid data", "errors": form.errors},
                status=400,
            )
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_emergency_contact(request, contact_id):
    """
    AJAX endpoint to delete an emergency contact
    """
    try:
        employee = request.user.employee_profile
        from .models import EmergencyContact

        contact = EmergencyContact.objects.get(id=contact_id, employee=employee)
        contact.delete()

        return JsonResponse({"status": "success", "message": "Emergency contact deleted successfully"})
    except EmergencyContact.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Contact not found"}, status=404)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_emergency_contact(request, contact_id):
    """
    AJAX endpoint to update an emergency contact
    """
    try:
        employee = request.user.employee_profile
        from .forms import EmergencyContactForm
        from .models import EmergencyContact

        contact = EmergencyContact.objects.get(id=contact_id, employee=employee)

        form = EmergencyContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()

            return JsonResponse(
                {
                    "status": "success",
                    "message": "Emergency contact updated successfully",
                    "contact": {
                        "id": contact.id,
                        "name": contact.name,
                        "phone_number": contact.phone_number,
                        "relationship": contact.relationship,
                        "is_primary": contact.is_primary,
                    },
                }
            )
        else:
            return JsonResponse(
                {"status": "error", "message": "Invalid data", "errors": form.errors},
                status=400,
            )
    except EmergencyContact.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Contact not found"}, status=404)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
def employee_id_card(request):
    """
    View to display the employee's ID card.
    """
    try:
        if hasattr(request.user, "employee_profile"):
            employee = request.user.employee_profile
        else:
            messages.error(request, "Employee profile not found.")
            return redirect("dashboard")

    except Exception:
        messages.error(request, "Employee profile not found.")
        return redirect("dashboard")

    context = {
        "employee": employee,
        "company": employee.company,
    }
    return render(request, "employees/id_card.html", context)

import calendar
import random
import json
from datetime import date, datetime, timedelta

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill

from accounts.models import User
from companies.models import Holiday
from employees.models import (
    Attendance,
    Employee,
    HandbookSection,
    LeaveBalance,
    LeaveRequest,
    Payslip,
    PolicySection,
)

from .decorators import admin_required, manager_required
from .error_handling import (
    safe_get_employee_profile,
)
from .forms import ForgotPasswordForm, OTPVerificationForm, ResetPasswordForm
from .models import PasswordResetOTP
from .utils import save_pdf_to_model
from employees.payroll_utils import calculate_payslip_breakdown, num2words_indian, num2words_flexible


@login_required
def dashboard(request):
    """Role-based Dashboard - Different views for Admin, Manager, and Employee"""

    """Role-based Dashboard - Different views for Admin, Manager, and Employee"""

    if request.user.role == User.Role.SUPERADMIN:
        return redirect("superadmin:dashboard")

    if request.user.role == User.Role.COMPANY_ADMIN:
        return admin_dashboard(request)

    role = request.user.role

    if role == User.Role.COMPANY_ADMIN:
        return admin_dashboard(request)

    elif role == User.Role.MANAGER:
        return manager_dashboard(request)

    else:
        # Default to employee dashboard

        return employee_dashboard(request)


@login_required
@manager_required
def manager_dashboard(request):
    """Manager Dashboard - Restricts view to own team"""
    from companies.models import Announcement

    if not hasattr(request.user, "employee_profile"):
        messages.error(request, "Manager profile not found.")

        return redirect("personal_home")

    manager_profile = request.user.employee_profile

    today = timezone.localtime().date()

    # Get direct reports (Team)
    # Filter out exited employees (keep if active or exit date is today/future)
    team_members = Employee.objects.filter(manager=request.user).filter(Q(is_active=True) | Q(exit_date__gte=today))

    team_ids = team_members.values_list("id", flat=True)

    # 1. Team Stats

    total_team_members = team_members.count()

    # Today's Attendance for Team

    team_attendance = Attendance.objects.filter(employee__in=team_ids, date=today).select_related(
        "employee", "employee__user"
    )

    present_count = team_attendance.filter(status="PRESENT").count()

    wfh_count = team_attendance.filter(status__in=["WFH", "ON_DUTY"]).count()

    leave_count = team_attendance.filter(status="LEAVE").count()

    # Late Arrivals in Team

    office_start = datetime.strptime("09:00", "%H:%M").time()

    late_count = 0

    for att in team_attendance:
        if att.clock_in:
            local_clock_in = timezone.localtime(att.clock_in).time()

            if local_clock_in > office_start:
                late_count += 1

    # 2. Pending Team Leave Requests

    pending_leaves = (
        LeaveRequest.objects.filter(employee__in=team_ids, status="PENDING")
        .select_related("employee", "employee__user")
        .order_by("-created_at")
    )

    # 3. Team Recent Activity (e.g. recent clock-ins)

    recent_activity = team_attendance.order_by("-clock_in")[:5]

    # 4. Announcements (Company-wide or Location-specific)
    announcements = (
        Announcement.objects.filter(company=manager_profile.company, is_active=True)
        .filter(Q(location__isnull=True) | Q(location=manager_profile.location))
        .order_by("-created_at")[:5]
    )

    # 5. Celebrations
    # 5. Celebrations
    # Filter company employees for celebrations (exclude exited)
    company_employees = Employee.objects.filter(company=manager_profile.company).filter(
        Q(is_active=True) | Q(exit_date__gte=today)
    )

    # Birthdays
    birthdays = company_employees.filter(dob__month=today.month, dob__day=today.day)

    # Work Anniversaries
    work_anniversaries = company_employees.filter(
        date_of_joining__month=today.month,
        date_of_joining__day=today.day,
        date_of_joining__lt=today,
    )

    # 6. Upcoming Holidays

    upcoming_holidays = (
        Holiday.objects.filter(company=manager_profile.company, date__gte=today, is_active=True)
        .filter(
            Q(location__name__iexact="Global")
            | Q(location__name__iexact="All Locations")
            | Q(location=manager_profile.location)
        )
        .order_by("date")[:5]
    )

    context = {
        "title": "Manager Dashboard",
        "manager": manager_profile,
        "today": today,
        "team_members": team_members,
        "total_team_members": total_team_members,
        "present_count": present_count,
        "wfh_count": wfh_count,
        "leave_count": leave_count,
        "late_count": late_count,
        "pending_leaves": pending_leaves,
        "recent_activity": recent_activity,
        "announcements": announcements,
        "birthdays": birthdays,
        "work_anniversaries": work_anniversaries,
        "upcoming_holidays": upcoming_holidays,
    }

    return render(request, "core/manager_dashboard.html", context)


@login_required
@admin_required
def admin_dashboard(request):
    """Real-time Admin Dashboard - Track employees' attendance, leaves, WFH, overtime"""

    if not hasattr(request.user, "company") or not request.user.company:
        # Check if SuperAdmin is spoofing a company
        if request.user.role == User.Role.SUPERADMIN and "active_company_id" in request.session:
            from companies.models import Company

            try:
                company = Company.objects.get(pk=request.session["active_company_id"])
                # Temporarily attach company to user instance for this request
                request.user.company = company
            except Company.DoesNotExist:
                return redirect("superadmin:dashboard")
        else:
            return render(request, "core/dashboard.html", {"title": "Dashboard"})

    from datetime import time as dt_time
    from datetime import timedelta

    from companies.models import Holiday, Location

    today = timezone.localtime().date()
    current_time = timezone.localtime()

    # Get all employees for the company
    company = request.user.company
    location_id = request.GET.get("location")
    employees = Employee.objects.filter(company=company)

    if location_id:
        employees = employees.filter(location_id=location_id)

    # Filter out inactive employees who have exited before today
    employees = employees.filter(Q(is_active=True) | Q(exit_date__gte=today))

    total_employees = employees.count()

    # Today's attendance records
    today_attendance = Attendance.objects.filter(employee__company=company, date=today).select_related(
        "employee", "employee__user"
    )

    if location_id:
        today_attendance = today_attendance.filter(employee__location_id=location_id)

    # Calculate stats
    late_arrivals_list = []
    early_departures_list = []
    on_duty_list = []
    on_time = 0
    work_from_office = 0
    remote_clockins = 0

    # Define office start time (9:00 AM)
    office_start = dt_time(9, 0)

    for att in today_attendance:
        if att.clock_in:
            clock_in_time = timezone.localtime(att.clock_in).time()
            # Late arrivals
            if att.is_late:
                late_arrivals_list.append(att)
            else:
                on_time += 1

            # Work from Office (Present + Half Day)
            if att.status in ["PRESENT", "HALF_DAY"]:
                work_from_office += 1

            # Remote clock-ins
            if att.location_in and att.status == "WFH":
                remote_clockins += 1

        # On Duty (Interpreted as Currently Clocked In / Active based on user request)
        if (att.clock_in and not att.clock_out) or att.status == "ON_DUTY":
            on_duty_list.append(att)

        # Early Departures
        if att.is_early_departure:
            early_departures_list.append(att)

    late_arrivals = len(late_arrivals_list)
    early_departures = len(early_departures_list)
    on_duty_count = len(on_duty_list)

    # --- Department Performance Logic ---
    # Get distinct departments with improved deduplication
    departments_raw = employees.values_list("department", flat=True).distinct()
    departments_map = {}  # Maps normalized name to original names

    # Build mapping of normalized names to original names
    for dept in departments_raw:
        if dept and dept.strip():
            normalized = dept.strip()
            if normalized not in departments_map:
                departments_map[normalized] = []
            departments_map[normalized].append(dept)

    # Get sorted unique departments (normalized)
    departments_list = sorted(departments_map.keys())

    department_performance = []

    for normalized_dept in departments_list:
        # Get all original department names that map to this normalized name
        original_dept_names = departments_map[normalized_dept]

        # Filter employees by any of the original department names
        dept_emps = employees.filter(department__in=original_dept_names)
        dept_total = dept_emps.count()

        # Count present (any positive attendance status) using original names
        dept_present = today_attendance.filter(
            employee__department__in=original_dept_names,
            status__in=["PRESENT", "WFH", "ON_DUTY", "HALF_DAY"],
        ).count()

        percentage = 0
        if dept_total > 0:
            percentage = (dept_present / dept_total) * 100

        department_performance.append(
            {
                "name": normalized_dept,
                "present": dept_present,
                "total": dept_total,
                "percentage": round(percentage, 1),
            }
        )

    # Pending leave requests
    pending_leave_requests = (
        LeaveRequest.objects.filter(employee__company=request.user.company, status="PENDING")
        .select_related("employee", "employee__user")
        .order_by("-created_at")[:10]
    )

    # Recent overtime requests (using leave requests with specific types or create OvertimeRequest model)
    overtime_requests = []

    # Get current month's calendar data for employees
    import calendar as cal

    # Get month/year from params or default to current
    try:
        current_month = int(request.GET.get("month", today.month))
        current_year = int(request.GET.get("year", today.year))
    except ValueError:
        current_month = today.month
        current_year = today.year

    # Calculate Prev/Next Month
    first_day_curr = date(current_year, current_month, 1)

    # Prev Month
    prev_month_date = first_day_curr - timedelta(days=1)
    prev_month = prev_month_date.month
    prev_year = prev_month_date.year

    # Next Month
    if current_month == 12:
        next_month = 1
        next_year = current_year + 1
    else:
        next_month = current_month + 1
        next_year = current_year

    # Get number of days in selected month
    num_days = cal.monthrange(current_year, current_month)[1]

    # Start and End Date for Query
    month_start = date(current_year, current_month, 1)
    month_end = date(current_year, current_month, num_days)

    # Get employees for calendar view (show all active employees)
    calendar_employees = employees

    # Build calendar data for each employee
    employee_calendar_data = []
    for emp in calendar_employees:
        emp_data = {"employee": emp, "days": []}

        # Get attendance for the selected month
        month_attendance = Attendance.objects.filter(employee=emp, date__range=[month_start, month_end])

        att_map = {att.date.day: att for att in month_attendance}

        # Sick Leave Map
        # Find approved SL requests that overlap with this month
        sick_leaves = LeaveRequest.objects.filter(
            employee=emp,
            status="APPROVED",
            leave_type="SL",
            start_date__lte=month_end,
            end_date__gte=month_start,
        )

        # Create a set of dates that are sick leaves
        sick_leave_dates = set()
        for sl in sick_leaves:
            # Intersection of leave range and month range
            s = max(sl.start_date, month_start)
            e = min(sl.end_date, month_end)
            curr = s
            while curr <= e:
                sick_leave_dates.add(curr.day)
                curr += timedelta(days=1)

        for day in range(1, num_days + 1):
            day_date = date(current_year, current_month, day)
            att = att_map.get(day)

            status_class = ""
            if att:
                if att.status == "WFH":
                    status_class = "wfh"
                elif att.status == "WEEKLY_OFF":
                    status_class = "weekly-off"
                elif att.status == "LEAVE":
                    # Check if it is sick leave
                    if day in sick_leave_dates:
                        status_class = "sick-leave"
                    else:
                        status_class = "paid-leave"
                elif att.status == "ABSENT":
                    status_class = "no-attendance"
                elif att.status == "HOLIDAY":
                    status_class = "holiday"
                elif att.status in ["PRESENT", "ON_DUTY", "HALF_DAY"]:
                    status_class = "present"
                else:
                    status_class = "present"  # Default for any other status with clock-in
            else:
                # No attendance record - determine what it should be
                if day_date > today:
                    status_class = "future"
                else:
                    # Check if it's a holiday for this employee's location
                    is_holiday = Holiday.objects.filter(
                        company=emp.company,
                        location=emp.location,
                        date=day_date,
                        is_active=True,
                    ).exists()

                    if is_holiday:
                        status_class = "holiday"
                    # Check if it's a weekly off for this employee
                    elif emp.is_week_off(day_date):
                        status_class = "weekly-off"
                    else:
                        # No record and not holiday/weekoff = absent
                        status_class = "no-attendance"

            emp_data["days"].append({"day": day, "status": status_class, "date": day_date})

        employee_calendar_data.append(emp_data)

    # Get departments and locations for filter
    departments = employees.values_list("department", flat=True).distinct()
    locations = Location.objects.filter(company=company, is_active=True)

    # --- Announcements Data (Next 30 Days) ---
    future_date = today + timedelta(days=30)

    # 1. Upcoming Birthdays
    upcoming_birthdays = []

    for emp in employees:
        if emp.dob:
            # Create birthday for current year
            try:
                this_year_bday = emp.dob.replace(year=today.year)
            except ValueError:
                # Leap year edge case (Feb 29 on non-leap year -> Feb 28 or Mar 1)
                this_year_bday = emp.dob.replace(year=today.year, day=28)

            # If birthday passed this year, check next year
            if this_year_bday < today:
                try:
                    next_birthday = emp.dob.replace(year=today.year + 1)
                except ValueError:
                    next_birthday = emp.dob.replace(year=today.year + 1, day=28)
            else:
                next_birthday = this_year_bday

            if today <= next_birthday <= future_date:
                days_left = (next_birthday - today).days
                upcoming_birthdays.append(
                    {
                        "employee": emp,
                        "date": next_birthday,
                        "display_date": next_birthday,
                        "is_today": days_left == 0,
                        "days_left": days_left,
                    }
                )

    # Sort by nearest date
    upcoming_birthdays.sort(key=lambda x: x["days_left"])

    # 2. Work Anniversaries
    upcoming_anniversaries = []
    for emp in employees:
        if emp.date_of_joining:
            # Calculate years completed
            years_completed = today.year - emp.date_of_joining.year

            # Anniv for current year
            try:
                this_year_anniv = emp.date_of_joining.replace(year=today.year)
            except ValueError:
                this_year_anniv = emp.date_of_joining.replace(year=today.year, day=28)

            if this_year_anniv < today:
                try:
                    next_anniv = emp.date_of_joining.replace(year=today.year + 1)
                    years_completed += 1  # It will be next year's anniversary
                except ValueError:
                    next_anniv = emp.date_of_joining.replace(year=today.year + 1, day=28)
                    years_completed += 1
            else:
                next_anniv = this_year_anniv

            if today <= next_anniv <= future_date and years_completed > 0:
                days_left = (next_anniv - today).days
                upcoming_anniversaries.append(
                    {
                        "employee": emp,
                        "date": next_anniv,
                        "years": years_completed,
                        "is_today": days_left == 0,
                        "days_left": days_left,
                    }
                )

    upcoming_anniversaries.sort(key=lambda x: x["days_left"])

    # 3. Announcements
    from companies.models import Announcement

    announcements = (
        Announcement.objects.filter(company=request.user.company, is_active=True)
        .select_related("location")
        .order_by("-created_at")[:10]
    )

    # 4. Upcoming Holidays
    # Using the Holiday model directly
    upcoming_holidays_qs = (
        Holiday.objects.filter(
            company=request.user.company,
            date__gte=today,
            date__lte=future_date,
            is_active=True,
        )
        .select_related("location")
        .order_by("date")
    )

    # Context updates

    context = {
        "title": "Admin Dashboard",
        "today": today,
        "total_employees": total_employees,
        "late_arrivals": late_arrivals,
        "early_departures": early_departures,
        "on_duty_count": on_duty_count,
        "late_arrivals_list": late_arrivals_list,
        "early_departures_list": early_departures_list,
        "on_duty_list": on_duty_list,
        "department_performance": department_performance,
        "on_time": on_time,
        "work_from_office": work_from_office,
        "remote_clockins": remote_clockins,
        "pending_leave_requests": pending_leave_requests,
        "announcements": announcements,
        "upcoming_birthdays": upcoming_birthdays,
        "upcoming_anniversaries": upcoming_anniversaries,
        "upcoming_holidays": upcoming_holidays_qs,
        "employee_calendar_data": employee_calendar_data,
        "current_month": cal.month_name[current_month],
        "current_year": current_year,
        "prev_month": prev_month,
        "prev_year": prev_year,
        "next_month": next_month,
        "next_year": next_year,
        "num_days": num_days,
        "departments": departments,
        "locations": locations,
        "location_filter": int(location_id) if location_id and location_id.isdigit() else None,
    }

    # Check for Admin's own celebration
    try:
        admin_emp = request.user.employee_profile
        celebration_type = None

        # Birthday Check
        if admin_emp.dob:
            try:
                # Handle leap years
                bday_this_year = admin_emp.dob.replace(year=today.year)
            except ValueError:
                bday_this_year = admin_emp.dob.replace(year=today.year, day=28)

            if bday_this_year == today:
                celebration_type = "birthday"

        # Anniversary Check
        if admin_emp.date_of_joining:
            try:
                anniv_this_year = admin_emp.date_of_joining.replace(year=today.year)
            except ValueError:
                anniv_this_year = admin_emp.date_of_joining.replace(year=today.year, day=28)

            if anniv_this_year == today and today.year > admin_emp.date_of_joining.year:
                celebration_type = "anniversary" if not celebration_type else "both"

        if celebration_type:
            context["celebration_type"] = celebration_type
            context["years_service"] = today.year - admin_emp.date_of_joining.year if admin_emp.date_of_joining else 0

    except Exception as e:
        logger.debug("Error checking celebration dates for admin dashboard", error=str(e))

    return render(request, "core/admin_dashboard.html", context)


@login_required
@admin_required
def search_employees_api(request):
    """API endpoint for searching employees by name"""
    from django.http import JsonResponse
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        query = request.GET.get("q", "").strip()
        logger.info(f"Search query received: '{query}' from user: {request.user.email}")

        # Search with just 1 character (like autocomplete)
        if not query or len(query) < 1:
            logger.info("Query too short, returning empty results")
            return JsonResponse({"employees": [], "debug": "Query too short"})

        # Get company
        company = request.user.company
        logger.info(f"Searching in company: {company.name if company else 'None'}")

        if not company:
            logger.error("User has no company assigned")
            return JsonResponse({"employees": [], "error": "No company assigned"})

        # Search employees by name (first name or last name)
        employees = (
            Employee.objects.filter(company=company)
            .filter(
                Q(user__first_name__icontains=query) | Q(user__last_name__icontains=query) | Q(badge_id__icontains=query)
            )
            .select_related("user", "location", "manager")[:10]
        )

        logger.info(f"Found {employees.count()} employees matching query")

        results = []
        for emp in employees:
            result = {
                "id": emp.id,
                "name": emp.user.get_full_name() or "No Name",
                "employee_id": emp.badge_id or f"EMP-{emp.id}",
                "department": emp.department or "N/A",
                "location": emp.location.name if emp.location else "N/A",
                "designation": emp.designation or "N/A",
                "email": emp.user.email,
                "phone": emp.mobile_number or "N/A",
                "profile_url": f"/employees/{emp.id}/detail/",
            }
            results.append(result)
            logger.info(f"Added employee: {result['name']}")

        logger.info(f"Returning {len(results)} results")
        return JsonResponse({"employees": results, "count": len(results)})
        
    except Exception as e:
        logger.error(f"Error in search_employees_api: {str(e)}")
        return JsonResponse({"employees": [], "error": str(e)}, status=500)


@login_required
def employee_dashboard(request):
    """Employee Personal Dashboard - Their own attendance, leaves, stats"""
    employee = safe_get_employee_profile(request.user)
    if not employee:
        messages.error(request, "Employee profile not found.")
        return redirect("personal_home")

    today = timezone.localtime().date()

    # Today's attendance
    attendance = Attendance.objects.filter(employee=employee, date=today).first()

    # --- Comprehensive Attendance History (Last 30 Days) ---
    end_date = today
    start_date = today - timedelta(days=30)
    
    # Fetch existing records
    attendance_records = {att.date: att for att in Attendance.objects.filter(employee=employee, date__range=[start_date, end_date])}
    
    # Fetch leaves
    leaves = LeaveRequest.objects.filter(employee=employee, status='APPROVED', start_date__lte=end_date, end_date__gte=start_date)
    leave_dates = {}
    for l in leaves:
        curr = max(l.start_date, start_date)
        while curr <= min(l.end_date, end_date):
            leave_dates[curr] = "LEAVE"
            curr += timedelta(days=1)
            
    # Fetch Holidays
    from companies.models import Holiday
    holiday_q = Q(location__isnull=True)
    if employee.location:
        holiday_q |= Q(location=employee.location)
    holidays = Holiday.objects.filter(company=employee.company, date__range=[start_date, end_date], is_active=True).filter(holiday_q)
    holiday_dates = {h.date: h.name for h in holidays}
    
    history = []
    curr_date = end_date
    while curr_date >= start_date:
        if curr_date in attendance_records:
            history.append(attendance_records[curr_date])
        else:
            if employee.is_week_off(curr_date):
                status = "WEEKLY_OFF"
            elif curr_date in holiday_dates:
                status = "HOLIDAY"
            elif curr_date in leave_dates:
                status = "LEAVE"
            elif curr_date == today:
                status = "NOT_LOGGED_IN"
            else:
                status = "MISSED"
                
            history.append({
                'date': curr_date,
                'status': status,
                'status_display': status.replace('_', ' ').title(),
                'clock_in': None,
                'clock_out': None,
                'effective_hours': "-",
                'id': None
            })
        curr_date -= timedelta(days=1)

    # Calculate stats from history
    total_days = len(history)
    present_days = 0
    wfh_days = 0
    leave_days = 0
    absent_days = 0
    total_seconds = 0
    sessions_with_time = 0

    for item in history:
        if isinstance(item, Attendance):
            status = item.status
            if item.clock_in and item.clock_out:
                total_seconds += (item.clock_out - item.clock_in).total_seconds()
                sessions_with_time += 1
        else:
            status = item.get('status')

        if status == 'PRESENT':
            present_days += 1
        elif status == 'WFH':
            wfh_days += 1
        elif status == 'LEAVE':
            leave_days += 1
        elif status in ['ABSENT', 'MISSED']:
            absent_days += 1

    avg_hours = "00:00"
    if sessions_with_time > 0:
        avg_sec = total_seconds / sessions_with_time
        h = int(avg_sec // 3600)
        m = int((avg_sec % 3600) // 60)
        avg_hours = f"{h:02d}:{m:02d}"

    # Weekly stats (current week - Monday to Sunday)
    week_start = today - timedelta(days=today.weekday())  # Monday
    week_end = week_start + timedelta(days=6)  # Sunday

    # Filter history for this week
    week_history = [item for item in history if week_start <= (item.date if isinstance(item, Attendance) else item['date']) <= week_end]
    
    week_present = 0
    week_wfh = 0
    week_leave = 0
    week_absent = 0
    week_total = 0 # Expected work days

    for item in week_history:
        date = item.date if isinstance(item, Attendance) else item['date']
        status = item.status if isinstance(item, Attendance) else item['status']
        
        if status == 'PRESENT':
            week_present += 1
            week_total += 1
        elif status == 'WFH':
            week_wfh += 1
            week_total += 1
        elif status == 'LEAVE':
            week_leave += 1
            week_total += 1
        elif status in ['ABSENT', 'MISSED']:
            week_absent += 1
            week_total += 1
        elif status in ['WEEKLY_OFF', 'HOLIDAY']:
            pass # Don't count in total expected work days

    week_attendance_rate = 0
    if week_total > 0:
        week_attendance_rate = round(((week_present + week_wfh) / week_total) * 100)

    # Yearly stats - Simplified based on existing records + estimate
    year_start = today.replace(month=1, day=1)
    year_attendance = Attendance.objects.filter(employee=employee, date__gte=year_start, date__lte=today)

    year_present = year_attendance.filter(status="PRESENT").count()
    year_wfh = year_attendance.filter(status="WFH").count()
    year_leave = year_attendance.filter(status="LEAVE").count()
    
    # Estimate total working days so far (excluding weekends)
    total_days_so_far = (today - year_start).days + 1
    # Very rough estimate: 5/7 of days are working days
    # Better: count actual non-week-off days if possible, but for now let's keep it simple or use database
    year_total = year_attendance.exclude(status__in=["WEEKLY_OFF", "HOLIDAY"]).count()
    
    # If year_total is less than historical records, it's definitely undercounting missed days
    # But without full history for year, we can't be precise. 
    # Let's at least use what we have.

    year_attendance_rate = 0
    if year_total > 0:
        year_attendance_rate = round(((year_present + year_wfh) / year_total) * 100)


    # Leave balance
    leave_balance = getattr(employee, "leave_balance", None)
    if leave_balance:
        # Force refresh from database to get latest data
        leave_balance.refresh_from_db()

    # Recent leave requests
    recent_leave_requests = LeaveRequest.objects.filter(employee=employee).order_by("-created_at")[:5]

    # --- Announcements & Celebrations Data ---

    from companies.models import Announcement, Holiday

    # Get all company employees for celebrations
    company_employees = Employee.objects.filter(company=employee.company, is_active=True)

    # 1. Announcements
    announcements = Announcement.objects.filter(company=employee.company, is_active=True).order_by("-created_at")[:5]

    # 2. Upcoming Birthdays & Anniversaries
    future_date = today + timedelta(days=30)
    upcoming_birthdays = []
    upcoming_anniversaries = []

    for emp in company_employees:
        # Birthday
        if emp.dob:
            try:
                this_year_bday = emp.dob.replace(year=today.year)
            except ValueError:
                this_year_bday = emp.dob.replace(year=today.year, day=28)

            if this_year_bday < today:
                try:
                    next_birthday = emp.dob.replace(year=today.year + 1)
                except ValueError:
                    next_birthday = emp.dob.replace(year=today.year + 1, day=28)
            else:
                next_birthday = this_year_bday

            if today <= next_birthday <= future_date:
                days_left = (next_birthday - today).days
                upcoming_birthdays.append(
                    {
                        "employee": emp,
                        "date": next_birthday,
                        "display_date": next_birthday,
                        "is_today": days_left == 0,
                        "days_left": days_left,
                    }
                )

        # Work Anniversary
        if emp.date_of_joining:
            years_completed = today.year - emp.date_of_joining.year
            try:
                this_year_anniv = emp.date_of_joining.replace(year=today.year)
            except ValueError:
                this_year_anniv = emp.date_of_joining.replace(year=today.year, day=28)

            if this_year_anniv < today:
                try:
                    next_anniv = emp.date_of_joining.replace(year=today.year + 1)
                    years_completed += 1
                except ValueError:
                    next_anniv = emp.date_of_joining.replace(year=today.year + 1, day=28)
                    years_completed += 1
            else:
                next_anniv = this_year_anniv

            if today <= next_anniv <= future_date and years_completed > 0:
                days_left = (next_anniv - today).days
                upcoming_anniversaries.append(
                    {
                        "employee": emp,
                        "date": next_anniv,
                        "years": years_completed,
                        "is_today": days_left == 0,
                        "days_left": days_left,
                    }
                )

    upcoming_birthdays.sort(key=lambda x: x["days_left"])
    upcoming_anniversaries.sort(key=lambda x: x["days_left"])

    # 3. Upcoming Holidays
    upcoming_holidays = (
        Holiday.objects.filter(
            company=employee.company,
            date__gte=today,
            date__lte=future_date,
            is_active=True,
        )
        .select_related("location")
        .order_by("date")
    )

    context = {
        "title": "My Dashboard",
        "employee": employee,
        "attendance": attendance,
        "today": today,
        "avg_hours": avg_hours,
        "total_days": total_days,
        "present_days": present_days,
        "wfh_days": wfh_days,
        "leave_days": leave_days,
        # Weekly stats
        "week_present": week_present,
        "week_wfh": week_wfh,
        "week_leave": week_leave,
        "week_absent": week_absent,
        "week_total": week_total,
        "week_attendance_rate": week_attendance_rate,
        "week_start_date": week_start,
        "week_end_date": week_end,
        # Yearly stats
        "year_present": year_present,
        "year_wfh": year_wfh,
        "year_leave": year_leave,
        "year_total": year_total,
        "year_attendance_rate": year_attendance_rate,
        "year_start_date": year_start,
        # Other data
        "leave_balance": leave_balance,
        "recent_leave_requests": recent_leave_requests,
        "attendance_history": history,
        "announcements": announcements,
        "upcoming_birthdays": upcoming_birthdays,
        "upcoming_anniversaries": upcoming_anniversaries,
        "upcoming_holidays": upcoming_holidays,
    }

    # Check for celebrations
    celebration_type = None
    if employee.dob:
        try:
            # Handle leap years
            bday_this_year = employee.dob.replace(year=today.year)
        except ValueError:
            bday_this_year = employee.dob.replace(year=today.year, day=28)

        if bday_this_year == today:
            celebration_type = "birthday"

    if employee.date_of_joining:
        try:
            anniv_this_year = employee.date_of_joining.replace(year=today.year)
        except ValueError:
            anniv_this_year = employee.date_of_joining.replace(year=today.year, day=28)

        if anniv_this_year == today and today.year > employee.date_of_joining.year:
            celebration_type = "anniversary" if not celebration_type else "both"

    context["celebration_type"] = celebration_type
    context["years_service"] = today.year - employee.date_of_joining.year if employee.date_of_joining else 0

    return render(request, "core/employee_dashboard.html", context)


@login_required
def personal_home(request):
    from companies.models import Announcement
    context = {}
    if hasattr(request.user, "employee_profile"):
        employee = request.user.employee_profile
        today = timezone.localdate()

        # Today's attendance
        attendance = Attendance.objects.filter(employee=employee, date=today).first()
        context["attendance"] = attendance

        # --- Comprehensive Attendance History (Last 30 Days) ---
        end_date = today
        start_date = today - timedelta(days=30)
        
        # Fetch existing records
        attendance_records = {att.date: att for att in Attendance.objects.filter(employee=employee, date__range=[start_date, end_date])}
        
        # Fetch leaves
        leaves = LeaveRequest.objects.filter(employee=employee, status='APPROVED', start_date__lte=end_date, end_date__gte=start_date)
        leave_dates = {}
        for l in leaves:
            curr = max(l.start_date, start_date)
            while curr <= min(l.end_date, end_date):
                leave_dates[curr] = "LEAVE"
                curr += timedelta(days=1)
                
        # Fetch Holidays
        holiday_q = Q(location__isnull=True)
        if employee.location:
            holiday_q |= Q(location=employee.location)
        holidays = Holiday.objects.filter(company=employee.company, date__range=[start_date, end_date], is_active=True).filter(holiday_q)
        holiday_dates = {h.date: h.name for h in holidays}
        
        history = []
        curr_date = end_date
        while curr_date >= start_date:
            if curr_date in attendance_records:
                history.append(attendance_records[curr_date])
            else:
                if employee.is_week_off(curr_date):
                    status = "WEEKLY_OFF"
                elif curr_date in holiday_dates:
                    status = "HOLIDAY"
                elif curr_date in leave_dates:
                    status = "LEAVE"
                elif curr_date == today:
                    status = "NOT_LOGGED_IN"
                else:
                    status = "MISSED"
                    
                history.append({
                    'date': curr_date,
                    'status': status,
                    'status_display': status.replace('_', ' ').title(),
                    'clock_in': None,
                    'clock_out': None,
                    'effective_hours': "-",
                    'id': None
                })
            curr_date -= timedelta(days=1)

        # Calculate stats from history
        total_seconds = 0
        sessions_with_time = 0
        for att in [item for item in history if isinstance(item, Attendance)]:
            if att.clock_in and att.clock_out:
                total_seconds += (att.clock_out - att.clock_in).total_seconds()
                sessions_with_time += 1

        avg_hours = "00:00"
        if sessions_with_time > 0:
            avg_sec = total_seconds / sessions_with_time
            h = int(avg_sec // 3600)
            m = int((avg_sec % 3600) // 60)
            avg_hours = f"{h:02d}:{m:02d}"

        context["avg_hours"] = avg_hours
        context["on_time_percentage"] = "100%"  # Stub for now

        context["attendance_history"] = history


        # Announcements - current month


        announcements = (
            Announcement.objects.filter(company=employee.company, is_active=True)
            .filter(Q(location__isnull=True) | Q(location=employee.location))
            .order_by("-created_at")[:5]
        )
        context["announcements"] = announcements

        # Current month start and end dates
        current_month = today.month
        current_year = today.year
        from calendar import monthrange

        _, last_day = monthrange(current_year, current_month)
        month_start = today.replace(day=1)
        month_end = today.replace(day=last_day)

        # Celebrations - Birthdays this month (all dates in current month)
        company_employees = Employee.objects.filter(company=employee.company)
        birthdays = company_employees.filter(dob__month=current_month).order_by("dob__day")
        context["birthdays"] = birthdays

        # Today's specific celebrations
        context["today"] = today
        context["todays_birthdays"] = company_employees.filter(dob__month=today.month, dob__day=today.day)
        context["todays_anniversaries"] = company_employees.filter(
            date_of_joining__month=today.month, date_of_joining__day=today.day
        ).exclude(date_of_joining__year=today.year)

        # Work Anniversaries this month (all dates in current month)
        work_anniversaries = (
            company_employees.filter(date_of_joining__month=current_month)
            .exclude(date_of_joining__year=current_year)
            .order_by("date_of_joining__day")
        )
        context["work_anniversaries"] = work_anniversaries

        # Holidays - All holidays in current month (past and upcoming)

        upcoming_holidays = (
            Holiday.objects.filter(
                company=employee.company,
                date__gte=month_start,
                date__lte=month_end,
                is_active=True,
            )
            .filter(Q(location__isnull=True) | Q(location=employee.location))
            .order_by("date")
        )
        context["upcoming_holidays"] = upcoming_holidays

        # Shift Timings & Timeline Data
        if employee.assigned_shift:
            context["assigned_shift"] = employee.assigned_shift
            shift = employee.assigned_shift

            # Grace Usage Stats
            grace_used_count = Attendance.objects.filter(
                employee=employee,
                date__month=today.month,
                date__year=today.year,
                is_grace_used=True,
            ).count()
            context["grace_used_count"] = grace_used_count
            context["late_logins_remaining"] = max(0, employee.assigned_shift.allowed_late_logins - grace_used_count)

            # Timeline Calculations
            # Define Start and End (in minutes from midnight)
            def to_minutes(t):
                return t.hour * 60 + t.minute

            shift_start_min = to_minutes(shift.start_time)
            shift_end_min = to_minutes(shift.end_time)
            total_duration = shift_end_min - shift_start_min
            if total_duration <= 0:
                total_duration += 24 * 60  # Overnight shift

            timeline_items = []

            from employees.models import AttendanceSession

            sessions = AttendanceSession.objects.filter(employee=employee, date=today).order_by("session_number")

            if sessions.exists():
                session_list = list(sessions)
                for i, session in enumerate(session_list):
                    is_first = i == 0
                    is_last_recorded = i == len(session_list) - 1

                    # Clock In Node
                    if session.clock_in:
                        login_time = timezone.localtime(session.clock_in).time()
                        login_min = to_minutes(login_time)
                        offset = login_min - shift_start_min
                        if offset < 0:
                            offset += 24 * 60
                        percent = (offset / total_duration) * 100
                        percent = max(0, min(percent, 100))

                        # Dot class logic: 1st clock-in shows time and session type
                        if is_first:
                            dot_class = "web" if session.session_type == "WEB" else "remote"
                            show_time = True
                        else:
                            dot_class = "clock-in-dot"
                            show_time = False

                        timeline_items.append(
                            {
                                "type": "login",
                                "time": login_time,
                                "label": f"Login {session.session_number}",
                                "percent": percent,
                                "dot_class": dot_class,
                                "show_time": show_time,
                                "is_late": attendance.is_late if is_first else False,
                            }
                        )

                    # Clock Out Node
                    if session.clock_out:
                        logout_time = timezone.localtime(session.clock_out).time()
                        logout_min = to_minutes(logout_time)
                        offset = logout_min - shift_start_min
                        if offset < 0:
                            offset += 24 * 60
                        percent = (offset / total_duration) * 100
                        percent = max(0, min(percent, 100))

                        # Dot class logic: Last clock-out shows time
                        # Special Case: Only show time for the VERY LAST clockout of the day if session_count == max_sessions
                        # or if it's the last one recorded and they are not clocked in.
                        is_clocked_in = attendance.is_currently_clocked_in if attendance else False

                        if is_last_recorded and not is_clocked_in:
                            dot_class = "logout"
                            show_time = True
                        else:
                            dot_class = "clock-out-dot"
                            show_time = False

                        timeline_items.append(
                            {
                                "type": "logout",
                                "time": logout_time,
                                "label": f"Logout {session.session_number}",
                                "percent": percent,
                                "dot_class": dot_class,
                                "show_time": show_time,
                                "is_early": attendance.is_early_departure if is_last_recorded else False,
                            }
                        )

                # If currently clocked in, or not reached max sessions,
                # we don't necessarily need to add the hollow End node if we have recorded sessions,
                # but it helps to fill the bar.
                if not sessions.filter(
                    clock_out__isnull=False,
                    session_number=attendance.max_daily_sessions if attendance else 3,
                ).exists():
                    if not timeline_items or timeline_items[-1]["percent"] < 100:
                        timeline_items.append(
                            {
                                "type": "logout",
                                "time": shift.end_time,
                                "label": "End",
                                "percent": 100,
                                "dot_class": "hollow",
                                "show_time": True,
                            }
                        )

            else:
                # No sessions yet
                timeline_items.append(
                    {
                        "type": "login",
                        "time": shift.start_time,
                        "label": "Start",
                        "percent": 0,
                        "dot_class": "hollow",
                        "show_time": True,
                    }
                )
                timeline_items.append(
                    {
                        "type": "logout",
                        "time": shift.end_time,
                        "label": "End",
                        "percent": 100,
                        "dot_class": "hollow",
                        "show_time": True,
                    }
                )

            context["timeline_items"] = timeline_items

        else:
            # Fallback
            from companies.models import ShiftTiming

            shift_timing, _ = ShiftTiming.objects.get_or_create(company=employee.company)
            context["shift_timing"] = shift_timing
            context["timeline_items"] = []  # Empty for default

        # Add leave balance to context
        try:
            leave_balance = employee.leave_balance
            leave_balance.refresh_from_db()  # Force refresh to get latest data
            context["leave_balance"] = leave_balance
        except Exception:
            # Create leave balance if it doesn't exist
            from employees.models import LeaveBalance
            leave_balance = LeaveBalance.objects.create(
                employee=employee,
                casual_leave_allocated=0.0,
                sick_leave_allocated=0.0
            )
            context["leave_balance"] = leave_balance

    return render(request, "core/personal_home.html", context)


# --- Me Section Stubs ---


@login_required
def my_profile(request):
    return redirect("employee_profile")


@login_required
def my_leaves(request):
    try:
        employee = request.user.employee_profile
        # Force refresh from database to get latest data
        employee.refresh_from_db()

    except Exception:
        # Graceful fallback or auto-create (reusing logic from profile view might be better)
        messages.error(request, "Employee profile not found.")
        return redirect("personal_home")

    # Get or create balance (accrual handled by command, but ensure existence)
    balance, created = LeaveBalance.objects.get_or_create(employee=employee)
    
    # Force refresh balance from database to get latest data after bulk upload
    balance.refresh_from_db()

    if request.method == "POST":
        leave_type = request.POST.get("leave_type")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        reason = request.POST.get("reason")
        duration = request.POST.get("duration", "FULL")  # Get duration from form, default to FULL

        # Basic Validation
        if not (leave_type and start_date and end_date):
            messages.error(request, "All fields are required.")
            return redirect("my_leaves")

        try:
            # Parse dates to ensure validity
            try:
                # Handle potentially different date formats if needed, but YYYY-MM-DD is standard for input type=date
                s_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
                e_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format used.")
                return redirect("my_leaves")

            if e_dt < s_dt:
                messages.error(request, "End date cannot be before start date.")
                return redirect("my_leaves")

            # Validate duration for single-day leaves
            is_single_day = s_dt == e_dt
            if not is_single_day and duration in ["FIRST_HALF", "SECOND_HALF"]:
                messages.error(request, "Half-day options are only available for single-day leaves.")
                return redirect("my_leaves")

            # Create Leave Request
            leave_request = LeaveRequest.objects.create(
                employee=employee,
                leave_type=leave_type,
                start_date=start_date,
                end_date=end_date,
                reason=reason,
                duration=duration,  # Use the duration from form
                status="PENDING",
            )

            # Send Email Notifications
            try:
                from core.email_utils import send_leave_request_notification

                result = send_leave_request_notification(leave_request)

                if not result.get("hr", False):
                    messages.warning(
                        request,
                        "Leave request submitted, but email notification to HR failed. Please notify HR manually.",
                    )

            except Exception as mail_err:
                import logging

                logger = logging.getLogger(__name__)
                logger.error(f"Error sending email: {mail_err}")
                messages.warning(
                    request,
                    "Leave request submitted, but email notification system encountered an error.",
                )

            messages.success(request, "Leave request submitted successfully.")

        except Exception as e:
            messages.error(request, f"Error submitting request: {str(e)}")
            # Log the error for admin debug
            logger.error(f"Leave Application Error: {e}")

        return redirect("my_leaves")

    recent_requests = LeaveRequest.objects.filter(employee=employee).order_by("-created_at")[:5]

    return render(
        request,
        "employees/leave_dashboard.html",
        {"title": "My Leaves", "balance": balance, "recent_requests": recent_requests},
    )


@login_required
def cancel_leave_request(request, pk):
    from django.shortcuts import get_object_or_404

    leave_request = get_object_or_404(LeaveRequest, pk=pk)

    # Ensure the user owns this request
    if leave_request.employee.user != request.user:
        messages.error(request, "You are not authorized to cancel this request.")
        return redirect("my_leaves")

    if leave_request.status == "PENDING":
        leave_request.delete()
        messages.success(request, "Leave request cancelled successfully.")
    else:
        messages.error(request, "Only pending leave requests can be cancelled.")

    return redirect("my_leaves")


@login_required
def my_finance(request):
    try:
        employee = request.user.employee_profile

    except Exception:
        messages.error(request, "Employee profile not found.")

        return redirect("personal_home")

    payslips = Payslip.objects.filter(employee=employee).order_by("-month")

    return render(
        request,
        "employees/finance_dashboard.html",
        {"title": "My Finance", "employee": employee, "payslips": payslips},
    )


@login_required
def employee_holidays(request):
    """Employee view to see company holidays filtered by their location"""

    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "Company information not found.")

        return redirect("personal_home")

    # Get employee profile to access location
    employee = safe_get_employee_profile(request.user)
    if not employee:
        messages.error(request, "Employee profile not found.")

        return redirect("personal_home")

    # Get current year or selected year

    current_year = timezone.now().year

    year_filter = int(request.GET.get("year", current_year))

    # Get holidays for the company and year

    # Filter by employee's location OR global holidays

    query_filter = Q(location__name__iexact="Global") | Q(location__name__iexact="All Locations")

    if employee.location:
        query_filter |= Q(location=employee.location)

    holidays = (
        Holiday.objects.filter(company=request.user.company, year=year_filter, is_active=True)
        .filter(query_filter)
        .order_by("date")
    )

    # Get available years

    years = (
        Holiday.objects.filter(company=request.user.company).values_list("year", flat=True).distinct().order_by("-year")
    )

    # Group holidays by month

    holidays_by_month = {}

    for holiday in holidays:
        month_name = holiday.date.strftime("%B %Y")

        if month_name not in holidays_by_month:
            holidays_by_month[month_name] = []

        holidays_by_month[month_name].append(holiday)

    # Statistics

    total_holidays = holidays.count()

    mandatory_count = holidays.filter(holiday_type="MANDATORY").count()

    optional_count = holidays.filter(holiday_type="OPTIONAL").count()

    # Upcoming holidays

    today = timezone.now().date()

    upcoming_holidays = holidays.filter(date__gte=today)[:5]

    return render(
        request,
        "core/employee_holidays.html",
        {
            "title": "Holiday Calendar",
            "holidays": holidays,
            "holidays_by_month": holidays_by_month,
            "years": years,
            "year_filter": year_filter,
            "total_holidays": total_holidays,
            "mandatory_count": mandatory_count,
            "optional_count": optional_count,
            "upcoming_holidays": upcoming_holidays,
            "employee_location": employee.location.name if employee.location else "Not Assigned",
        },
    )


@login_required
def handbook(request):
    try:
        # Auto-initialize default sections if empty (for demo)

        if not HandbookSection.objects.exists():
            HandbookSection.objects.create(
                title="Company Rules",
                order=1,
                content="<h4>Working Hours</h4><p>Standard working hours are 9:00 AM to 6:00 PM.</p>",
            )

            HandbookSection.objects.create(
                title="Conduct",
                order=2,
                content="<h4>Code of Conduct</h4><p>Employees are expected to maintain professionalism.</p>",
            )

        sections = HandbookSection.objects.filter(is_active=True)

        return render(
            request,
            "core/handbook.html",
            {"title": "Employee Handbook", "sections": sections},
        )

    except Exception as e:
        messages.error(request, f"Error loading handbook: {e}")

        return redirect("personal_home")


@login_required
def policy(request):
    try:
        # Auto-initialize default sections if empty

        if not PolicySection.objects.exists():
            PolicySection.objects.create(
                title="Leave Policy",
                order=1,
                content="<h4>Annual Leave</h4><p>Employees are entitled to 25 days of annual leave.</p>",
            )

            PolicySection.objects.create(
                title="HR Policy",
                order=2,
                content="<h4>Recruitment</h4><p>We are an equal opportunity employer.</p>",
            )

        sections = PolicySection.objects.filter(is_active=True)

        return render(
            request,
            "core/policy.html",
            {"title": "Company Policy", "sections": sections},
        )

    except Exception as e:
        messages.error(request, f"Error loading policy: {e}")

        return redirect("personal_home")


# --- Employees Section Stubs ---


@login_required
def org_chart(request):
    # Ensure user has a company
    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "You are not linked to any company.")
        return redirect("dashboard")

    # Fetch only ACTIVE employees for this company
    employees = Employee.objects.filter(
        company=request.user.company, employment_status="ACTIVE", is_active=True
    ).select_related("user", "manager")

    # Helper to detect cycles
    def creates_cycle(parent_node, child_node):
        """Check if adding child_node to parent_node would create a cycle"""
        queue = [child_node]
        visited = {child_node["id"]}
        while queue:
            curr = queue.pop(0)
            if curr["id"] == parent_node["id"]:
                return True
            for grandchild in curr["direct_reports"]:
                if grandchild["id"] not in visited:
                    visited.add(grandchild["id"])
                    queue.append(grandchild)
        return False

    try:
        # Build dictionary Key=USER_ID (not Employee ID)
        nodes = {}
        for emp in employees:
            if emp.user:
                nodes[emp.user.id] = {
                    "id": emp.user.id,
                    "employee": emp,
                    "user": emp.user,
                    "direct_reports": [],
                    "is_superadmin": False,
                }

        roots = []
        for emp in employees:
            if not emp.user:
                continue

            current_node = nodes[emp.user.id]
            manager_user = emp.manager  # User object

            if manager_user and manager_user.id != emp.user.id:
                # Case A: Manager is in the company (exists in nodes)
                if manager_user.id in nodes:
                    manager_node = nodes[manager_user.id]
                    if not creates_cycle(manager_node, current_node):
                        if current_node not in manager_node["direct_reports"]:
                            manager_node["direct_reports"].append(current_node)
                    else:
                        if current_node not in roots:
                            roots.append(current_node)
                else:
                    # Case B: Manager is External (SuperAdmin)
                    if manager_user.role == User.Role.SUPERADMIN:
                        if manager_user.id not in nodes:
                            nodes[manager_user.id] = {
                                "id": manager_user.id,
                                "employee": None,
                                "user": manager_user,
                                "direct_reports": [],
                                "is_superadmin": True,
                            }
                            roots.append(nodes[manager_user.id])

                        manager_node = nodes[manager_user.id]
                        if not creates_cycle(manager_node, current_node):
                            if current_node not in manager_node["direct_reports"]:
                                manager_node["direct_reports"].append(current_node)
                    else:
                        if current_node not in roots:
                            roots.append(current_node)
            else:
                if current_node not in roots:
                    roots.append(current_node)

        # Filter roots to ensure no children
        child_ids = set()
        for uid, node in nodes.items():
            for child in node["direct_reports"]:
                child_ids.add(child["id"])

        final_roots = [node for uid, node in nodes.items() if uid not in child_ids]

        return render(
            request,
            "core/org_chart.html",
            {
                "title": "Organisation Chart",
                "roots": final_roots,
                "company": request.user.company,
            },
        )
    except Exception as e:
        import traceback

        logger.error(f"Admin Org Chart Error: {str(e)}\n{traceback.format_exc()}")
        return render(
            request,
            "core/org_chart.html",
            {
                "title": "Organisation Chart",
                "roots": [],
                "company": request.user.company,
                "error": "Could not generate hierarchical view. Please check reporting structures.",
            },
        )


@login_required
def employee_org_chart(request):
    """Employee-facing org chart view - same as org_chart but accessible to all employees"""
    # Ensure user has a company
    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "You are not linked to any company.")
        return redirect("dashboard")

    # Fetch only ACTIVE employees for this company
    all_employees = Employee.objects.filter(
        company=request.user.company, employment_status="ACTIVE", is_active=True
    ).select_related("user", "manager")

    # Filter employees based on Role
    final_employees = []

    # Helper structure for subtree logic
    emp_map = {e.user.id: e for e in all_employees}
    subordinates_map = {}
    for emp in all_employees:
        if emp.manager_id:
            if emp.manager_id not in subordinates_map:
                subordinates_map[emp.manager_id] = []
            subordinates_map[emp.manager_id].append(emp)

    def get_subtree(user_id, visited=None):
        """Recursively get all subordinates with cycle protection"""
        if visited is None:
            visited = set()

        if user_id in visited:
            return []
        visited.add(user_id)

        result = []
        if user_id in subordinates_map:
            for sub in subordinates_map[user_id]:
                result.append(sub)
                result.extend(get_subtree(sub.user.id, visited))
        return result

    # Determine context
    current_emp = getattr(request.user, "employee_profile", None)
    role = getattr(request.user, "role", None)

    if current_emp:
        if role == User.Role.EMPLOYEE:
            # Employee View: Department + Reporting Manager
            dept = current_emp.department
            manager_id = current_emp.manager_id

            added_ids = set()

            # 1. Add Department Colleagues
            for emp in all_employees:
                if emp.department == dept:
                    final_employees.append(emp)
                    added_ids.add(emp.id)

            # 2. Add Reporting Manager (if not already added)
            if manager_id and manager_id in emp_map:
                mgr = emp_map[manager_id]
                if mgr.id not in added_ids:
                    final_employees.append(mgr)
                    added_ids.add(mgr.id)

        elif role in [User.Role.MANAGER, User.Role.COMPANY_ADMIN]:
            # Manager/Admin View
            # Rule: "his team and his reporting manager"

            if current_emp.manager_id:
                # Restricted View
                added_ids = set()

                # 1. Reporting Manager
                if current_emp.manager_id in emp_map:
                    mgr = emp_map[current_emp.manager_id]
                    final_employees.append(mgr)
                    added_ids.add(mgr.id)

                # 2. Self
                if current_emp.id not in added_ids:
                    final_employees.append(current_emp)
                    added_ids.add(current_emp.id)

                # 3. Team (Subtree)
                team = get_subtree(request.user.id)
                for member in team:
                    if member.id not in added_ids:
                        final_employees.append(member)
                        added_ids.add(member.id)
            else:
                # Top Level / CEO / No Manager -> Show All
                final_employees = list(all_employees)
        else:
            # Unknown role -> Show All
            final_employees = list(all_employees)
    else:
        # No profile (e.g. SuperAdmin or Error) -> Show All
        final_employees = list(all_employees)

    # Helper to detect cycles
    def creates_cycle(parent_node, child_node):
        """Check if adding child_node to parent_node would create a cycle"""
        # BFS to see if parent is reachable from child (meaning child is already an ancestor of parent)
        queue = [child_node]
        visited = {child_node["id"]}
        while queue:
            curr = queue.pop(0)
            if curr["id"] == parent_node["id"]:
                return True
            for grandchild in curr["direct_reports"]:
                if grandchild["id"] not in visited:
                    visited.add(grandchild["id"])
                    queue.append(grandchild)
        return False

    try:
        # Use the filtered list for node building
        employees = final_employees

        # Build dictionary Key=USER_ID (not Employee ID)
        nodes = {}
        for emp in employees:
            if emp.user:
                nodes[emp.user.id] = {
                    "id": emp.user.id,
                    "employee": emp,
                    "user": emp.user,
                    "direct_reports": [],
                    "is_superadmin": False,
                }

        roots = []
        for emp in employees:
            if not emp.user:
                continue

            current_node = nodes[emp.user.id]
            manager_user = emp.manager  # User object

            if manager_user and manager_user.id != emp.user.id:
                # Case A: Manager is in the company (exists in nodes)
                if manager_user.id in nodes:
                    manager_node = nodes[manager_user.id]
                    # Check for cycle before adding
                    if not creates_cycle(manager_node, current_node):
                        if current_node not in manager_node["direct_reports"]:
                            manager_node["direct_reports"].append(current_node)
                    else:
                        # Cycle detected - do not link, treat as root to safely display
                        if current_node not in roots:
                            roots.append(current_node)

                else:
                    # Case B: Manager is External (SuperAdmin) or not in the filtered set
                    if manager_user.role == User.Role.SUPERADMIN:
                        # Create external manager node if needed
                        if manager_user.id not in nodes:
                            nodes[manager_user.id] = {
                                "id": manager_user.id,
                                "employee": None,
                                "user": manager_user,
                                "direct_reports": [],
                                "is_superadmin": True,
                            }
                            roots.append(nodes[manager_user.id])

                        manager_node = nodes[manager_user.id]
                        if not creates_cycle(manager_node, current_node):
                            if current_node not in manager_node["direct_reports"]:
                                manager_node["direct_reports"].append(current_node)
                    else:
                        # Manager not in current view, add as root to prevent orphan
                        if current_node not in roots:
                            roots.append(current_node)
            else:
                # No manager or self-managed (shouldn't happen)
                if current_node not in roots:
                    roots.append(current_node)

        # Filter roots to ensure no children - building from child_ids set is safer
        child_ids = set()
        for uid, node in nodes.items():
            for child in node["direct_reports"]:
                child_ids.add(child["id"])

        final_roots = [node for uid, node in nodes.items() if uid not in child_ids]

        return render(
            request,
            "core/org_chart.html",
            {
                "title": "Organisation Chart",
                "roots": final_roots,
                "company": request.user.company,
            },
        )
    except Exception as e:
        import traceback

        logger.error(f"Org Chart Error: {str(e)}\n{traceback.format_exc()}")
        # Fallback to simple list if tree fails
        return render(
            request,
            "core/org_chart.html",
            {
                "title": "Organisation Chart",
                "roots": [],
                "company": request.user.company,
                "error": "Could not generate hierarchical view. Please check reporting structures.",
            },
        )


@login_required
@manager_required
def attendance_analytics(request):
    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "Restricted access.")
        return redirect("dashboard")

    today = timezone.localtime().date()

    # Calculate date ranges
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    # Handle beginning of year for last month calculation
    if month_start.month == 1:
        last_month_start = date(month_start.year - 1, 12, 1)
        last_month_end = date(month_start.year - 1, 12, 31)
    else:
        last_month_start = (month_start - timedelta(days=1)).replace(day=1)
        last_month_end = month_start - timedelta(days=1)

    # Filter for Manager vs Admin
    if request.user.role == User.Role.MANAGER:
        # Get direct reports
        manager_profile = safe_get_employee_profile(request.user)
        if manager_profile:
            employees = Employee.objects.filter(manager=request.user)
        else:
            employees = Employee.objects.none()
    else:
        # Admin gets all company employees
        employees = Employee.objects.filter(company=request.user.company)

    # Filter out employees who left before the current month
    # Show active employees OR employees who exited this month (or later)
    employees = employees.filter(Q(is_active=True) | Q(exit_date__gte=month_start))

    total_employees = employees.count()
    employee_ids = employees.values_list("id", flat=True)

    # Today's stats filtered by accessible employees
    attendance_today = Attendance.objects.filter(employee__in=employee_ids, date=today)

    present_today = attendance_today.filter(status="PRESENT").count()
    absent_today = attendance_today.filter(status="ABSENT").count()
    leave_today = attendance_today.filter(status="LEAVE").count()
    wfh_today = attendance_today.filter(status="WFH").count()
    on_duty_today = attendance_today.filter(status="ON_DUTY").count()

    # Calculate late arrivals and early departures for today
    late_arrivals_today = 0
    early_departures_today = 0

    for att in attendance_today:
        if att.is_late:
            late_arrivals_today += 1
        if att.is_early_departure:
            early_departures_today += 1

    # Calculate percentages
    present_pct = (present_today / total_employees * 100) if total_employees > 0 else 0

    # This week's stats
    week_attendance = Attendance.objects.filter(employee__in=employee_ids, date__gte=week_start, date__lte=today)

    week_present = week_attendance.filter(status="PRESENT").count()
    week_absent = week_attendance.filter(status="ABSENT").count()
    week_wfh = week_attendance.filter(status="WFH").count()

    # This month's stats
    month_attendance = Attendance.objects.filter(employee__in=employee_ids, date__gte=month_start, date__lte=today)

    month_present = month_attendance.filter(status="PRESENT").count()
    month_absent = month_attendance.filter(status="ABSENT").count()
    month_wfh = month_attendance.filter(status="WFH").count()

    # Last month's stats for comparison
    last_month_attendance = Attendance.objects.filter(
        employee__in=employee_ids, date__gte=last_month_start, date__lte=last_month_end
    )

    last_month_present = last_month_attendance.filter(status="PRESENT").count()

    # Calculate month-over-month change
    if last_month_present > 0:
        mom_change = ((month_present - last_month_present) / last_month_present) * 100
    else:
        mom_change = 0

    # Department-wise breakdown
    departments = employees.values_list("department", flat=True).distinct()
    dept_stats = []

    for dept in departments:
        if dept:
            dept_emps = employees.filter(department=dept).values_list("id", flat=True)
            dept_att = attendance_today.filter(employee__in=dept_emps)
            dept_present = dept_att.filter(status="PRESENT").count()
            dept_total = len(dept_emps)
            dept_pct = (dept_present / dept_total * 100) if dept_total > 0 else 0

            dept_stats.append(
                {
                    "name": dept,
                    "present": dept_present,
                    "total": dept_total,
                    "percentage": round(dept_pct, 1),
                }
            )

    # Sort by percentage
    dept_stats.sort(key=lambda x: x["percentage"], reverse=True)

    return render(
        request,
        "core/attendance_analytics.html",
        {
            "title": "Attendance Analytics",
            "total_employees": total_employees,
            "present_today": present_today,
            "absent_today": absent_today,
            "leave_today": leave_today,
            "wfh_today": wfh_today,
            "on_duty_today": on_duty_today,
            "present_pct": round(present_pct, 1),
            "late_arrivals_today": late_arrivals_today,
            "early_departures_today": early_departures_today,
            # Weekly stats
            "week_present": week_present,
            "week_absent": week_absent,
            "week_wfh": week_wfh,
            # Monthly stats
            "month_present": month_present,
            "month_absent": month_absent,
            "month_wfh": month_wfh,
            "mom_change": round(mom_change, 1),
            # Department stats
            "dept_stats": dept_stats,
            "today": today,
        },
    )


@login_required
@manager_required
def attendance_report(request):
    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "Restricted access.")
        return redirect("dashboard")

    today = timezone.localtime().date()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    location_id = request.GET.get("location")

    from companies.models import Location

    # Calculate payroll cycle dates (28th to 27th)
    if month == 1:
        # January: Dec 28 to Jan 27
        start_date = date(year - 1, 12, 28)
        end_date = date(year, 1, 27)
    else:
        # Other months: Previous month 28th to current month 27th
        start_date = date(year, month - 1, 28)
        end_date = date(year, month, 27)

    # Only show data up to current date
    if end_date > today:
        end_date = today

    # Generate date range for payroll cycle (only up to current date)
    date_range = []
    current_date = start_date
    while current_date <= end_date:
        date_range.append(current_date)
        current_date += timedelta(days=1)

    # Filter Employees based on Role
    if request.user.role == User.Role.MANAGER:
        manager_profile = safe_get_employee_profile(request.user)
        if manager_profile:
            employees = Employee.objects.filter(manager=request.user).select_related("user", "manager", "location")
        else:
            employees = Employee.objects.none()
    else:
        employees = Employee.objects.filter(company=request.user.company).select_related("user", "manager", "location")

    if location_id:
        employees = employees.filter(location_id=location_id)

    # Filter out employees who left before the report period
    # Show active employees OR employees who exited on or after the start date
    employees = employees.filter(Q(is_active=True) | Q(exit_date__gte=start_date))

    locations = Location.objects.filter(company=request.user.company, is_active=True)
    employee_ids = employees.values_list("id", flat=True)

    # Get all attendance records for the period
    attendances = Attendance.objects.filter(employee__in=employee_ids, date__gte=start_date, date__lte=end_date)

    # Get all holidays for the period and company
    holidays = Holiday.objects.filter(
        company=request.user.company,
        date__gte=start_date,
        date__lte=end_date,
        is_active=True,
    ).select_related("location")

    # Create holiday map by location and date
    holiday_map = {}
    for holiday in holidays:
        if holiday.location_id not in holiday_map:
            holiday_map[holiday.location_id] = {}
        holiday_map[holiday.location_id][holiday.date] = holiday

    # Map attendance by employee and date
    att_map = {}
    for att in attendances:
        if att.employee_id not in att_map:
            att_map[att.employee_id] = {}
        att_map[att.employee_id][att.date] = att

    reports = []
    total_stats = {
        "present": 0,
        "absent": 0,
        "leave": 0,
        "half_day": 0,
        "weekly_off": 0,
        "holiday": 0,
    }

    for emp in employees:
        emp_data = {
            "employee": emp,
            "days": [],
            "stats": {
                "present": 0,
                "absent": 0,
                "leave": 0,
                "half_day": 0,
                "weekly_off": 0,
                "holiday": 0,
            },
        }

        for dt in date_range:
            att = att_map.get(emp.id, {}).get(dt)

            # Determine status based on actual attendance records and clock-in data
            if att:
                # Attendance record exists - check if employee actually clocked in
                if att.clock_in:
                    # Employee clocked in - determine status
                    if att.status == "WFH":
                        status_code = "WFH"  # Will be counted as Present
                    elif att.status == "PRESENT":
                        status_code = "PRESENT"
                    elif att.status == "HYBRID":
                        status_code = "PRESENT"  # Hybrid counted as present
                    elif att.status == "HALF_DAY":
                        status_code = "HALF_DAY"
                    else:
                        # Has clock_in but other status (shouldn't happen normally)
                        status_code = "PRESENT"
                else:
                    # Attendance record exists but no clock_in - check the status
                    if att.status == "LEAVE":
                        status_code = "LEAVE"
                    elif att.status == "WEEKLY_OFF":
                        status_code = "WEEKLY_OFF"
                    elif att.status == "HOLIDAY":
                        status_code = "HOLIDAY"
                    elif att.status == "HALF_DAY":
                        status_code = "HALF_DAY"
                    else:
                        # No clock_in and not leave/holiday/weekoff = absent
                        status_code = "ABSENT"
            else:
                # No attendance record - determine what it should be
                # Check if it's a holiday for this employee's location
                if emp.location_id and emp.location_id in holiday_map and dt in holiday_map[emp.location_id]:
                    status_code = "HOLIDAY"
                # Check if it's a weekoff for this employee
                elif emp.is_week_off(dt):
                    status_code = "WEEKLY_OFF"
                else:
                    # No record and not holiday/weekoff = absent
                    status_code = "ABSENT"

            # Map status to display value and count
            display_val = "-"

            if status_code == "PRESENT":
                display_val = "P"
                emp_data["stats"]["present"] += 1
                total_stats["present"] += 1
            elif status_code == "WFH":
                # Count WFH as present
                display_val = "P"
                emp_data["stats"]["present"] += 1
                total_stats["present"] += 1
            elif status_code == "ABSENT":
                display_val = "A"
                emp_data["stats"]["absent"] += 1
                total_stats["absent"] += 1
            elif status_code == "LEAVE":
                display_val = "L"
                emp_data["stats"]["leave"] += 1
                total_stats["leave"] += 1
            elif status_code == "HALF_DAY":
                display_val = "HD"
                emp_data["stats"]["half_day"] += 1
                total_stats["half_day"] += 1
            elif status_code == "WEEKLY_OFF":
                display_val = "WO"
                emp_data["stats"]["weekly_off"] += 1
                total_stats["weekly_off"] += 1
            elif status_code == "HOLIDAY":
                display_val = "H"
                emp_data["stats"]["holiday"] += 1
                total_stats["holiday"] += 1

            emp_data["days"].append(display_val)

        # Calculate working days and attendance percentage
        working_days = len(date_range) - emp_data["stats"]["weekly_off"] - emp_data["stats"]["holiday"]
        present_days = emp_data["stats"]["present"]  # WFH is already counted as present

        emp_data["working_days"] = working_days
        emp_data["present_days"] = present_days
        emp_data["attendance_percentage"] = round((present_days / working_days * 100) if working_days > 0 else 0, 1)

        reports.append(emp_data)

    # Create days display (show date with day number)
    days_display = []
    for dt in date_range:
        days_display.append({"day": dt.day, "month_short": dt.strftime("%b"), "date": dt})

    return render(
        request,
        "core/attendance_report.html",
        {
            "title": "Attendance Detailed Report",
            "reports": reports,
            "year": year,
            "month": month,
            "month_name": calendar.month_name[month],
            "days_display": days_display,
            "num_days": len(date_range),
            "start_date": start_date,
            "end_date": end_date,
            "cycle_label": f"{start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}",
            "locations": locations,
            "location_filter": int(location_id) if location_id else None,
            "total_stats": total_stats,
            "total_employees": len(reports),
        },
    )


@login_required
def download_attendance(request):
    if not hasattr(request.user, "company") or not request.user.company:
        return HttpResponse("Unauthorized", status=403)

    today = timezone.localtime().date()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    location_id = request.GET.get("location")

    # Calculate payroll cycle dates (28th to 27th)
    if month == 1:
        start_date = date(year - 1, 12, 28)
        end_date = date(year, 1, 27)
    else:
        start_date = date(year, month - 1, 28)
        end_date = date(year, month, 27)

    # Only show data up to current date
    if end_date > today:
        end_date = today

    # File name
    filename = f"Attendance_Payroll_Cycle_{start_date.strftime('%d%b')}_to_{end_date.strftime('%d%b%Y')}.xlsx"

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {calendar.month_name[month]} {year}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")

    # 1. Define Headers
    headers = [
        "Employee Number",
        "Employee Name",
        "Job Title",
        "Department",
        "Location",
        "Reporting Manager",
    ]

    # Dynamic Date Headers for payroll cycle (only up to current date)
    date_cols = []
    current_date = start_date
    while current_date <= end_date:
        col_name = current_date.strftime("%d-%b")
        headers.append(col_name)
        date_cols.append(current_date)
        current_date += timedelta(days=1)

    # Summary Headers (simplified)
    summary_headers = [
        "Total Days",
        "Present",
        "Half Day",
        "Weekly Offs",
        "Holidays",
        "Leave",
        "Absent Days",
        "Working Days",
        "Attendance %",
        "Late Arrival Days",
    ]
    headers.extend(summary_headers)

    # Write Headers
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 2. Fetch Data

    employees = Employee.objects.filter(company=request.user.company).select_related("user", "manager", "location")

    # Filter out employees who left before the report period
    employees = employees.filter(Q(is_active=True) | Q(exit_date__gte=start_date))

    if location_id:
        employees = employees.filter(location_id=location_id)

    attendances = Attendance.objects.filter(
        employee__company=request.user.company, date__gte=start_date, date__lte=end_date
    )

    # Get holidays for the period
    holidays = Holiday.objects.filter(
        company=request.user.company,
        date__gte=start_date,
        date__lte=end_date,
        is_active=True,
    ).select_related("location")

    # Create holiday map by location and date
    holiday_map = {}
    for holiday in holidays:
        if holiday.location_id not in holiday_map:
            holiday_map[holiday.location_id] = {}
        holiday_map[holiday.location_id][holiday.date] = holiday

    # Map attendance by employee and date
    att_map = {}
    for att in attendances:
        if att.employee_id not in att_map:
            att_map[att.employee_id] = {}
        att_map[att.employee_id][att.date] = att

    # 3. Write Rows
    row_num = 2

    for emp in employees:
        # Basic Info
        ws.cell(row=row_num, column=1, value=emp.badge_id)
        ws.cell(row=row_num, column=2, value=emp.user.get_full_name())
        ws.cell(row=row_num, column=3, value=emp.designation)
        ws.cell(row=row_num, column=4, value=emp.department)
        ws.cell(row=row_num, column=5, value=emp.location.name if emp.location else "N/A")
        ws.cell(
            row=row_num,
            column=6,
            value=emp.manager.get_full_name() if emp.manager else "-",
        )

        # Stats Counters (simplified)
        stats = {
            "present": 0,
            "absent": 0,
            "leave": 0,
            "half_day": 0,
            "weekly_off": 0,
            "holiday": 0,
            "late_arrival": 0,
        }

        # Date Columns
        col_idx = 7
        for dt in date_cols:
            att = att_map.get(emp.id, {}).get(dt)

            # Determine status using same logic as report view - check actual clock-in
            if att:
                # Attendance record exists - check if employee actually clocked in
                if att.clock_in:
                    # Employee clocked in - determine status
                    if att.status == "WFH":
                        status_code = "WFH"  # Will be counted as Present
                    elif att.status == "PRESENT":
                        status_code = "PRESENT"
                    elif att.status == "HYBRID":
                        status_code = "PRESENT"  # Hybrid counted as present
                    elif att.status == "HALF_DAY":
                        status_code = "HALF_DAY"
                    else:
                        # Has clock_in but other status (shouldn't happen normally)
                        status_code = "PRESENT"
                else:
                    # Attendance record exists but no clock_in - check the status
                    if att.status == "LEAVE":
                        status_code = "LEAVE"
                    elif att.status == "WEEKLY_OFF":
                        status_code = "WEEKLY_OFF"
                    elif att.status == "HOLIDAY":
                        status_code = "HOLIDAY"
                    elif att.status == "HALF_DAY":
                        status_code = "HALF_DAY"
                    else:
                        # No clock_in and not leave/holiday/weekoff = absent
                        status_code = "ABSENT"
            else:
                # No attendance record - determine what it should be
                if emp.location_id and emp.location_id in holiday_map and dt in holiday_map[emp.location_id]:
                    status_code = "HOLIDAY"
                elif emp.is_week_off(dt):
                    status_code = "WEEKLY_OFF"
                else:
                    status_code = "ABSENT"

            # Map status to display value and count
            display_val = "-"
            if status_code == "PRESENT":
                display_val = "P"
                stats["present"] += 1
                if att and att.is_late:
                    display_val += " (L)"
                    stats["late_arrival"] += 1
            elif status_code == "WFH":
                # Count WFH as present
                display_val = "P"
                stats["present"] += 1
            elif status_code == "ABSENT":
                display_val = "A"
                stats["absent"] += 1
            elif status_code == "LEAVE":
                display_val = "L"
                stats["leave"] += 1
            elif status_code == "HALF_DAY":
                display_val = "HD"
                stats["half_day"] += 1
            elif status_code == "WEEKLY_OFF":
                display_val = "WO"
                stats["weekly_off"] += 1
            elif status_code == "HOLIDAY":
                display_val = "H"
                stats["holiday"] += 1

            cell = ws.cell(row=row_num, column=col_idx, value=display_val)
            cell.alignment = Alignment(horizontal="center")
            col_idx += 1

        # Summary Columns
        total_days = len(date_cols)
        working_days = total_days - stats["weekly_off"] - stats["holiday"]
        present_days = stats["present"]  # WFH is already counted as present
        attendance_percentage = round((present_days / working_days * 100) if working_days > 0 else 0, 1)

        ws.cell(row=row_num, column=col_idx, value=total_days)
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=stats["present"])
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=stats["half_day"])
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=stats["weekly_off"])
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=stats["holiday"])
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=stats["leave"])
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=stats["absent"])
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=working_days)
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=f"{attendance_percentage}%")
        col_idx += 1
        ws.cell(row=row_num, column=col_idx, value=stats["late_arrival"])
        col_idx += 1

        row_num += 1

    # Return Excel File
    import io

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# --- Leaves Section ---


@login_required
def leave_requests(request):
    """Admin view for managing leave requests"""

    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "Restricted access.")

        return redirect("dashboard")

    # Handle approve/reject actions

    if request.method == "POST":
        action = request.POST.get("action")

        leave_id = request.POST.get("leave_id")

        admin_comment = request.POST.get("admin_comment", "")

        try:
            leave_request = LeaveRequest.objects.get(id=leave_id, employee__company=request.user.company)

            if action == "approve":
                prev_status = leave_request.status

                # Only transition and deduct if this wasn't already approved
                if prev_status != "APPROVED":
                    # Use the model's approve_leave method to handle all balance updates properly
                    if leave_request.approve_leave(request.user, approval_type="FULL"):
                        leave_request.admin_comment = admin_comment
                        leave_request.save()

                    messages.success(
                        request,
                        f"Leave request approved for {leave_request.employee.user.get_full_name()}",
                    )
                else:
                    messages.info(request, "Leave request was already approved earlier.")

                # Send Approval Email
                try:
                    from core.email_utils import send_leave_approval_notification

                    if not send_leave_approval_notification(leave_request):
                        messages.warning(request, "Leave approved, but email notification failed.")
                except Exception as e:
                    import logging

                    logger = logging.getLogger(__name__)
                    logger.error(f"Error sending approval email: {e}")

            elif action == "reject":
                if not admin_comment:
                    messages.error(request, "Admin comment is mandatory for rejection.")

                    return redirect("leave_requests")

                leave_request.status = "REJECTED"

                leave_request.approved_by = request.user

                leave_request.approved_at = timezone.now()

                leave_request.rejection_reason = admin_comment

                leave_request.admin_comment = admin_comment

                leave_request.save()

                messages.success(
                    request,
                    f"Leave request rejected for {leave_request.employee.user.get_full_name()}",
                )

                # Send Rejection Email
                try:
                    from core.email_utils import send_leave_rejection_notification

                    if not send_leave_rejection_notification(leave_request):
                        messages.warning(request, "Leave rejected, but email notification failed.")
                except Exception as e:
                    import logging

                    logger = logging.getLogger(__name__)
                    logger.error(f"Error sending rejection email: {e}")

        except LeaveRequest.DoesNotExist:
            messages.error(request, "Leave request not found.")

        except Exception as e:
            messages.error(request, f"Error processing request: {e}")

        return redirect("leave_requests")

    # Filter parameters

    status_filter = request.GET.get("status", "PENDING")

    employee_filter = request.GET.get("employee", "")

    department_filter = request.GET.get("department", "")

    leave_type_filter = request.GET.get("leave_type", "")

    # Base query

    leave_requests = LeaveRequest.objects.filter(employee__company=request.user.company).select_related(
        "employee__user", "employee__manager", "approved_by"
    )

    # Apply filters

    if status_filter:
        leave_requests = leave_requests.filter(status=status_filter)

    if employee_filter:
        leave_requests = leave_requests.filter(
            Q(employee__user__first_name__icontains=employee_filter)
            | Q(employee__user__last_name__icontains=employee_filter)
            | Q(employee__badge_id__icontains=employee_filter)
        )

    if department_filter:
        leave_requests = leave_requests.filter(employee__department__icontains=department_filter)

    if leave_type_filter:
        leave_requests = leave_requests.filter(leave_type=leave_type_filter)

    # Get unique departments for filter dropdown

    departments = Employee.objects.filter(company=request.user.company).values_list("department", flat=True).distinct()

    return render(
        request,
        "core/leave_requests.html",
        {
            "title": "Leave Requests",
            "leave_requests": leave_requests,
            "departments": departments,
            "status_filter": status_filter,
            "employee_filter": employee_filter,
            "department_filter": department_filter,
            "leave_type_filter": leave_type_filter,
        },
    )


@login_required
@manager_required
def leave_history(request):
    """View for leave history with filtering and export"""

    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "Restricted access.")

        return redirect("dashboard")

    # Filter parameters

    employee_filter = request.GET.get("employee", "")

    department_filter = request.GET.get("department", "")

    leave_type_filter = request.GET.get("leave_type", "")

    status_filter = request.GET.get("status", "")

    year_filter = request.GET.get("year", "")

    month_filter = request.GET.get("month", "")

    # Base query based on Role

    if request.user.role == User.Role.MANAGER:
        manager_profile = safe_get_employee_profile(request.user)
        if manager_profile:
            employees = Employee.objects.filter(manager=request.user)
        else:
            employees = Employee.objects.none()

        leave_history = LeaveRequest.objects.filter(employee__in=employees).select_related(
            "employee__user",
            "employee__manager",
            "employee__leave_balance",
            "approved_by",
        )

    else:
        # Admin gets all

        leave_history = LeaveRequest.objects.filter(employee__company=request.user.company).select_related(
            "employee__user",
            "employee__manager",
            "employee__leave_balance",
            "approved_by",
        )

    # Apply filters

    if employee_filter:
        leave_history = leave_history.filter(
            Q(employee__user__first_name__icontains=employee_filter)
            | Q(employee__user__last_name__icontains=employee_filter)
            | Q(employee__badge_id__icontains=employee_filter)
        )

    if department_filter:
        leave_history = leave_history.filter(employee__department__icontains=department_filter)

    if leave_type_filter:
        leave_history = leave_history.filter(leave_type=leave_type_filter)

    if status_filter:
        leave_history = leave_history.filter(status=status_filter)

    if year_filter:
        leave_history = leave_history.filter(start_date__year=int(year_filter))

    if month_filter:
        leave_history = leave_history.filter(start_date__month=int(month_filter))

    # Get unique departments and years for filters

    departments = Employee.objects.filter(company=request.user.company).values_list("department", flat=True).distinct()

    years = LeaveRequest.objects.filter(employee__company=request.user.company).dates("start_date", "year").distinct()

    return render(
        request,
        "core/leave_history.html",
        {
            "title": "Leave History",
            "leave_history": leave_history,
            "departments": departments,
            "years": [y.year for y in years],
            "employee_filter": employee_filter,
            "department_filter": department_filter,
            "leave_type_filter": leave_type_filter,
            "status_filter": status_filter,
            "year_filter": year_filter,
            "month_filter": month_filter,
        },
    )


# --- Payroll Section Stubs ---


@login_required
@admin_required
def payroll_dashboard(request):
    """Admin Payroll Dashboard - Manage employee payslips"""
    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "Restricted access.")
        return redirect("dashboard")

    company = request.user.company
    today = timezone.localtime().date()
    
    # Month/Year selection
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))
    
    # Get all active employees
    employees = Employee.objects.filter(company=company, is_active=True).select_related("user")
    
    # Get payslips for the selected month/year
    # month_date is used for filtering. We use the first of the month.
    month_filter = date(selected_year, selected_month, 1)
    
    existing_payslips = Payslip.objects.filter(
        employee__company=company,
        month__month=selected_month,
        month__year=selected_year
    )
    
    # Create a map for easy lookup
    payslip_map = {slip.employee_id: slip for slip in existing_payslips}
    
    # Enhance employee list with payslip info
    employee_data = []
    for emp in employees:
        employee_data.append({
            "employee": emp,
            "payslip": payslip_map.get(emp.id)
        })
    
    months = []
    for i in range(1, 13):
        months.append({"value": i, "name": calendar.month_name[i]})
        
    years = range(today.year - 2, today.year + 2)

    # Add days in month to context for default worked days
    days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    
    context = {
        "title": "Payroll Dashboard",
        "employee_data": employee_data,
        "selected_month": selected_month,
        "selected_year": selected_year,
        "months": months,
        "years": years,
        "days_in_month": days_in_month,
    }
    
    return render(request, "core/payroll_dashboard.html", context)

@login_required
@admin_required
def upload_payslip(request):
    """API or View to upload/generate a payslip for an employee"""
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        month = int(request.POST.get("month"))
        year = int(request.POST.get("year"))
        net_salary = request.POST.get("net_salary") or 0
        annual_ctc = request.POST.get("annual_ctc")
        uan = request.POST.get("uan")
        pdf_file = request.FILES.get("pdf_file")
        
        try:
            employee = Employee.objects.get(id=employee_id, company=request.user.company)
            
            # Reflect changes to employee finance profile
            if annual_ctc:
                employee.annual_ctc = float(annual_ctc)
            if uan is not None:
                employee.uan = uan
            employee.save()
            
            month_date = date(year, month, 1)
            
            payslip, created = Payslip.objects.get_or_create(
                employee=employee,
                month=month_date,
                defaults={"net_salary": net_salary}
            )
            
            if not created:
                payslip.net_salary = net_salary
            
            if pdf_file:
                payslip.pdf_file = pdf_file
            
            payslip.save()
            
            messages.success(request, f"Payslip for {employee.user.get_full_name()} saved successfully.")
        except Exception as e:
            messages.error(request, f"Error saving payslip: {str(e)}")
            
    return redirect(f"{reverse('payroll_dashboard')}?month={request.POST.get('month')}&year={request.POST.get('year')}")

@login_required
@admin_required
def calculate_generated_payslip(request):
    """API to calculate payslip breakdown without saving"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            employee_id = data.get("employee_id")
            annual_ctc = data.get("annual_ctc")
            worked_days = data.get("worked_days")
            month = int(data.get("month"))
            year = int(data.get("year"))
            
            # Prioritize the pf_enabled flag from request if provided
            pf_enabled = data.get("pf_enabled")
            if pf_enabled is None:
                if employee_id:
                    employee = Employee.objects.get(id=employee_id, company=request.user.company)
                    pf_enabled = employee.pf_enabled
                else:
                    pf_enabled = True
            
            # Days in month
            total_days = calendar.monthrange(year, month)[1]
            
            # Get employee's location if available
            location = None
            if employee_id:
                try:
                    employee = Employee.objects.get(id=employee_id, company=request.user.company)
                    location = employee.location
                except Employee.DoesNotExist:
                    pass

            breakdown = calculate_payslip_breakdown(annual_ctc, worked_days, total_days, pf_enabled, location=location)
            return JsonResponse({"status": "success", "breakdown": breakdown})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)

@login_required
@admin_required
def process_payslip_generation(request):
    """View to calculate, save and generate PDF payslip"""
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        annual_ctc = request.POST.get("annual_ctc")
        worked_days = float(request.POST.get("worked_days") or 0)
        month = int(request.POST.get("month"))
        year = int(request.POST.get("year"))
        
        try:
            employee = Employee.objects.get(id=employee_id, company=request.user.company)
            
            # Update employee's annual CTC if changed in the calculator
            if annual_ctc:
                new_ctc = float(str(annual_ctc).replace(",", ""))
                if float(employee.annual_ctc or 0) != new_ctc:
                    employee.annual_ctc = new_ctc
                    employee.save()
            
            total_days = calendar.monthrange(year, month)[1]
            month_date = date(year, month, 1)
            
            # Pass employee's PF status and location
            breakdown = calculate_payslip_breakdown(annual_ctc, worked_days, total_days, employee.pf_enabled, location=employee.location)
            
            payslip, created = Payslip.objects.get_or_create(
                employee=employee,
                month=month_date
            )
            
            # Update fields
            payslip.basic = breakdown['basic']
            payslip.hra = breakdown['hra']
            payslip.lta = breakdown['lta']
            payslip.other_allowance = breakdown['other_allowance']
            # Map location specific allowances
            payslip.conveyance_allowance = breakdown.get('conveyance', 0.0)
            payslip.special_allowance = breakdown.get('medical', 0.0)
            payslip.monthly_gross = breakdown['full_monthly_gross']
            payslip.gross_salary = breakdown['gross_monthly']
            payslip.employee_pf = breakdown['employee_pf']
            payslip.employer_pf = breakdown['employer_pf']
            payslip.professional_tax = breakdown['professional_tax']
            payslip.net_salary = breakdown['net_salary']
            payslip.worked_days = worked_days
            payslip.total_days = total_days
            payslip.save()
            
            # Generate PDF
            # Determine currency name for words
            currency_name = "Rupees"
            if employee.location:
                if employee.location.country_code == 'BD' or employee.location.currency == 'BDT':
                    currency_name = "Taka"
                elif employee.location.country_code == 'US' or employee.location.currency == 'USD':
                    currency_name = "Dollars"
                elif employee.location.currency:
                    currency_name = employee.location.currency

            # Prepare branding info
            cname_upper = employee.company.name.upper()
            branding = {
                'name': 'PETABYTZ TECHNOLOGY SERVICES PVT LTD',
                'address': 'PLOT NO 201 & 202, 1ST FLOOR, DMR CORPORATE, KAVURI HILLS RD, HYDERABAD, TELANGANA 500081.'
            }
            if "SOFTSTANDARD" in cname_upper or "RMINDS" in cname_upper:
                branding['name'] = 'SOFTSTANDARD SOLUTIONS'
            
            # Use location address if available
            if employee.location and employee.location.address_line1:
                loc = employee.location
                addr = f"{loc.address_line1}"
                if loc.address_line2:
                    addr += f", {loc.address_line2}"
                addr += f", {loc.city}"
                if loc.state:
                    addr += f", {loc.state}"
                if loc.postal_code:
                    addr += f" {loc.postal_code}"
                branding['address'] = addr

            context = {
                'payslip': payslip,
                'company': employee.company,
                'branding': branding,
                'net_salary_words': num2words_flexible(payslip.net_salary, currency_name)
            }
            filename = f"payslip_{employee.badge_id}_{month_date.strftime('%b_%Y')}.pdf"
            save_pdf_to_model(payslip, 'employees/payslip_pdf.html', context, filename)
            
            messages.success(request, f"Payslip for {employee.user.get_full_name()} generated successfully.")
        except Exception as e:
            messages.error(request, f"Error generating payslip: {str(e)}")
            
    return redirect(f"{reverse('payroll_dashboard')}?month={request.POST.get('month')}&year={request.POST.get('year')}")


@login_required
@admin_required
def download_payslip_template(request):
    """Download Excel template for bulk payslip upload"""
    import openpyxl
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payslip Template"
    
    headers = ["Employee ID", "Name", "Department", "Designation", "Monthly Gross", "Net Pay", "Payable Days"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        
    # Add some sample data from active employees
    employees = Employee.objects.filter(company=request.user.company, is_active=True).select_related('user')[:5]
    for row_num, emp in enumerate(employees, 2):
        ws.cell(row=row_num, column=1, value=emp.badge_id)
        ws.cell(row=row_num, column=2, value=emp.user.get_full_name())
        ws.cell(row=row_num, column=3, value=emp.department or "N/A")
        ws.cell(row=row_num, column=4, value=emp.designation or "N/A")
        ws.cell(row=row_num, column=5, value=float(emp.annual_ctc or 0) / 12)
        # Net pay is not easily pre-calculated without knowing worked days
        ws.cell(row=row_num, column=7, value=30) 
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Bulk_Payslip_Template.xlsx"'
    wb.save(response)
    return response


@login_required
@admin_required
def bulk_upload_payslips(request):
    """Refined bulk upload from Excel"""
    if request.method == "POST" and request.FILES.get("excel_file"):
        import openpyxl
        from datetime import date
        
        excel_file = request.FILES["excel_file"]
        month = int(request.POST.get("month"))
        year = int(request.POST.get("year"))
        month_date = date(year, month, 1)
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Map headers to column indices
            header_map = {}
            for cell in ws[1]:
                if cell.value:
                    header_map[cell.value.strip().lower()] = cell.column - 1
            
            # Required columns
            needed = ['employee id', 'monthly gross', 'payable days']
            for col in needed:
                if col not in header_map:
                    messages.error(request, f"Required column missing: {col}")
                    return redirect(f"{reverse('payroll_dashboard')}?month={month}&year={year}")
            
            success_count = 0
            error_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                badge_id = str(row[header_map['employee id']]) if row[header_map['employee id']] else None
                monthly_gross = row[header_map['monthly gross']]
                payable_days = row[header_map['payable days']]
                
                if not badge_id or monthly_gross is None or payable_days is None:
                    continue
                
                try:
                    employee = Employee.objects.get(badge_id=badge_id, company=request.user.company)
                    annual_ctc = float(monthly_gross) * 12
                    worked_days = float(payable_days)
                    total_days = calendar.monthrange(year, month)[1]
                    
                    # Calculate breakdown
                    breakdown = calculate_payslip_breakdown(
                        annual_ctc, worked_days, total_days, 
                        employee.pf_enabled, location=employee.location
                    )
                    
                    payslip, created = Payslip.objects.get_or_create(
                        employee=employee,
                        month=month_date
                    )
                    
                    # Update fields
                    payslip.basic = breakdown['basic']
                    payslip.hra = breakdown['hra']
                    payslip.lta = breakdown['lta']
                    payslip.other_allowance = breakdown['other_allowance']
                    payslip.conveyance_allowance = breakdown.get('conveyance', 0.0)
                    payslip.special_allowance = breakdown.get('medical', 0.0)
                    payslip.monthly_gross = breakdown['full_monthly_gross']
                    payslip.gross_salary = breakdown['gross_monthly']
                    payslip.employee_pf = breakdown['employee_pf']
                    payslip.employer_pf = breakdown['employer_pf']
                    payslip.professional_tax = breakdown['professional_tax']
                    payslip.net_salary = breakdown['net_salary']
                    payslip.worked_days = worked_days
                    payslip.total_days = total_days
                    payslip.save()
                    
                    # Generate PDF
                    currency_name = "Rupees"
                    if employee.location:
                        if employee.location.country_code == 'BD' or employee.location.currency == 'BDT':
                            currency_name = "Taka"
                        elif employee.location.country_code == 'US' or employee.location.currency == 'USD':
                            currency_name = "Dollars"
                        elif employee.location.currency:
                            currency_name = employee.location.currency

                    cname_upper = employee.company.name.upper()
                    branding = {
                        'name': 'PETABYTZ TECHNOLOGY SERVICES PVT LTD',
                        'address': 'PLOT NO 201 & 202, 1ST FLOOR, DMR CORPORATE, KAVURI HILLS RD, HYDERABAD, TELANGANA 500081.'
                    }
                    if "SOFTSTANDARD" in cname_upper or "RMINDS" in cname_upper:
                        branding['name'] = 'SOFTSTANDARD SOLUTIONS'
                    
                    if employee.location and employee.location.address_line1:
                        loc = employee.location
                        addr = f"{loc.address_line1}"
                        if loc.address_line2: addr += f", {loc.address_line2}"
                        addr += f", {loc.city}"
                        if loc.state: addr += f", {loc.state}"
                        if loc.postal_code: addr += f" {loc.postal_code}"
                        branding['address'] = addr

                    context = {
                        'payslip': payslip,
                        'company': employee.company,
                        'branding': branding,
                        'net_salary_words': num2words_flexible(payslip.net_salary, currency_name)
                    }
                    filename = f"payslip_{employee.badge_id}_{month_date.strftime('%b_%Y')}.pdf"
                    save_pdf_to_model(payslip, 'employees/payslip_pdf.html', context, filename)
                    
                    success_count += 1
                except Employee.DoesNotExist:
                    error_count += 1
                except Exception as e:
                    error_count += 1
                    
            messages.success(request, f"Successfully processed {success_count} payslips. {error_count} errors.")
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            
    return redirect(f"{reverse('payroll_dashboard')}?month={request.POST.get('month', date.today().month)}&year={request.POST.get('year', date.today().year)}")




# --- Configuration Section ---


@login_required
@admin_required
def holidays(request):
    """Holiday Configuration - Manage company holidays with location support"""

    if not hasattr(request.user, "company") or not request.user.company:
        messages.error(request, "Restricted access.")

        return redirect("dashboard")

    from datetime import datetime

    from companies.models import Location

    # Handle POST requests (Add/Edit/Delete/Import)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            try:
                Holiday.objects.create(
                    company=request.user.company,
                    name=request.POST.get("name"),
                    date=datetime.strptime(request.POST.get("date"), "%Y-%m-%d").date(),
                    location=Location.objects.get(id=request.POST.get("location")),
                    holiday_type=request.POST.get("holiday_type", "MANDATORY"),
                    description=request.POST.get("description", ""),
                    created_by=request.user.get_full_name(),
                )

                messages.success(request, f"Holiday '{request.POST.get('name')}' added successfully!")

            except Exception as e:
                messages.error(request, f"Error adding holiday: {e}")

        elif action == "edit":
            try:
                holiday_id = request.POST.get("holiday_id")

                holiday = Holiday.objects.get(id=holiday_id, company=request.user.company)

                holiday.name = request.POST.get("name")

                holiday.date = datetime.strptime(request.POST.get("date"), "%Y-%m-%d").date()

                holiday.location = Location.objects.get(id=request.POST.get("location"))

                holiday.holiday_type = request.POST.get("holiday_type")

                holiday.description = request.POST.get("description", "")

                holiday.save()

                messages.success(request, f"Holiday '{holiday.name}' updated successfully!")

            except Holiday.DoesNotExist:
                messages.error(request, "Holiday not found.")

            except Exception as e:
                messages.error(request, f"Error updating holiday: {e}")

        elif action == "delete":
            try:
                holiday_id = request.POST.get("holiday_id")

                holiday = Holiday.objects.get(id=holiday_id, company=request.user.company)

                holiday_name = holiday.name

                holiday.delete()

                messages.success(request, f"Holiday '{holiday_name}' deleted successfully!")

            except Holiday.DoesNotExist:
                messages.error(request, "Holiday not found.")

            except Exception as e:
                messages.error(request, f"Error deleting holiday: {e}")

        elif action == "import_excel":
            # Handle Excel import

            if "excel_file" in request.FILES:
                try:
                    from datetime import datetime

                    import openpyxl

                    file = request.FILES["excel_file"]

                    wb = openpyxl.load_workbook(file)

                    ws = wb.active

                    imported_count = 0

                    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                        if row[0]:  # Check if name exists
                            try:
                                # Expected columns: Name, Date, Location, Type, Description

                                holiday_date = (
                                    row[1]
                                    if isinstance(row[1], date)
                                    else datetime.strptime(str(row[1]), "%Y-%m-%d").date()
                                )

                                Holiday.objects.create(
                                    company=request.user.company,
                                    name=row[0],
                                    date=holiday_date,
                                    location=Location.objects.get(name__iexact=row[2])
                                    if row[2]
                                    else Location.objects.get(name="Global"),
                                    holiday_type=row[3] if row[3] else "MANDATORY",
                                    description=row[4] if len(row) > 4 and row[4] else "",
                                    created_by=request.user.get_full_name(),
                                )

                                imported_count += 1

                            except Exception:
                                continue

                    messages.success(request, f"Successfully imported {imported_count} holidays!")

                except Exception as e:
                    messages.error(request, f"Error importing Excel file: {e}")

            else:
                messages.error(request, "No file uploaded.")

        return redirect("holidays")

    # Filter parameters

    year_filter = request.GET.get("year", timezone.now().year)

    location_filter = request.GET.get("location", "")

    type_filter = request.GET.get("type", "")

    # Base query

    holidays_list = Holiday.objects.filter(company=request.user.company)

    # Apply filters

    if year_filter:
        holidays_list = holidays_list.filter(year=int(year_filter))

    if location_filter:
        holidays_list = holidays_list.filter(location_id=location_filter)

    if type_filter:
        holidays_list = holidays_list.filter(holiday_type=type_filter)

    # Get unique years for filter

    years = (
        Holiday.objects.filter(company=request.user.company).values_list("year", flat=True).distinct().order_by("-year")
    )

    # Statistics

    total_holidays = holidays_list.count()

    mandatory_count = holidays_list.filter(holiday_type="MANDATORY").count()

    optional_count = holidays_list.filter(holiday_type="OPTIONAL").count()

    locations = Location.objects.filter(company=request.user.company, is_active=True)
    location_groups = {}
    for loc in locations:
        location_holidays = holidays_list.filter(location=loc)
        if location_holidays.exists():
            location_groups[loc.name] = location_holidays

    return render(
        request,
        "core/holidays.html",
        {
            "title": "Holiday Configuration",
            "holidays": holidays_list,
            "location_groups": location_groups,
            "years": years,
            "year_filter": int(year_filter) if year_filter else timezone.now().year,
            "location_filter": location_filter,
            "type_filter": type_filter,
            "total_holidays": total_holidays,
            "mandatory_count": mandatory_count,
            "optional_count": optional_count,
            "locations": locations,
            "type_choices": Holiday.HOLIDAY_TYPE_CHOICES,
        },
    )


@login_required
@admin_required
def download_holiday_template(request):
    """Download Excel template for holiday import"""

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Holiday Template"

    # Headers

    headers = ["Name", "Date (YYYY-MM-DD)", "Location", "Type", "Description"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(horizontal="center")

    # Sample data
    ws.append(["Republic Day", "2025-01-26", "India", "MANDATORY", "National Holiday"])
    ws.append(["Independence Day", "2025-07-04", "US", "MANDATORY", "National Holiday"])
    ws.append(["Diwali", "2025-10-20", "India", "MANDATORY", "Festival of Lights"])

    # Adjust column widths

    ws.column_dimensions["A"].width = 25

    ws.column_dimensions["B"].width = 20

    ws.column_dimensions["C"].width = 15

    ws.column_dimensions["D"].width = 15

    ws.column_dimensions["E"].width = 40

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    response["Content-Disposition"] = 'attachment; filename="Holiday_Import_Template.xlsx"'

    wb.save(response)

    return response


@login_required
@admin_required
def export_holidays(request):
    """Export holidays to Excel"""

    if not hasattr(request.user, "company") or not request.user.company:
        return HttpResponse("Unauthorized", status=403)

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    year = request.GET.get("year", timezone.now().year)

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = f"Holidays {year}"

    # Headers

    headers = ["Holiday Name", "Date", "Day", "Location", "Type", "Description"]

    header_fill = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(horizontal="center")

    # Data

    holidays = Holiday.objects.filter(company=request.user.company, year=year).order_by("date")

    for row_num, holiday in enumerate(holidays, 2):
        ws.cell(row=row_num, column=1, value=holiday.name)

        ws.cell(row=row_num, column=2, value=holiday.date.strftime("%d-%b-%Y"))

        ws.cell(row=row_num, column=3, value=holiday.date.strftime("%A"))

        ws.cell(row=row_num, column=4, value=holiday.location.name)

        ws.cell(row=row_num, column=5, value=holiday.get_holiday_type_display())

        ws.cell(row=row_num, column=6, value=holiday.description or "")

    # Adjust column widths

    ws.column_dimensions["A"].width = 30

    ws.column_dimensions["B"].width = 15

    ws.column_dimensions["C"].width = 12

    ws.column_dimensions["D"].width = 20

    ws.column_dimensions["E"].width = 20

    ws.column_dimensions["F"].width = 40

    filename = f"Holidays_{request.user.company.name}_{year}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response


@login_required
@admin_required
def company_leaves(request):
    return render(request, "core/stub.html", {"title": "Company Leave Configuration"})


# -------------------------------------------------------------------------
# Forgot Password / OTP Flow
# -------------------------------------------------------------------------


def forgot_password_view(request):
    if request.method == "POST":
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            user = User.objects.get(email=email)

            # Generate 6-digit OTP
            otp = str(random.randint(100000, 999999))

            # Save OTP
            # Save OTP
            PasswordResetOTP.objects.create(user=user, otp=otp)

            # ALWAYS PRINT OTP FOR DEBUGGING/DEV
            # print("==========================================")
            # print(f"Generated OTP for {email}: {otp}")
            # print("==========================================")

            # Send Email
            try:
                from django.core.mail import EmailMultiAlternatives
                from django.template.loader import render_to_string
                from django.utils.html import strip_tags

                subject = "Password Reset OTP - Petabytz HRMS"
                context = {"otp": otp}
                html_content = render_to_string("accounts/emails/password_reset_otp_email.html", context)
                text_content = strip_tags(html_content)

                # Use EMAIL_HOST_USER as from_email for Microsoft 365 compatibility
                from_email = settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL

                email_msg = EmailMultiAlternatives(subject, text_content, from_email, [email])
                email_msg.attach_alternative(html_content, "text/html")
                email_msg.send(fail_silently=False)

                messages.success(request, f"OTP sent to {email}. Please check your inbox.")
            except Exception as e:
                import traceback

                logger.error(f"Error sending email: {e}")
                logger.error(traceback.format_exc())
                messages.error(request, f"Failed to send OTP email. Error: {str(e)}")

            request.session["reset_email"] = email
            return redirect("verify_otp")
    else:
        form = ForgotPasswordForm()

    return render(request, "core/forgot_password.html", {"form": form})


def verify_otp_view(request):
    email = request.session.get("reset_email")
    if not email:
        messages.error(request, "Session expired. Please try again.")
        return redirect("forgot_password")

    if request.method == "POST":
        form = OTPVerificationForm(request.POST)
        if form.is_valid():
            otp_input = form.cleaned_data["otp"]
            user = User.objects.get(email=email)

            # Verify OTP (check latest for user)
            otp_record = PasswordResetOTP.objects.filter(user=user).last()

            if otp_record and otp_record.otp == otp_input:
                otp_record.is_verified = True
                otp_record.save()

                request.session["otp_verified"] = True
                messages.success(request, "OTP Verified Successfully.")
                return redirect("reset_password")
            else:
                messages.error(request, "Invalid OTP.")
    else:
        form = OTPVerificationForm()

    return render(request, "core/verify_otp.html", {"form": form, "email": email})


def reset_password_view(request):
    email = request.session.get("reset_email")
    is_verified = request.session.get("otp_verified")

    if not email or not is_verified:
        messages.error(request, "Unauthorized access. Please verify OTP first.")
        return redirect("forgot_password")

    if request.method == "POST":
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data["new_password"]
            user = User.objects.get(email=email)

            user.set_password(new_password)
            user.save()

            # Clear session
            del request.session["reset_email"]
            del request.session["otp_verified"]

            messages.success(request, "Password reset successfully. You can now login.")
            return redirect("login")  # Assuming 'login' is the name of your login url
    else:
        form = ResetPasswordForm()

    return render(request, "core/reset_password.html", {"form": form})


# ==================== Notification Views ====================


@login_required
def get_notifications(request):
    """
    API endpoint to get notifications for the current user
    """
    from django.http import JsonResponse

    from accounts.models import User

    from .models import Notification

    # Only allow managers, admins, and HR department users to see notifications
    emp = getattr(request.user, "employee_profile", None)
    is_hr = emp and str(getattr(emp, "department", "")).upper() == "HR"

    if request.user.role not in [User.Role.COMPANY_ADMIN, User.Role.MANAGER] and not is_hr:
        return JsonResponse({"notifications": [], "count": 0})

    notifications = Notification.objects.filter(recipient=request.user).order_by("-created_at")[:20]

    notifications_data = []
    for notif in notifications:
        notifications_data.append(
            {
                "id": notif.id,
                "type": notif.notification_type,
                "message": notif.message,
                "is_read": notif.is_read,
                "created_at": notif.created_at.isoformat(),
                "url": get_notification_url(notif),
            }
        )

    unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()

    return JsonResponse({"notifications": notifications_data, "count": unread_count})


@login_required
def mark_notification_read(request, notification_id):
    """
    Mark a notification as read
    """
    from django.http import JsonResponse
    from django.utils import timezone

    from .models import Notification

    try:
        notification = Notification.objects.get(id=notification_id, recipient=request.user)
        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save()

        unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()

        return JsonResponse({"success": True, "unread_count": unread_count})
    except Notification.DoesNotExist:
        return JsonResponse({"success": False, "error": "Notification not found"}, status=404)


@login_required
def mark_all_notifications_read(request):
    """
    Mark all notifications as read for the current user
    """
    from django.http import JsonResponse
    from django.utils import timezone

    from .models import Notification

    if request.method == "POST":
        Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True, read_at=timezone.now())

        return JsonResponse({"success": True, "unread_count": 0})

    return JsonResponse({"success": False, "error": "Invalid request method"}, status=400)


def get_notification_url(notification):
    """
    Helper function to get the URL for a notification based on its type
    """
    from django.urls import reverse

    if notification.notification_type == "LEAVE_REQUEST":
        return reverse("leave_requests")
    elif notification.notification_type == "REGULARIZATION_REQUEST":
        return reverse("regularization_list")

    return "#"
